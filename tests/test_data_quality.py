from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text

from app.main import app
from app.modules.data_quality import service
from app.modules.data_quality.service import get_product_quality_report


client = TestClient(app)


def create_product_table(conn):
    conn.execute(
        text(
            """
            CREATE TABLE amazon_product_info (
                id INTEGER PRIMARY KEY,
                store_site TEXT,
                msku TEXT,
                asin TEXT,
                listing TEXT,
                brand TEXT,
                sales_status TEXT,
                product_name TEXT,
                updated_at TEXT
            )
            """
        )
    )


def create_listing_owner_table(conn):
    conn.execute(
        text(
            """
            CREATE TABLE amazon_listing_owner_config (
                id INTEGER PRIMARY KEY,
                store_site TEXT,
                listing TEXT,
                owner TEXT
            )
            """
        )
    )


def create_store_site_table(conn):
    conn.execute(
        text(
            """
            CREATE TABLE amazon_store_site (
                id INTEGER PRIMARY KEY,
                store_site TEXT
            )
            """
        )
    )


def test_get_product_quality_report_counts_missing_fields(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        create_product_table(conn)
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, store_site, msku, asin, listing, brand, sales_status, product_name, updated_at
                )
                VALUES
                    (1, 'SAYOLA:US', 'MSKU-001', NULL, 'L1', 'BrandA', '在售', 'Product 1', '2026-06-01'),
                    (2, 'SAYOLA:US', 'MSKU-002', '   ', NULL, '', '停售', '', '2026-06-02'),
                    (3, 'SAYOLA:US', 'MSKU-003', 'B003', 'L3', 'BrandC', NULL, 'Product 3', '2026-06-03')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    report = get_product_quality_report()

    assert report["total"] == 3
    issues = {issue["key"]: issue for issue in report["issues"]}
    assert [issue["key"] for issue in report["field_issues"]] == [
        "missing_asin",
        "missing_listing",
        "missing_brand",
        "missing_sales_status",
        "missing_product_name",
    ]
    assert [issue["key"] for issue in report["relation_issues"]] == [
        "missing_listing_owner_config",
        "orphan_listing_owner_config",
        "unknown_product_store_site",
        "unknown_listing_owner_store_site",
    ]
    assert [issues[key] for key in [
        "missing_asin",
        "missing_listing",
        "missing_brand",
        "missing_sales_status",
        "missing_product_name",
    ]] == [
        {
            "key": "missing_asin",
            "label": "缺 ASIN",
            "field": "asin",
            "count": 2,
            "rows": [
                {"id": 2, "store_site": "SAYOLA:US", "msku": "MSKU-002", "product_name": ""},
                {"id": 1, "store_site": "SAYOLA:US", "msku": "MSKU-001", "product_name": "Product 1"},
            ],
        },
        {
            "key": "missing_listing",
            "label": "缺 Listing",
            "field": "listing",
            "count": 1,
            "rows": [
                {"id": 2, "store_site": "SAYOLA:US", "msku": "MSKU-002", "product_name": ""}
            ],
        },
        {
            "key": "missing_brand",
            "label": "缺品牌",
            "field": "brand",
            "count": 1,
            "rows": [
                {"id": 2, "store_site": "SAYOLA:US", "msku": "MSKU-002", "product_name": ""}
            ],
        },
        {
            "key": "missing_sales_status",
            "label": "缺销售状态",
            "field": "sales_status",
            "count": 1,
            "rows": [
                {"id": 3, "store_site": "SAYOLA:US", "msku": "MSKU-003", "product_name": "Product 3"}
            ],
        },
        {
            "key": "missing_product_name",
            "label": "缺产品名称",
            "field": "product_name",
            "count": 1,
            "rows": [
                {"id": 2, "store_site": "SAYOLA:US", "msku": "MSKU-002", "product_name": ""}
            ],
        },
    ]


def test_get_product_quality_report_handles_missing_listing_owner_table_with_store_sites(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        create_product_table(conn)
        create_store_site_table(conn)
        conn.execute(text("INSERT INTO amazon_store_site (id, store_site) VALUES (1, 'SAYOLA:US')"))
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, store_site, msku, asin, listing, brand, sales_status, product_name, updated_at
                )
                VALUES
                    (1, 'SAYOLA:US', 'MSKU-001', 'B001', 'L1', 'BrandA', '在售', 'Product 1', '2026-06-01'),
                    (2, 'UNKNOWN:US', 'MSKU-002', 'B002', 'L2', 'BrandA', '在售', 'Product 2', '2026-06-02')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)
    service.clear_quality_report_cache()

    report = get_product_quality_report()
    issues = {issue["key"]: issue for issue in report["relation_issues"]}

    assert issues["missing_listing_owner_config"]["count"] == 0
    assert issues["orphan_listing_owner_config"]["count"] == 0
    assert issues["unknown_product_store_site"]["count"] == 1
    assert issues["unknown_product_store_site"]["rows"] == [
        {
            "id": 2,
            "store_site": "UNKNOWN:US",
            "msku": "MSKU-002",
            "listing": "L2",
            "product_name": "Product 2",
        }
    ]
    assert issues["unknown_listing_owner_store_site"]["count"] == 0


def test_get_product_quality_report_counts_listing_owner_integrity(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        create_product_table(conn)
        create_listing_owner_table(conn)
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, store_site, msku, asin, listing, brand, sales_status, product_name, updated_at
                )
                VALUES
                    (1, 'SAYOLA:US', 'MSKU-001', 'B001', 'L1', 'BrandA', '在售', 'Product 1', '2026-06-01'),
                    (2, 'SAYOLA:US', 'MSKU-002', 'B002', 'L2', 'BrandA', '在售', 'Product 2', '2026-06-02')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (id, store_site, listing, owner)
                VALUES
                    (1, 'SAYOLA:US', 'L1', 'Alice'),
                    (2, 'SAYOLA:US', 'L404', 'Bob')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)
    service.clear_quality_report_cache()

    report = get_product_quality_report()
    issues = {issue["key"]: issue for issue in report["issues"]}

    assert issues["missing_listing_owner_config"]["count"] == 1
    assert issues["missing_listing_owner_config"]["rows"] == [
        {
            "id": 2,
            "store_site": "SAYOLA:US",
            "msku": "MSKU-002",
            "listing": "L2",
            "product_name": "Product 2",
        }
    ]
    assert issues["orphan_listing_owner_config"]["count"] == 1
    assert issues["orphan_listing_owner_config"]["rows"] == [
        {"id": 2, "store_site": "SAYOLA:US", "listing": "L404", "owner": "Bob"}
    ]
    assert [issue["key"] for issue in report["relation_issues"]] == [
        "missing_listing_owner_config",
        "orphan_listing_owner_config",
        "unknown_product_store_site",
        "unknown_listing_owner_store_site",
    ]


def test_get_product_quality_report_counts_unknown_store_site_references(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        create_product_table(conn)
        create_listing_owner_table(conn)
        create_store_site_table(conn)
        conn.execute(text("INSERT INTO amazon_store_site (id, store_site) VALUES (1, 'SAYOLA:US')"))
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, store_site, msku, asin, listing, brand, sales_status, product_name, updated_at
                )
                VALUES
                    (1, 'SAYOLA:US', 'MSKU-001', 'B001', 'L1', 'BrandA', '在售', 'Product 1', '2026-06-01'),
                    (2, 'UNKNOWN:US', 'MSKU-002', 'B002', 'L2', 'BrandA', '在售', 'Product 2', '2026-06-02')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (id, store_site, listing, owner)
                VALUES
                    (1, 'SAYOLA:US', 'L1', 'Alice'),
                    (2, 'UNKNOWN:US', 'L2', 'Bob')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)
    service.clear_quality_report_cache()

    report = get_product_quality_report()
    issues = {issue["key"]: issue for issue in report["relation_issues"]}

    assert issues["unknown_product_store_site"]["count"] == 1
    assert issues["unknown_product_store_site"]["rows"] == [
        {
            "id": 2,
            "store_site": "UNKNOWN:US",
            "msku": "MSKU-002",
            "listing": "L2",
            "product_name": "Product 2",
        }
    ]
    assert issues["unknown_listing_owner_store_site"]["count"] == 1
    assert issues["unknown_listing_owner_store_site"]["rows"] == [
        {"id": 2, "store_site": "UNKNOWN:US", "listing": "L2", "owner": "Bob"}
    ]


def test_get_product_quality_report_filters_by_store_site(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        create_product_table(conn)
        create_listing_owner_table(conn)
        create_store_site_table(conn)
        conn.execute(
            text(
                """
                INSERT INTO amazon_store_site (id, store_site)
                VALUES
                    (1, 'SAYOLA:US'),
                    (2, 'RIVBOS:CA')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, store_site, msku, asin, listing, brand, sales_status, product_name, updated_at
                )
                VALUES
                    (1, 'SAYOLA:US', 'MSKU-001', NULL, 'L1', 'BrandA', '在售', 'Product 1', '2026-06-01'),
                    (2, 'RIVBOS:CA', 'MSKU-002', NULL, 'L2', 'BrandA', '在售', 'Product 2', '2026-06-02'),
                    (3, 'UNKNOWN:US', 'MSKU-003', 'B003', 'L3', 'BrandA', '在售', 'Product 3', '2026-06-03')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (id, store_site, listing, owner)
                VALUES
                    (1, 'SAYOLA:US', 'L404', 'Alice'),
                    (2, 'RIVBOS:CA', 'L404', 'Bob'),
                    (3, 'UNKNOWN:US', 'L3', 'Chris')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)
    service.clear_quality_report_cache()

    report = get_product_quality_report(store_site="SAYOLA:US")
    issues = {issue["key"]: issue for issue in report["issues"]}

    assert report["total"] == 1
    assert issues["missing_asin"]["count"] == 1
    assert issues["missing_asin"]["rows"] == [
        {"id": 1, "store_site": "SAYOLA:US", "msku": "MSKU-001", "product_name": "Product 1"}
    ]
    assert issues["orphan_listing_owner_config"]["count"] == 1
    assert issues["orphan_listing_owner_config"]["rows"] == [
        {"id": 1, "store_site": "SAYOLA:US", "listing": "L404", "owner": "Alice"}
    ]
    assert issues["unknown_product_store_site"]["count"] == 0
    assert issues["unknown_listing_owner_store_site"]["count"] == 0


def test_build_quality_issue_workbook_exports_issue_rows(monkeypatch):
    report = {
        "total": 1,
        "issues": [
            {
                "key": "missing_asin",
                "label": "缺 ASIN",
                "field": "asin",
                "count": 1,
                "rows": [
                    {
                        "id": 1,
                        "store_site": "SAYOLA:US",
                        "msku": "MSKU-001",
                        "product_name": "Product 1",
                    }
                ],
            }
        ],
    }

    content = service.build_quality_issue_workbook(report)

    from io import BytesIO
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == ("检查项", "记录ID", "店铺站点", "MSKU", "Listing", "负责人", "产品名称")
    assert rows[1] == ("缺 ASIN", 1, "SAYOLA:US", "MSKU-001", None, None, "Product 1")


def test_get_product_quality_report_reuses_cached_report(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        create_product_table(conn)
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, store_site, msku, asin, listing, brand, sales_status, product_name, updated_at
                )
                VALUES (1, 'SAYOLA:US', 'MSKU-001', NULL, 'L1', 'BrandA', '在售', 'Product 1', '2026-06-01')
                """
            )
        )

    connect_count = 0
    original_connect = engine.connect

    def counted_connect(*args, **kwargs):
        nonlocal connect_count
        connect_count += 1
        return original_connect(*args, **kwargs)

    monkeypatch.setattr(service, "get_engine", lambda: engine)
    monkeypatch.setattr(engine, "connect", counted_connect)
    service.clear_quality_report_cache()

    assert get_product_quality_report() == get_product_quality_report()
    assert connect_count == 1


def test_data_quality_page_renders_report(monkeypatch):
    monkeypatch.setattr(
        "app.modules.data_quality.routes.get_product_quality_report",
        lambda: {
            "total": 3,
            "issues": [
                {
                    "key": "missing_asin",
                    "label": "缺 ASIN",
                    "field": "asin",
                    "count": 2,
                    "rows": [
                        {
                            "id": 1,
                            "store_site": "SAYOLA:US",
                            "msku": "MSKU-001",
                            "product_name": "Product 1",
                        }
                    ],
                }
            ],
            "field_issues": [
                {
                    "key": "missing_asin",
                    "label": "缺 ASIN",
                    "field": "asin",
                    "count": 2,
                    "rows": [
                        {
                            "id": 1,
                            "store_site": "SAYOLA:US",
                            "msku": "MSKU-001",
                            "product_name": "Product 1",
                        }
                    ],
                }
            ],
            "relation_issues": [
                {
                    "key": "missing_listing_owner_config",
                    "label": "缺 Listing 负责人配置",
                    "field": "listing",
                    "count": 1,
                    "rows": [
                        {
                            "id": 2,
                            "store_site": "SAYOLA:US",
                            "msku": "MSKU-002",
                            "listing": "Listing A",
                            "product_name": "Product 2",
                        }
                    ],
                },
                {
                    "key": "orphan_listing_owner_config",
                    "label": "无产品使用的负责人配置",
                    "field": "listing",
                    "count": 1,
                    "rows": [
                        {
                            "id": 3,
                            "store_site": "SAYOLA:US",
                            "listing": "Listing B",
                            "owner": "Bob",
                        }
                    ],
                },
                {
                    "key": "unknown_product_store_site",
                    "label": "产品引用未知店铺站点",
                    "field": "store_site",
                    "count": 1,
                    "rows": [
                        {
                            "id": 4,
                            "store_site": "UNKNOWN:US",
                            "msku": "MSKU-004",
                            "listing": "Listing C",
                            "product_name": "Product 4",
                        }
                    ],
                },
            ],
        },
    )

    response = client.get("/data-quality")

    assert response.status_code == 200
    assert "数据质量" in response.text
    assert "共 3 条产品" in response.text
    assert "缺 ASIN" in response.text
    assert "MSKU-001" in response.text
    assert "/products/1" in response.text
    assert "字段完整性" in response.text
    assert "业务关系健康" in response.text
    assert "缺 Listing 负责人配置" in response.text
    assert "/listing-owners/new?store_site=SAYOLA%3AUS&amp;listing=Listing%20A" in response.text
    assert "/listing-owners?q=Listing%20B" in response.text
    assert "产品引用未知店铺站点" in response.text
    assert "/store-sites/new?store_site=UNKNOWN%3AUS" in response.text


def test_data_quality_export_route_downloads_workbook(monkeypatch):
    monkeypatch.setattr(
        "app.modules.data_quality.routes.get_product_quality_report",
        lambda: {
            "total": 1,
            "issues": [
                {
                    "key": "missing_asin",
                    "label": "缺 ASIN",
                    "field": "asin",
                    "count": 1,
                    "rows": [
                        {
                            "id": 1,
                            "store_site": "SAYOLA:US",
                            "msku": "MSKU-001",
                            "product_name": "Product 1",
                        }
                    ],
                }
            ],
        },
    )

    response = client.get("/data-quality/export")

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]


def test_data_quality_page_uses_workbench_layout(monkeypatch):
    monkeypatch.setattr(
        "app.modules.data_quality.routes.get_product_quality_report",
        lambda: {
            "total": 3,
            "issues": [
                {
                    "key": "missing_asin",
                    "label": "缺 ASIN",
                    "field": "asin",
                    "count": 2,
                    "rows": [
                        {
                            "id": 1,
                            "store_site": "SAYOLA:US",
                            "msku": "MSKU-001",
                            "product_name": "Product 1",
                        }
                    ],
                },
                {
                    "key": "missing_listing",
                    "label": "缺 Listing",
                    "field": "listing",
                    "count": 0,
                    "rows": [],
                },
            ],
            "field_issues": [
                {
                    "key": "missing_asin",
                    "label": "缺 ASIN",
                    "field": "asin",
                    "count": 2,
                    "rows": [
                        {
                            "id": 1,
                            "store_site": "SAYOLA:US",
                            "msku": "MSKU-001",
                            "product_name": "Product 1",
                        }
                    ],
                },
                {
                    "key": "missing_listing",
                    "label": "缺 Listing",
                    "field": "listing",
                    "count": 0,
                    "rows": [],
                },
            ],
            "relation_issues": [
                {
                    "key": "missing_listing_owner_config",
                    "label": "缺 Listing 负责人配置",
                    "field": "listing",
                    "count": 0,
                    "rows": [],
                }
            ],
        },
    )

    response = client.get("/data-quality")

    assert response.status_code == 200
    assert 'class="quality-workbench"' in response.text
    assert "检查项概览" in response.text
    assert "问题明细" in response.text
    assert "字段完整性" in response.text
    assert "业务关系健康" in response.text
    assert "优先处理" in response.text
    assert "已通过" in response.text


def test_data_quality_page_passes_store_site_filter(monkeypatch):
    captured = {}

    def fake_get_product_quality_report(store_site=None):
        captured["store_site"] = store_site
        return {
            "total": 1,
            "issues": [],
            "field_issues": [],
            "relation_issues": [],
        }

    monkeypatch.setattr(
        "app.modules.data_quality.routes.get_product_quality_report",
        fake_get_product_quality_report,
    )

    response = client.get("/data-quality?store_site=SAYOLA%3AUS")

    assert response.status_code == 200
    assert captured["store_site"] == "SAYOLA:US"
    assert "当前店铺站点：SAYOLA:US" in response.text
    assert 'href="/data-quality"' in response.text
