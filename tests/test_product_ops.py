from datetime import datetime
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.pool import StaticPool

from app.main import app


client = TestClient(app)


def test_product_ops_overview_renders_rows_and_tabs(monkeypatch):
    def fake_list_product_ops_rows(filters):
        assert filters.q == "RB"
        return {
            "rows": [
                {
                    "store_site": "RIVBOS:US",
                    "listing": "RBK004",
                    "owner": "张三",
                    "listing_status": "维护",
                    "project_group": "项目组A",
                    "allocation_msku_count": 12,
                    "product_msku_count": 8,
                    "forecast_month_count": 3,
                    "forecast_units_total": 1200,
                    "data_status": "正常",
                    "purchase_readiness_label": "可进入采购判断",
                    "purchase_readiness_reasons": "",
                }
            ],
            "total": 1,
            "page": 1,
            "page_size": 50,
            "pages": 1,
        }

    monkeypatch.setattr("app.modules.product_ops.routes.list_product_ops_rows", fake_list_product_ops_rows)

    response = client.get("/product-ops", params={"q": "RB"})

    assert response.status_code == 200
    assert "产品经营管理" in response.text
    assert "经营总览" in response.text
    assert "销占比管理" in response.text
    assert "销售预估管理" in response.text
    assert "RIVBOS:US" in response.text
    assert "项目组A" in response.text
    assert "销占比SKU数" in response.text
    assert "产品信息数" in response.text
    assert "未来预估" in response.text
    assert "采购准备度" in response.text
    assert "可进入采购判断" in response.text
    assert "/?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/listing-profile?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/allocations?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/forecasts?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/import-preview" in response.text
    assert "/product-ops/export?q=RB" in response.text
    assert "操作" in response.text


def test_product_ops_filter_actions_use_responsive_header():
    css = Path("app/static/css/app.css").read_text(encoding="utf-8")
    template_paths = [
        Path("app/templates/product_ops/overview.html"),
        Path("app/templates/product_ops/allocations.html"),
        Path("app/templates/product_ops/forecasts.html"),
    ]

    for template_path in template_paths:
        template = template_path.read_text(encoding="utf-8")
        assert "product-ops-filter-heading" in template
        assert "md:flex-row md:items-end md:justify-between" not in template

    assert ".product-ops-filter-heading" in css
    assert "@media (min-width: 1180px)" in css
    assert ".product-ops-filter-heading .product-ops-filter-actions" in css


def test_listing_profile_page_renders_aggregated_listing_context(monkeypatch):
    monkeypatch.setattr(
        "app.modules.product_ops.routes.get_listing_profile",
        lambda store_site, listing: {
            "store_site": store_site,
            "listing": listing,
            "overview": {
                "store_site": "RIVBOS:US",
                "listing": "RBK004",
                "owner": "张三",
                "listing_status": "维护",
                "listing_maintainer": "李四",
                "project_group": "项目组A",
                "product_msku_count": 2,
                "allocation_msku_count": 2,
                "forecast_month_count": 2,
                "forecast_units_total": 2500,
                "data_status": "正常",
            },
            "issue_labels": ["正常"],
            "health_summary": {
                "status": "warning",
                "status_label": "需维护",
                "message": "1 项待维护",
            },
            "purchase_readiness": {
                "status": "review",
                "label": "需人工确认",
                "message": "存在异常数据，需确认后再进入采购判断",
                "reasons": ["销占比为0"],
            },
            "health_items": [
                {
                    "key": "product_info",
                    "label": "产品信息",
                    "status": "normal",
                    "status_label": "正常",
                    "message": "已关联 2 个产品信息",
                    "action_label": "查看产品信息",
                    "action_url": "/?store_site=RIVBOS%3AUS&listing=RBK004",
                },
                {
                    "key": "sales_allocation",
                    "label": "销占比",
                    "status": "warning",
                    "status_label": "异常",
                    "message": "存在销占比为0的数据",
                    "action_label": "维护销占比",
                    "action_url": "/product-ops/allocations?store_site=RIVBOS%3AUS&listing=RBK004",
                },
            ],
            "allocation_rows": [
                {
                    "store_site": "RIVBOS:US",
                    "listing": "RBK004",
                    "msku": "RBK004-Pink",
                    "sku": "SKU-1",
                    "style": "粉色款",
                    "style_sales_ratio": 0.35,
                    "sku_sales_ratio": 0.2,
                    "stocking_position": "备货",
                }
            ],
            "forecast_rows": [
                {
                    "store_site": "RIVBOS:US",
                    "listing": "RBK004",
                    "forecast_month": "2026-07",
                    "forecast_units": 1200,
                }
            ],
        },
        raising=False,
    )

    response = client.get("/product-ops/listing-profile", params={"store_site": "RIVBOS:US", "listing": "RBK004"})

    assert response.status_code == 200
    assert "Listing 经营档案" in response.text
    assert "RIVBOS:US" in response.text
    assert "RBK004" in response.text
    assert "张三" in response.text
    assert "产品信息数" in response.text
    assert "2" in response.text
    assert "RBK004-Pink" in response.text
    assert "35.00%" in response.text
    assert "2026-07" in response.text
    assert "1200" in response.text
    assert "经营健康度" in response.text
    assert "采购准备度" in response.text
    assert "需人工确认" in response.text
    assert "需维护" in response.text
    assert "已关联 2 个产品信息" in response.text
    assert "存在销占比为0的数据" in response.text
    assert "维护销占比" in response.text
    assert "/?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/allocations?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/allocations/import-template?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/forecasts/import-template?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text


def test_product_ops_import_preview_page_renders_summary(monkeypatch):
    monkeypatch.setattr(
        "app.modules.product_ops.routes.build_product_ops_import_preview",
        lambda: {
            "base_dir": "待合并/src_data",
            "allocation": {
                "label": "销占比",
                "file_count": 8,
                "total_rows": 7036,
                "duplicate_key_count": 0,
                "missing_key_count": 0,
                "owner_unmatched_count": 74,
                "owner_ambiguous_count": 0,
                "owner_resolved_by_allocation_count": 0,
                "owner_unresolved_ambiguous_count": 0,
                "issue_rows": [
                    {
                        "issue": "未匹配负责人配置",
                        "file": "KAYO-销占比参数表-20260601.xlsx",
                        "row_number": 1710,
                        "key": "rivbos:JP / RSY007 / JP-RSY007",
                    }
                ],
            },
            "forecast": {
                "label": "销售预估",
                "file_count": 5,
                "total_rows": 2259,
                "duplicate_key_count": 2,
                "missing_key_count": 0,
                "store_site_missing_count": 638,
                "store_site_explicit_count": 0,
                "store_site_inferred_count": 486,
                "owner_unmatched_count": 84,
                "owner_ambiguous_count": 638,
                "owner_resolved_by_allocation_count": 486,
                "owner_unresolved_ambiguous_count": 152,
                "issue_rows": [
                    {
                        "issue": "重复唯一键",
                        "file": "CHUNMEI-销售预估参数表 20260608.xlsx",
                        "row_number": 198,
                        "key": "US / SG106 / 2025-02",
                    }
                ],
            },
        },
    )

    response = client.get("/product-ops/import-preview")

    assert response.status_code == 200
    assert "导入预览" in response.text
    assert "销占比" in response.text
    assert "销售预估" in response.text
    assert "7036" in response.text
    assert "2259" in response.text
    assert "未匹配负责人配置" in response.text
    assert "重复唯一键" in response.text
    assert "源表缺店铺站点" in response.text
    assert "已明确店铺站点" in response.text
    assert "自动补店铺站点" in response.text
    assert "可由销占比定位" in response.text
    assert "仍需人工处理" in response.text
    assert "确认写入经营数据" in response.text
    assert 'action="/product-ops/import-preview/commit"' in response.text
    assert "/product-ops/import-preview" in response.text


def test_sales_allocation_management_renders_rows(monkeypatch):
    monkeypatch.setattr(
        "app.modules.product_ops.routes.list_sales_allocations",
        lambda filters: {
            "rows": [
                {
                    "id": 1,
                    "store_site": "RIVBOS:US",
                    "site": "US",
                    "listing": "RBK004",
                    "msku": "RBK004-2 W Pink",
                    "sku": "RBK004-2",
                    "owner": "张三",
                    "style": "粉色款",
                    "style_sales_ratio": 0.35,
                    "sku_sales_ratio": 0.2,
                    "stocking_position": "备货",
                    "total_shipping_days": 95,
                    "updated_at": "2026-06-10",
                }
            ],
            "total": 1,
            "page": 1,
            "page_size": 50,
            "pages": 1,
        },
    )

    response = client.get("/product-ops/allocations")

    assert response.status_code == 200
    assert "销占比管理" in response.text
    assert "RIVBOS:US" in response.text
    assert "RBK004-2 W Pink" in response.text
    assert "35.00%" in response.text
    assert "20.00%" in response.text
    assert "0.35" not in response.text
    assert "/product-ops?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/listing-profile?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/forecasts?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/import-preview" in response.text
    assert "/product-ops/allocations/export" in response.text
    assert "/product-ops/allocations/import-template" in response.text
    assert "/product-ops/allocations/import" in response.text
    assert "操作" in response.text
    assert "备货" in response.text


def test_sales_allocation_management_filters_zero_ratio_and_exposes_bulk_form(monkeypatch):
    captured = {}

    def fake_list(filters):
        captured["filters"] = filters
        return {
            "rows": [
                {
                    "id": 1,
                    "store_site": "RIVBOS:US",
                    "site": "US",
                    "listing": "RBK004",
                    "msku": "RBK004-2 W Pink",
                    "sku": "RBK004-2",
                    "owner": "张三",
                    "style": "粉色款",
                    "style_sales_ratio": 0,
                    "sku_sales_ratio": 0.2,
                    "stocking_position": "备货",
                    "total_shipping_days": 95,
                    "updated_at": "2026-06-10",
                }
            ],
            "total": 1,
            "page": 1,
            "page_size": 50,
            "pages": 1,
        }

    monkeypatch.setattr("app.modules.product_ops.routes.list_sales_allocations", fake_list)

    response = client.get("/product-ops/allocations", params={"ratio_status": "zero"})

    assert response.status_code == 200
    assert captured["filters"].ratio_status == "zero"
    assert "销占比为0" in response.text
    assert "/product-ops?data_status=missing_allocation" in response.text
    assert 'action="/product-ops/allocations/bulk-update"' in response.text
    assert 'name="row_ids" value="1"' in response.text
    assert 'name="style_sales_ratio"' in response.text
    assert 'name="sku_sales_ratio"' in response.text
    assert 'name="stocking_position"' in response.text


def test_sales_allocation_bulk_update_route_parses_percent_and_redirects(monkeypatch):
    captured = {}

    def fake_bulk_update(row_ids, updates, changed_by="system"):
        captured["row_ids"] = row_ids
        captured["updates"] = updates
        captured["changed_by"] = changed_by
        return {"updated": 2, "skipped": 0, "requested": 2}

    monkeypatch.setattr("app.modules.product_ops.routes.bulk_update_sales_allocations", fake_bulk_update)

    response = client.post(
        "/product-ops/allocations/bulk-update",
        data={
            "row_ids": ["1", "2"],
            "style_sales_ratio": "35%",
            "sku_sales_ratio": "0.12",
            "stocking_position": "少备",
            "store_site": "RIVBOS:US",
            "listing": "RBK004",
            "ratio_status": "zero",
            "page_size": "50",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert response.headers["location"] == "/product-ops/allocations?store_site=RIVBOS%3AUS&listing=RBK004&ratio_status=zero&page_size=50"
    assert captured == {
        "row_ids": [1, 2],
        "updates": {
            "style_sales_ratio": 0.35,
            "sku_sales_ratio": 0.12,
            "stocking_position": "少备",
        },
        "changed_by": "test-admin",
    }


def test_sales_allocation_import_template_route_uses_current_filters(monkeypatch):
    captured = {}

    def fake_export(filters):
        captured["filters"] = filters
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "销占比维护"
        sheet.append(["店铺站点", "站点", "Listing", "MSKU", "SKU"])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    monkeypatch.setattr("app.modules.product_ops.routes.export_sales_allocation_maintenance_template", fake_export, raising=False)

    response = client.get(
        "/product-ops/allocations/import-template",
        params={"q": "Pink", "store_site": "RIVBOS:US", "listing": "RBK004"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert 'filename="sales_allocation_maintenance_template.xlsx"' in response.headers["content-disposition"]
    assert captured["filters"].q == "Pink"
    assert captured["filters"].store_site == "RIVBOS:US"
    assert captured["filters"].listing == "RBK004"


def test_sales_allocation_maintenance_import_page_renders_upload_form():
    response = client.get("/product-ops/allocations/import")

    assert response.status_code == 200
    assert "销占比批量维护" in response.text
    assert 'action="/product-ops/allocations/import/preview"' in response.text
    assert 'name="file"' in response.text
    assert "/product-ops/allocations/import-template" in response.text


def test_sales_allocation_maintenance_preview_route_exposes_commit_when_valid(monkeypatch):
    monkeypatch.setattr(
        "app.modules.product_ops.routes.preview_sales_allocation_maintenance_import",
        lambda content: {
            "total_rows": 2,
            "valid_count": 2,
            "insert_count": 1,
            "update_count": 1,
            "error_count": 0,
            "valid_rows": [
                {
                    "row_number": 2,
                    "action": "更新",
                    "store_site": "RIVBOS:US",
                    "listing": "RBK004",
                    "msku": "RBK004-Pink",
                    "sku": "SKU-1",
                }
            ],
            "error_rows": [],
        },
    )
    monkeypatch.setattr("app.modules.product_ops.routes.save_import_upload", lambda content: "token123")

    response = client.post(
        "/product-ops/allocations/import/preview",
        files={"file": ("allocation.xlsx", b"xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    assert "销占比维护预览" in response.text
    assert "新增 1 行" in response.text
    assert "更新 1 行" in response.text
    assert 'name="import_token" value="token123"' in response.text
    assert "确认写入销占比" in response.text


def test_sales_allocation_maintenance_commit_route_writes_uploaded_file(monkeypatch):
    captured = {}
    monkeypatch.setattr("app.modules.product_ops.routes.load_import_upload", lambda token: b"xlsx")

    def fake_commit(content, changed_by):
        captured["content"] = content
        captured["changed_by"] = changed_by
        return {
            "success": True,
            "inserted_count": 1,
            "updated_count": 2,
            "skipped_count": 0,
            "message": "写入完成",
            "preview": {"total_rows": 3, "valid_count": 3, "insert_count": 1, "update_count": 2, "error_count": 0, "valid_rows": [], "error_rows": []},
        }

    monkeypatch.setattr("app.modules.product_ops.routes.commit_sales_allocation_maintenance_import", fake_commit)

    response = client.post("/product-ops/allocations/import/commit", data={"import_token": "token123"})

    assert response.status_code == 200
    assert "写入完成" in response.text
    assert "新增 1 行，更新 2 行，跳过 0 行" in response.text
    assert "/operation-logs?table_name=amazon_sales_allocation" in response.text
    assert captured == {"content": b"xlsx", "changed_by": "test-admin"}


def test_sales_forecast_management_renders_rows(monkeypatch):
    monkeypatch.setattr(
        "app.modules.product_ops.routes.list_sales_forecasts",
        lambda filters: {
            "rows": [
                {
                    "id": 1,
                    "store_site": "RIVBOS:US",
                    "site": "US",
                    "listing": "RBK004",
                    "forecast_month": "2026-07",
                    "forecast_units": 1200,
                    "updated_at": "2026-06-10",
                }
            ],
            "total": 1,
            "page": 1,
            "page_size": 50,
            "pages": 1,
        },
    )

    response = client.get("/product-ops/forecasts")

    assert response.status_code == 200
    assert "销售预估管理" in response.text
    assert "RIVBOS:US" in response.text
    assert "RBK004" in response.text
    assert "2026-07" in response.text
    assert "1200" in response.text
    assert "/product-ops?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/listing-profile?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/allocations?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/import-preview" in response.text
    assert "/product-ops/forecasts/export" in response.text
    assert "/product-ops/forecasts/import-template" in response.text
    assert "/product-ops/forecasts/import" in response.text
    assert "操作" in response.text


def test_sales_forecast_management_filters_zero_units_and_exposes_management_forms(monkeypatch):
    captured = {}

    def fake_list(filters):
        captured["filters"] = filters
        return {
            "rows": [
                {
                    "id": 1,
                    "store_site": "RIVBOS:US",
                    "site": "US",
                    "listing": "RBK004",
                    "forecast_month": "2026-07",
                    "forecast_units": 0,
                    "updated_at": "2026-06-10",
                }
            ],
            "total": 1,
            "page": 1,
            "page_size": 50,
            "pages": 1,
        }

    monkeypatch.setattr("app.modules.product_ops.routes.list_sales_forecasts", fake_list)

    response = client.get(
        "/product-ops/forecasts",
        params={"store_site": "RIVBOS:US", "listing": "RBK004", "forecast_status": "zero"},
    )

    assert response.status_code == 200
    assert captured["filters"].forecast_status == "zero"
    assert "预估为0" in response.text
    assert "/product-ops?store_site=RIVBOS%3AUS&amp;listing=RBK004&amp;data_status=missing_forecast" in response.text
    assert 'action="/product-ops/forecasts/bulk-update"' in response.text
    assert 'name="row_ids" value="1"' in response.text
    assert 'name="forecast_units"' in response.text
    assert 'action="/product-ops/forecasts/upsert"' in response.text
    assert 'name="forecast_month"' in response.text


def test_sales_forecast_bulk_update_route_updates_selected_rows(monkeypatch):
    captured = {}

    def fake_bulk_update(row_ids, forecast_units, changed_by="system"):
        captured["row_ids"] = row_ids
        captured["forecast_units"] = forecast_units
        captured["changed_by"] = changed_by
        return {"updated": 2, "skipped": 0, "requested": 2}

    monkeypatch.setattr("app.modules.product_ops.routes.bulk_update_sales_forecasts", fake_bulk_update)

    response = client.post(
        "/product-ops/forecasts/bulk-update",
        data={
            "row_ids": ["1", "2"],
            "forecast_units": "1200",
            "store_site": "RIVBOS:US",
            "listing": "RBK004",
            "forecast_status": "zero",
            "page_size": "50",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert response.headers["location"] == "/product-ops/forecasts?store_site=RIVBOS%3AUS&listing=RBK004&forecast_status=zero&page_size=50"
    assert captured == {"row_ids": [1, 2], "forecast_units": 1200, "changed_by": "test-admin"}


def test_sales_forecast_upsert_route_adds_listing_month(monkeypatch):
    captured = {}

    def fake_upsert(payload, changed_by="system"):
        captured["payload"] = payload
        captured["changed_by"] = changed_by
        return {"action": "inserted", "id": 3}

    monkeypatch.setattr("app.modules.product_ops.routes.upsert_sales_forecast", fake_upsert)

    response = client.post(
        "/product-ops/forecasts/upsert",
        data={
            "store_site": "RIVBOS:US",
            "site": "US",
            "listing": "RBK004",
            "forecast_month": "2026-09",
            "forecast_units": "1300",
            "page_size": "50",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert response.headers["location"] == "/product-ops/forecasts?store_site=RIVBOS%3AUS&site=US&listing=RBK004&page_size=50"
    assert captured == {
        "payload": {
            "store_site": "RIVBOS:US",
            "site": "US",
            "listing": "RBK004",
            "forecast_month": "2026-09",
            "forecast_units": 1300,
        },
        "changed_by": "test-admin",
    }


def test_sales_forecast_import_template_route_uses_current_filters(monkeypatch):
    captured = {}

    def fake_export(filters):
        captured["filters"] = filters
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "销售预估维护"
        sheet.append(["店铺站点", "站点", "Listing", "月份", "Listing月度预估销量"])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    monkeypatch.setattr("app.modules.product_ops.routes.export_sales_forecast_maintenance_template", fake_export, raising=False)

    response = client.get(
        "/product-ops/forecasts/import-template",
        params={"q": "RB", "store_site": "RIVBOS:US", "site": "US", "listing": "RBK004"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert 'filename="sales_forecast_maintenance_template.xlsx"' in response.headers["content-disposition"]
    assert captured["filters"].q == "RB"
    assert captured["filters"].store_site == "RIVBOS:US"
    assert captured["filters"].site == "US"
    assert captured["filters"].listing == "RBK004"


def test_sales_forecast_maintenance_import_page_renders_upload_form():
    response = client.get("/product-ops/forecasts/import")

    assert response.status_code == 200
    assert "销售预估批量维护" in response.text
    assert 'action="/product-ops/forecasts/import/preview"' in response.text
    assert 'name="file"' in response.text
    assert "/product-ops/forecasts/import-template" in response.text


def test_sales_forecast_maintenance_preview_route_exposes_commit_when_valid(monkeypatch):
    monkeypatch.setattr(
        "app.modules.product_ops.routes.preview_sales_forecast_maintenance_import",
        lambda content: {
            "total_rows": 2,
            "valid_count": 2,
            "insert_count": 1,
            "update_count": 1,
            "error_count": 0,
            "valid_rows": [
                {
                    "row_number": 2,
                    "action": "更新",
                    "store_site": "RIVBOS:US",
                    "listing": "RBK004",
                    "forecast_month": "2026-07",
                    "forecast_units": 1200,
                }
            ],
            "error_rows": [],
        },
    )
    monkeypatch.setattr("app.modules.product_ops.routes.save_import_upload", lambda content: "token123")

    response = client.post(
        "/product-ops/forecasts/import/preview",
        files={"file": ("forecast.xlsx", b"xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    assert "销售预估维护预览" in response.text
    assert "新增 1 行" in response.text
    assert "更新 1 行" in response.text
    assert 'name="import_token" value="token123"' in response.text
    assert "确认写入销售预估" in response.text


def test_sales_forecast_maintenance_commit_route_writes_uploaded_file(monkeypatch):
    captured = {}
    monkeypatch.setattr("app.modules.product_ops.routes.load_import_upload", lambda token: b"xlsx")

    def fake_commit(content, changed_by):
        captured["content"] = content
        captured["changed_by"] = changed_by
        return {
            "success": True,
            "inserted_count": 1,
            "updated_count": 2,
            "skipped_count": 0,
            "message": "写入完成",
            "preview": {"total_rows": 3, "valid_count": 3, "insert_count": 1, "update_count": 2, "error_count": 0, "valid_rows": [], "error_rows": []},
        }

    monkeypatch.setattr("app.modules.product_ops.routes.commit_sales_forecast_maintenance_import", fake_commit)

    response = client.post("/product-ops/forecasts/import/commit", data={"import_token": "token123"})

    assert response.status_code == 200
    assert "写入完成" in response.text
    assert "新增 1 行，更新 2 行，跳过 0 行" in response.text
    assert "/operation-logs?table_name=amazon_sales_forecast" in response.text
    assert captured == {"content": b"xlsx", "changed_by": "test-admin"}


def test_product_ops_overview_export_route_uses_current_filters(monkeypatch):
    captured = {}

    def fake_export(filters):
        captured["filters"] = filters
        return b"xlsx"

    monkeypatch.setattr("app.modules.product_ops.routes.export_product_ops_rows_to_xlsx", fake_export, raising=False)

    response = client.get(
        "/product-ops/export",
        params={
            "q": "RB",
            "store_site": "RIVBOS:US",
            "listing": "RBK004",
            "brand": "RIVBOS",
            "data_status": "missing_product",
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert 'filename="product_ops_overview.xlsx"' in response.headers["content-disposition"]
    assert captured["filters"].q == "RB"
    assert captured["filters"].store_site == "RIVBOS:US"
    assert captured["filters"].listing == "RBK004"
    assert captured["filters"].brand == "RIVBOS"
    assert captured["filters"].data_status == "missing_product"


def test_sales_allocation_export_route_downloads_filtered_workbook(monkeypatch):
    captured = {}

    def fake_export(filters):
        captured["filters"] = filters
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "销占比"
        sheet.append(["店铺站点", "Listing", "MSKU", "款式销占比"])
        sheet.append(["RIVBOS:US", "RBK004", "RBK004-Pink", 0.35])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    monkeypatch.setattr("app.modules.product_ops.routes.export_sales_allocations_to_xlsx", fake_export, raising=False)

    response = client.get(
        "/product-ops/allocations/export",
        params={"q": "Pink", "store_site": "RIVBOS:US", "listing": "RBK004"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert 'filename="sales_allocations.xlsx"' in response.headers["content-disposition"]
    assert captured["filters"].q == "Pink"
    assert captured["filters"].store_site == "RIVBOS:US"
    assert captured["filters"].listing == "RBK004"
    workbook = load_workbook(BytesIO(response.content))
    assert workbook["销占比"].cell(row=2, column=1).value == "RIVBOS:US"


def test_sales_forecast_export_route_downloads_filtered_workbook(monkeypatch):
    captured = {}

    def fake_export(filters):
        captured["filters"] = filters
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "销售预估"
        sheet.append(["店铺站点", "站点", "Listing", "月份", "销量"])
        sheet.append(["RIVBOS:US", "US", "RBK004", "2026-07-01", 1200])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    monkeypatch.setattr("app.modules.product_ops.routes.export_sales_forecasts_to_xlsx", fake_export, raising=False)

    response = client.get(
        "/product-ops/forecasts/export",
        params={"q": "RB", "store_site": "RIVBOS:US", "site": "US", "listing": "RBK004"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert 'filename="sales_forecasts.xlsx"' in response.headers["content-disposition"]
    assert captured["filters"].q == "RB"
    assert captured["filters"].store_site == "RIVBOS:US"
    assert captured["filters"].site == "US"
    assert captured["filters"].listing == "RBK004"
    workbook = load_workbook(BytesIO(response.content))
    assert workbook["销售预估"].cell(row=2, column=5).value == 1200


def test_product_ops_gaps_page_renders_grouped_issues(monkeypatch):
    monkeypatch.setattr(
        "app.modules.product_ops.routes.list_product_ops_gaps",
        lambda: {
            "summary": {
                "missing_product": 2,
                "missing_allocation": 1,
                "missing_forecast": 1,
                "zero_allocation_ratio": 1,
                "zero_forecast_units": 1,
            },
            "groups": [
                {
                    "key": "missing_product",
                    "label": "缺产品信息",
                    "rows": [
                        {
                            "store_site": "SEEKWAY:JP",
                            "listing": "SWS001",
                            "owner": "KAYO",
                            "allocation_msku_count": 112,
                            "forecast_month_count": 48,
                            "product_msku_count": 0,
                            "data_status": "缺产品信息",
                        }
                    ],
                },
                {
                    "key": "missing_allocation",
                    "label": "缺销占比",
                    "rows": [
                        {
                            "store_site": "RIVMOUNT:CA",
                            "listing": "SWS002",
                            "owner": "ASH",
                            "allocation_msku_count": 0,
                            "forecast_month_count": 20,
                            "product_msku_count": 0,
                            "data_status": "缺销占比/缺产品信息",
                        }
                    ],
                },
                {
                    "key": "missing_forecast",
                    "label": "缺销售预估",
                    "rows": [
                        {
                            "store_site": "RIVBOS:US",
                            "listing": "RBK004",
                            "owner": "张三",
                            "allocation_msku_count": 10,
                            "forecast_month_count": 0,
                            "product_msku_count": 10,
                            "data_status": "缺销售预估",
                        }
                    ],
                },
                {
                    "key": "zero_allocation_ratio",
                    "label": "销占比为0",
                    "rows": [
                        {
                            "store_site": "RIVMOUNT:US",
                            "listing": "SWS003",
                            "owner": "ASH",
                            "allocation_msku_count": 2,
                            "forecast_month_count": 12,
                            "product_msku_count": 2,
                            "data_status": "销占比为0",
                        }
                    ],
                },
                {
                    "key": "zero_forecast_units",
                    "label": "销售预估为0",
                    "rows": [
                        {
                            "store_site": "RIVBOS:CA",
                            "listing": "RYS007",
                            "owner": "KAYO",
                            "allocation_msku_count": 3,
                            "forecast_month_count": 2,
                            "product_msku_count": 3,
                            "data_status": "销售预估为0",
                        }
                    ],
                },
            ],
        },
    )

    response = client.get("/product-ops/gaps")

    assert response.status_code == 200
    assert "经营问题工作台" in response.text
    assert "缺产品信息" in response.text
    assert "缺销占比" in response.text
    assert "缺销售预估" in response.text
    assert "销占比为0" in response.text
    assert "销售预估为0" in response.text
    assert "SEEKWAY:JP" in response.text
    assert "RIVBOS:CA" in response.text
    assert "/product-ops?store_site=SEEKWAY%3AJP&amp;listing=SWS001" in response.text
    assert "/product-ops/listing-profile?store_site=SEEKWAY%3AJP&amp;listing=SWS001" in response.text
    assert "/?store_site=SEEKWAY%3AJP&amp;listing=SWS001" in response.text
    assert "/product-ops/allocations?store_site=SEEKWAY%3AJP&amp;listing=SWS001" in response.text
    assert "/product-ops/forecasts?store_site=SEEKWAY%3AJP&amp;listing=SWS001" in response.text
    assert "/product-ops/allocations?store_site=RIVMOUNT%3AUS&amp;listing=SWS003&amp;ratio_status=zero" in response.text
    assert "/product-ops/forecasts?store_site=RIVBOS%3ACA&amp;listing=RYS007&amp;forecast_status=zero" in response.text
    assert "/product-ops/allocations/import-template?store_site=RIVMOUNT%3ACA&amp;listing=SWS002" in response.text
    assert "/product-ops/allocations/import" in response.text
    assert "/product-ops/forecasts/import-template?store_site=RIVBOS%3AUS&amp;listing=RBK004" in response.text
    assert "/product-ops/forecasts/import" in response.text
    assert "/product-ops/gaps/export" in response.text


def test_product_ops_gaps_export_route_downloads_workbook(monkeypatch):
    def fake_export():
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "缺口汇总"
        sheet.append(["类型", "数量"])
        sheet.append(["缺产品信息", 2])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    monkeypatch.setattr("app.modules.product_ops.routes.export_product_ops_gaps_to_xlsx", fake_export, raising=False)

    response = client.get("/product-ops/gaps/export")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert 'filename="product_ops_gaps.xlsx"' in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content))
    assert workbook["缺口汇总"].cell(row=2, column=2).value == 2


def test_get_listing_profile_aggregates_listing_context(monkeypatch):
    from app.modules.product_ops import service

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    listing_status TEXT,
                    listing_maintainer TEXT,
                    project_group TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    msku TEXT,
                    brand TEXT,
                    category_a TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_allocation (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    owner TEXT,
                    listing TEXT,
                    style TEXT,
                    msku TEXT,
                    sku TEXT,
                    scale_position TEXT,
                    style_sales_ratio REAL,
                    sku_sales_ratio REAL,
                    demand_position TEXT,
                    shipping_position TEXT,
                    stocking_position TEXT,
                    operation_min_order_days INTEGER,
                    total_shipping_days INTEGER,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_forecast (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    listing TEXT,
                    forecast_month TEXT,
                    forecast_units INTEGER,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (
                    id, store_site, listing, owner, listing_status, listing_maintainer, project_group
                )
                VALUES (1, 'RIVBOS:US', 'RBK004', '张三', '维护', '李四', '项目组A')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (id, store_site, listing, msku, brand, category_a)
                VALUES
                    (1, 'RIVBOS:US', 'RBK004', 'RBK004-Pink', 'RIVBOS', '配件'),
                    (2, 'RIVBOS:US', 'RBK004', 'RBK004-Blue', 'RIVBOS', '配件')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_allocation (
                    id, store_site, site, owner, listing, style, msku, sku,
                    style_sales_ratio, sku_sales_ratio, stocking_position, updated_at
                )
                VALUES
                    (1, 'RIVBOS:US', 'US', '张三', 'RBK004', '粉色款', 'RBK004-Pink', 'SKU-1', 0.35, 0.2, '备货', '2026-06-10'),
                    (2, 'RIVBOS:US', 'US', '张三', 'RBK004', '蓝色款', 'RBK004-Blue', 'SKU-2', 0, 0.1, '少备', '2026-06-10')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_forecast (
                    id, store_site, site, listing, forecast_month, forecast_units, updated_at
                )
                VALUES
                    (1, 'RIVBOS:US', 'US', 'RBK004', '2026-07', 1200, '2026-06-10'),
                    (2, 'RIVBOS:US', 'US', 'RBK004', '2026-08', 1300, '2026-06-10')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    profile = service.get_listing_profile("RIVBOS:US", "RBK004")

    assert profile["store_site"] == "RIVBOS:US"
    assert profile["listing"] == "RBK004"
    assert profile["overview"]["owner"] == "张三"
    assert profile["overview"]["product_msku_count"] == 2
    assert profile["overview"]["allocation_msku_count"] == 2
    assert profile["overview"]["forecast_month_count"] == 2
    assert profile["overview"]["forecast_units_total"] == 2500
    assert [row["msku"] for row in profile["allocation_rows"]] == ["RBK004-Blue", "RBK004-Pink"]
    assert [row["forecast_month"] for row in profile["forecast_rows"]] == ["2026-07", "2026-08"]
    assert profile["issue_labels"] == ["销占比为0"]
    assert profile["health_summary"]["status"] == "warning"
    assert profile["health_summary"]["status_label"] == "需维护"
    assert profile["purchase_readiness"]["status"] == "review"
    assert profile["purchase_readiness"]["label"] == "需人工确认"
    assert "销占比为0" in profile["purchase_readiness"]["reasons"]
    health_by_key = {item["key"]: item for item in profile["health_items"]}
    assert health_by_key["product_info"]["status_label"] == "正常"
    assert health_by_key["owner_config"]["status_label"] == "正常"
    assert health_by_key["sales_allocation"]["status_label"] == "异常"
    assert "销占比为0" in health_by_key["sales_allocation"]["message"]
    assert health_by_key["sales_forecast"]["status_label"] == "正常"


def test_sales_allocation_zero_ratio_filter_and_bulk_update_write_logs(monkeypatch):
    from app.modules.product_ops import service

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_allocation (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    owner TEXT,
                    listing TEXT,
                    style TEXT,
                    msku TEXT,
                    sku TEXT,
                    scale_position TEXT,
                    style_sales_ratio REAL,
                    sku_sales_ratio REAL,
                    demand_position TEXT,
                    shipping_position TEXT,
                    stocking_position TEXT,
                    operation_min_order_days INTEGER,
                    total_shipping_days INTEGER,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_operation_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    table_name TEXT,
                    record_id INTEGER,
                    operation_type TEXT,
                    changed_by TEXT,
                    change_data TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_allocation (
                    id, store_site, listing, msku, sku, style_sales_ratio, sku_sales_ratio, stocking_position
                )
                VALUES
                    (1, 'RIVBOS:US', 'RBK004', 'MSKU-1', 'SKU-1', 0, 0.20, '备货'),
                    (2, 'RIVBOS:US', 'RBK004', 'MSKU-2', 'SKU-2', 0.30, 0, '备货'),
                    (3, 'RIVBOS:US', 'RBK004', 'MSKU-3', 'SKU-3', 0.30, 0.10, '备货')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    page = service.list_sales_allocations(service.SalesAllocationFilters(ratio_status="zero"))

    assert [row["id"] for row in page["rows"]] == [1, 2]

    result = service.bulk_update_sales_allocations(
        [1, 2],
        {"style_sales_ratio": 0.35, "sku_sales_ratio": 0.12, "stocking_position": "少备"},
        changed_by="tester",
    )

    assert result == {"updated": 2, "skipped": 0, "requested": 2}
    with engine.connect() as conn:
        rows = [
            dict(row)
            for row in conn.execute(
                text(
                    """
                    SELECT id, style_sales_ratio, sku_sales_ratio, stocking_position
                    FROM amazon_sales_allocation
                    ORDER BY id
                    """
                )
            ).mappings()
        ]
        logs = [
            dict(row)
            for row in conn.execute(
                text(
                    """
                    SELECT table_name, record_id, operation_type, changed_by, change_data
                    FROM amazon_operation_log
                    ORDER BY record_id
                    """
                )
            ).mappings()
        ]

    assert rows == [
        {"id": 1, "style_sales_ratio": 0.35, "sku_sales_ratio": 0.12, "stocking_position": "少备"},
        {"id": 2, "style_sales_ratio": 0.35, "sku_sales_ratio": 0.12, "stocking_position": "少备"},
        {"id": 3, "style_sales_ratio": 0.30, "sku_sales_ratio": 0.10, "stocking_position": "备货"},
    ]
    assert [(log["table_name"], log["record_id"], log["operation_type"], log["changed_by"]) for log in logs] == [
        ("amazon_sales_allocation", 1, "BULK_UPDATE", "tester"),
        ("amazon_sales_allocation", 2, "BULK_UPDATE", "tester"),
    ]
    assert "style_sales_ratio" in logs[0]["change_data"]
    assert "sku_sales_ratio" in logs[1]["change_data"]


def test_sales_forecast_zero_units_filter_bulk_update_upsert_and_health(monkeypatch):
    from app.modules.product_ops import service

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    listing_status TEXT,
                    listing_maintainer TEXT,
                    project_group TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    msku TEXT,
                    brand TEXT,
                    category_a TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_allocation (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    owner TEXT,
                    listing TEXT,
                    style TEXT,
                    msku TEXT,
                    sku TEXT,
                    scale_position TEXT,
                    style_sales_ratio REAL,
                    sku_sales_ratio REAL,
                    demand_position TEXT,
                    shipping_position TEXT,
                    stocking_position TEXT,
                    operation_min_order_days INTEGER,
                    total_shipping_days INTEGER,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_forecast (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    listing TEXT,
                    forecast_month TEXT,
                    forecast_units REAL,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_operation_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    table_name TEXT,
                    record_id INTEGER,
                    operation_type TEXT,
                    changed_by TEXT,
                    change_data TEXT
                )
                """
            )
        )
        conn.execute(text("INSERT INTO amazon_listing_owner_config VALUES (1, 'RIVBOS:US', 'RBK004', '张三', '维护', '李四', '项目组A')"))
        conn.execute(text("INSERT INTO amazon_product_info VALUES (1, 'RIVBOS:US', 'RBK004', 'MSKU-1', 'RIVBOS', '配件')"))
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_allocation (
                    id, store_site, listing, msku, style_sales_ratio, sku_sales_ratio
                )
                VALUES (1, 'RIVBOS:US', 'RBK004', 'MSKU-1', 0.35, 0.2)
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_forecast (id, store_site, site, listing, forecast_month, forecast_units)
                VALUES
                    (1, 'RIVBOS:US', 'US', 'RBK004', '2026-07', 0),
                    (2, 'RIVBOS:US', 'US', 'RBK004', '2026-08', 100)
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    page = service.list_sales_forecasts(service.SalesForecastFilters(forecast_status="zero"))
    profile_before = service.get_listing_profile("RIVBOS:US", "RBK004")

    assert [row["id"] for row in page["rows"]] == [1]
    assert {item["key"]: item["status_label"] for item in profile_before["health_items"]}["sales_forecast"] == "异常"

    update_result = service.bulk_update_sales_forecasts([1], 1200, changed_by="tester")
    upsert_result = service.upsert_sales_forecast(
        {
            "store_site": "RIVBOS:US",
            "site": "US",
            "listing": "RBK004",
            "forecast_month": "2026-09",
            "forecast_units": 1300,
        },
        changed_by="tester",
    )
    profile_after = service.get_listing_profile("RIVBOS:US", "RBK004")

    assert update_result == {"updated": 1, "skipped": 0, "requested": 1}
    assert upsert_result["action"] == "inserted"
    assert {item["key"]: item["status_label"] for item in profile_after["health_items"]}["sales_forecast"] == "正常"
    with engine.connect() as conn:
        rows = [
            dict(row)
            for row in conn.execute(
                text(
                    """
                    SELECT forecast_month, forecast_units
                    FROM amazon_sales_forecast
                    ORDER BY forecast_month
                    """
                )
            ).mappings()
        ]
        logs = [
            dict(row)
            for row in conn.execute(
                text(
                    """
                    SELECT table_name, operation_type, changed_by, change_data
                    FROM amazon_operation_log
                    ORDER BY id
                    """
                )
            ).mappings()
        ]

    assert rows == [
        {"forecast_month": "2026-07", "forecast_units": 1200},
        {"forecast_month": "2026-08", "forecast_units": 100},
        {"forecast_month": "2026-09-01", "forecast_units": 1300},
    ]
    assert [(log["table_name"], log["operation_type"], log["changed_by"]) for log in logs] == [
        ("amazon_sales_forecast", "BULK_UPDATE", "tester"),
        ("amazon_sales_forecast", "MANUAL_INSERT", "tester"),
    ]
    assert "forecast_units" in logs[0]["change_data"]


def test_list_product_ops_gaps_groups_missing_and_zero_ratio_rows(monkeypatch):
    from app.modules.product_ops import service

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    listing_status TEXT,
                    listing_maintainer TEXT,
                    project_group TEXT,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    msku TEXT,
                    brand TEXT,
                    category_a TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_allocation (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    msku TEXT,
                    style_sales_ratio REAL,
                    sku_sales_ratio REAL,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_forecast (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    forecast_month TEXT,
                    forecast_units REAL
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (
                    id, store_site, listing, owner, listing_status, listing_maintainer, project_group
                )
                VALUES
                    (1, 'A:US', 'L1', '张三', '维护', '组长', '一组'),
                    (2, 'A:US', 'L2', '李四', '维护', '组长', '一组'),
                    (3, 'A:US', 'L3', '王五', '维护', '组长', '二组')
                """
            )
        )
        conn.execute(text("INSERT INTO amazon_product_info (id, store_site, listing, msku, brand, category_a) VALUES (1, 'A:US', 'L1', 'M1', 'B', 'C')"))
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_allocation (id, store_site, listing, msku, style_sales_ratio, sku_sales_ratio)
                VALUES
                    (1, 'A:US', 'L1', 'M1', 0.2, 0.1),
                    (2, 'A:US', 'L3', 'M3', 0, 0)
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_forecast (id, store_site, listing, forecast_month, forecast_units)
                VALUES
                    (1, 'A:US', 'L1', '2026-07-01', 0),
                    (2, 'A:US', 'L2', '2026-07-01', 100)
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    gaps = service.list_product_ops_gaps()

    assert gaps["summary"] == {
        "missing_product": 2,
        "missing_allocation": 1,
        "missing_forecast": 1,
        "zero_allocation_ratio": 1,
        "zero_forecast_units": 1,
    }
    groups = {group["key"]: group for group in gaps["groups"]}
    assert [row["listing"] for row in groups["missing_product"]["rows"]] == ["L2", "L3"]
    assert [row["listing"] for row in groups["missing_allocation"]["rows"]] == ["L2"]
    assert [row["listing"] for row in groups["missing_forecast"]["rows"]] == ["L3"]
    assert [row["listing"] for row in groups["zero_allocation_ratio"]["rows"]] == ["L3"]
    assert [row["listing"] for row in groups["zero_forecast_units"]["rows"]] == ["L1"]


def test_list_product_ops_gaps_counts_all_owner_rows_beyond_first_page(monkeypatch):
    from app.modules.product_ops import service

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    listing_status TEXT,
                    listing_maintainer TEXT,
                    project_group TEXT,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (
                    id, store_site, listing, owner, listing_status, listing_maintainer, project_group
                )
                VALUES (:id, :store_site, :listing, :owner, '维护', '组长', '项目组A')
                """
            ),
            [
                {
                    "id": index,
                    "store_site": "RIVBOS:US",
                    "listing": f"RB{index:03d}",
                    "owner": "张三",
                }
                for index in range(1, 202)
            ],
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    gaps = service.list_product_ops_gaps()

    assert gaps["summary"]["missing_product"] == 201
    assert gaps["summary"]["missing_allocation"] == 201
    assert gaps["summary"]["missing_forecast"] == 201
    assert len(gaps["groups"][0]["rows"]) == 20


def test_build_product_ops_import_preview_validates_local_workbooks(tmp_path, monkeypatch):
    from app.modules.product_ops import import_preview

    allocation_dir = tmp_path / "销占比参数"
    forecast_dir = tmp_path / "销售预估参数"
    allocation_dir.mkdir()
    forecast_dir.mkdir()

    allocation_wb = Workbook()
    allocation_ws = allocation_wb.active
    allocation_ws.title = "parameter"
    allocation_ws.append(["站点", "负责人", "Listing", "款式", "MSKU", "积加SKU", "店铺"])
    allocation_ws.append(["US", "张三", "RBK004", "粉色款", "RBK004-Pink", "SKU-1", "RIVBOS"])
    allocation_ws.append(["US", "张三", "MISSING", "蓝色款", "RBK004-Blue", "SKU-2", "RIVBOS"])
    allocation_ws.append(["US", "李四", "SK001", "黑色款", "SK001-Black", "SKU-3", "SEEKWAY"])
    allocation_wb.save(allocation_dir / "allocation.xlsx")

    forecast_wb = Workbook()
    forecast_ws = forecast_wb.active
    forecast_ws.title = "销量预估表-parameter"
    forecast_ws.append(["店铺/站点", "站点", "Listing", "月份", "Listing_月度预估销量"])
    forecast_ws.append(["RIVBOS:US", "US", "RBK004", datetime(2026, 7, 1), 1200])
    forecast_ws.append([None, "US", "SK001", datetime(2026, 8, 1), 800])
    forecast_ws.append([None, "US", "SK001", datetime(2026, 8, 1), 800])
    forecast_ws.append([None, "US", "SWS002", datetime(2026, 9, 1), 500])
    forecast_wb.save(forecast_dir / "forecast.xlsx")

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (id, store_site, listing)
                VALUES
                    (1, 'RIVBOS:US', 'RBK004'),
                    (2, 'SEEKWAY:US', 'SK001'),
                    (3, 'SIMARI:US', 'SK001'),
                    (4, 'RIVMOUNT:US', 'SWS002'),
                    (5, 'SIMARI:US', 'SWS002')
                """
            )
        )

    monkeypatch.setattr(import_preview, "get_engine", lambda: engine)

    preview = import_preview.build_product_ops_import_preview(tmp_path)

    assert preview["allocation"]["file_count"] == 1
    assert preview["allocation"]["total_rows"] == 3
    assert preview["allocation"]["missing_key_count"] == 0
    assert preview["allocation"]["duplicate_key_count"] == 0
    assert preview["allocation"]["owner_unmatched_count"] == 1
    assert preview["forecast"]["file_count"] == 1
    assert preview["forecast"]["total_rows"] == 4
    assert preview["forecast"]["store_site_explicit_count"] == 1
    assert preview["forecast"]["store_site_missing_count"] == 3
    assert preview["forecast"]["store_site_inferred_count"] == 2
    assert preview["forecast"]["duplicate_key_count"] == 2
    assert preview["forecast"]["owner_ambiguous_count"] == 3
    assert preview["forecast"]["owner_resolved_by_allocation_count"] == 2
    assert preview["forecast"]["owner_unresolved_ambiguous_count"] == 1
    assert any(row["issue"] == "未匹配负责人配置" for row in preview["allocation"]["issue_rows"])
    assert any(row["issue"] == "重复唯一键" for row in preview["forecast"]["issue_rows"])


def test_build_forecast_store_site_review_extracts_unresolved_rows(tmp_path, monkeypatch):
    from app.modules.product_ops import import_preview

    allocation_dir = tmp_path / "销占比参数"
    forecast_dir = tmp_path / "销售预估参数"
    allocation_dir.mkdir()
    forecast_dir.mkdir()

    allocation_wb = Workbook()
    allocation_ws = allocation_wb.active
    allocation_ws.title = "parameter"
    allocation_ws.append(["站点", "负责人", "Listing", "款式", "MSKU", "积加SKU", "店铺"])
    allocation_ws.append(["US", "张三", "SK001", "黑色款", "SK001-Black", "SKU-3", "SEEKWAY"])
    allocation_wb.save(allocation_dir / "allocation.xlsx")

    forecast_wb = Workbook()
    forecast_ws = forecast_wb.active
    forecast_ws.title = "销量预估表-parameter"
    forecast_ws.append(["店铺/站点", "站点", "Listing", "月份", "Listing_月度预估销量"])
    forecast_ws.append(["RIVBOS:US", "US", "RBK004", datetime(2026, 7, 1), 1200])
    forecast_ws.append([None, "US", "SK001", datetime(2026, 8, 1), 800])
    forecast_ws.append([None, "US", "SWS002", datetime(2026, 9, 1), 500])
    forecast_wb.save(forecast_dir / "forecast.xlsx")

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    listing_status TEXT,
                    project_group TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (
                    id, store_site, listing, owner, listing_status, project_group
                )
                VALUES
                    (1, 'RIVBOS:US', 'RBK004', '张三', '正常', '项目组A'),
                    (2, 'SEEKWAY:US', 'SK001', '李四', '维护', '项目组B'),
                    (3, 'SIMARI:US', 'SK001', '王五', '维护', '项目组C'),
                    (4, 'RIVMOUNT:US', 'SWS002', 'ASH', '维护', '备货2'),
                    (5, 'SIMARI:US', 'SWS002', 'None', '已停运', '无')
                """
            )
        )

    monkeypatch.setattr(import_preview, "get_engine", lambda: engine)

    review = import_preview.build_forecast_store_site_review(tmp_path)

    assert review["total"] == 1
    row = review["rows"][0]
    assert row["site"] == "US"
    assert row["listing"] == "SWS002"
    assert row["forecast_month"] == "2026-09"
    assert row["forecast_units"] == 500
    assert row["candidate_store_sites"] == "RIVMOUNT:US, SIMARI:US"
    assert "RIVMOUNT:US / ASH / 维护 / 备货2" in row["owner_candidates"]
    assert "SIMARI:US / - / 已停运 / 无" in row["owner_candidates"]


def test_forecast_store_site_review_page_renders_rows(monkeypatch):
    monkeypatch.setattr(
        "app.modules.product_ops.routes.build_forecast_store_site_review",
        lambda: {
            "base_dir": "待合并/src_data",
            "total": 1,
            "rows": [
                {
                    "file": "ASH-销售预估参数表 20260525.xlsx",
                    "row_number": 224,
                    "site": "CA",
                    "listing": "SWS002",
                    "forecast_month": "2025-05",
                    "forecast_units": 300,
                    "candidate_store_sites": "RIVMOUNT:CA, SIMARI:CA",
                    "owner_candidates": "RIVMOUNT:CA / ASH / 维护 / 备货2",
                }
            ],
        },
    )

    response = client.get("/product-ops/import-preview/forecast-store-site-review")

    assert response.status_code == 200
    assert "销售预估店铺站点待排查" in response.text
    assert "RIVMOUNT:CA, SIMARI:CA" in response.text
    assert "ASH-销售预估参数表" in response.text
    assert "/product-ops/import-preview/forecast-store-site-review.xlsx" in response.text


def test_forecast_store_site_review_export_downloads_workbook(monkeypatch):
    monkeypatch.setattr(
        "app.modules.product_ops.routes.build_forecast_store_site_review",
        lambda: {
            "base_dir": "待合并/src_data",
            "total": 1,
            "rows": [
                {
                    "file": "ASH-销售预估参数表 20260525.xlsx",
                    "row_number": 224,
                    "site": "CA",
                    "listing": "SWS002",
                    "forecast_month": "2025-05",
                    "forecast_units": 300,
                    "candidate_store_sites": "RIVMOUNT:CA, SIMARI:CA",
                    "owner_candidates": "RIVMOUNT:CA / ASH / 维护 / 备货2",
                }
            ],
        },
    )

    response = client.get("/product-ops/import-preview/forecast-store-site-review.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    assert rows[0] == (
        "文件",
        "行号",
        "站点",
        "Listing",
        "月份",
        "预估销量",
        "候选店铺站点",
        "候选负责人配置",
        "确认店铺站点",
    )
    assert rows[1][0] == "ASH-销售预估参数表 20260525.xlsx"
    assert rows[1][6] == "RIVMOUNT:CA, SIMARI:CA"
    assert rows[1][8] is None


def test_preview_forecast_store_site_corrections_validates_confirmed_store_site(monkeypatch):
    from app.modules.product_ops import import_preview

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["文件", "行号", "站点", "Listing", "月份", "预估销量", "候选店铺站点", "候选负责人配置", "确认店铺站点"])
    sheet.append(["forecast.xlsx", 2, "US", "RBK004", "2026-07", 1200, "RIVBOS:US", "", "RIVBOS:US"])
    sheet.append(["forecast.xlsx", 3, "US", "SWS002", "2026-08", 800, "RIVMOUNT:US, SIMARI:US", "", ""])
    sheet.append(["forecast.xlsx", 4, "US", "SWS002", "2026-09", 500, "RIVMOUNT:US, SIMARI:US", "", "UNKNOWN:US"])
    output = BytesIO()
    workbook.save(output)

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (id, store_site, listing)
                VALUES
                    (1, 'RIVBOS:US', 'RBK004'),
                    (2, 'RIVMOUNT:US', 'SWS002')
                """
            )
        )

    monkeypatch.setattr(import_preview, "get_engine", lambda: engine)

    preview = import_preview.preview_forecast_store_site_corrections(output.getvalue())

    assert preview["total_rows"] == 3
    assert preview["valid_count"] == 1
    assert preview["error_count"] == 2
    assert preview["valid_rows"][0]["confirmed_store_site"] == "RIVBOS:US"
    assert preview["valid_rows"][0]["listing"] == "RBK004"
    assert preview["error_rows"][0]["message"] == "确认店铺站点为空"
    assert preview["error_rows"][1]["message"] == "确认店铺站点未匹配负责人配置"


def test_commit_forecast_store_site_corrections_upserts_valid_rows(monkeypatch):
    from app.modules.product_ops import import_preview

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["文件", "行号", "站点", "Listing", "月份", "预估销量", "候选店铺站点", "候选负责人配置", "确认店铺站点"])
    sheet.append(["forecast.xlsx", 2, "US", "RBK004", "2026-07", 1200, "", "", "RIVBOS:US"])
    sheet.append(["forecast.xlsx", 3, "US", "SWS002", "2026-08", 800, "", "", "RIVMOUNT:US"])
    output = BytesIO()
    workbook.save(output)

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_forecast (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    listing TEXT,
                    forecast_month TEXT,
                    forecast_units REAL,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (id, store_site, listing)
                VALUES
                    (1, 'RIVBOS:US', 'RBK004'),
                    (2, 'RIVMOUNT:US', 'SWS002')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_forecast (
                    id, store_site, site, listing, forecast_month, forecast_units, updated_at
                )
                VALUES (1, 'RIVBOS:US', 'US', 'RBK004', '2026-07-01', 900, '2026-06-10')
                """
            )
        )

    monkeypatch.setattr(import_preview, "get_engine", lambda: engine)

    result = import_preview.commit_forecast_store_site_corrections(output.getvalue(), changed_by="tester")

    assert result["success"] is True
    assert result["inserted_count"] == 1
    assert result["updated_count"] == 1
    assert result["skipped_count"] == 0
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT store_site, site, listing, forecast_month, forecast_units
                FROM amazon_sales_forecast
                ORDER BY store_site, listing
                """
            )
        ).all()
    assert rows == [
        ("RIVBOS:US", "US", "RBK004", "2026-07-01", 1200),
        ("RIVMOUNT:US", "US", "SWS002", "2026-08-01", 800),
    ]


def test_commit_product_ops_source_import_writes_resolved_rows(tmp_path, monkeypatch):
    from app.modules.product_ops import import_preview

    allocation_dir = tmp_path / "销占比参数"
    forecast_dir = tmp_path / "销售预估参数"
    allocation_dir.mkdir()
    forecast_dir.mkdir()

    allocation_wb = Workbook()
    allocation_ws = allocation_wb.active
    allocation_ws.title = "parameter"
    allocation_ws.append(
        [
            "站点",
            "负责人",
            "Listing",
            "款式",
            "MSKU",
            "积加SKU",
            "店铺",
            "规模定位",
            "款式销占比",
            "SKU销占比",
            "需求定位",
            "发货定位",
            "备货定位",
            "运营保底下单天数",
            "总发货天数",
        ]
    )
    allocation_ws.append(["US", "张三", "RBK004", "粉色款", "RBK004-Pink", "SKU-1", "RIVBOS", "A", 0.6, 0.3, "高", "快", "备货", 30, 95])
    allocation_ws.append(["US", "李四", "SK001", "黑色款", "SK001-Black", "SKU-2", "SEEKWAY", "B", 0.4, 0.2, "中", "慢", "少备", 20, 60])
    allocation_ws.append(["US", "王五", "MISSING", "蓝色款", "MISSING-Blue", "SKU-3", "RIVBOS", "C", 0.2, 0.1, "低", "慢", "不备", 10, 45])
    allocation_wb.save(allocation_dir / "allocation.xlsx")

    forecast_wb = Workbook()
    forecast_ws = forecast_wb.active
    forecast_ws.title = "销量预估表-parameter"
    forecast_ws.append(["店铺/站点", "站点", "Listing", "月份", "Listing_月度预估销量"])
    forecast_ws.append(["RIVBOS:US", "US", "RBK004", "2026-07", 1200])
    forecast_ws.append([None, "US", "SK001", "2026-08", 800])
    forecast_ws.append([None, "US", "SWS002", "2026-09", 600])
    forecast_ws.append([None, "US", "WP001", "2026-10", 500])
    forecast_wb.save(forecast_dir / "forecast.xlsx")

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_allocation (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    owner TEXT,
                    listing TEXT,
                    style TEXT,
                    msku TEXT,
                    sku TEXT,
                    scale_position TEXT,
                    style_sales_ratio REAL,
                    sku_sales_ratio REAL,
                    demand_position TEXT,
                    shipping_position TEXT,
                    stocking_position TEXT,
                    operation_min_order_days INTEGER,
                    total_shipping_days INTEGER
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_forecast (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    listing TEXT,
                    forecast_month TEXT,
                    forecast_units REAL
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (id, store_site, listing)
                VALUES
                    (1, 'RIVBOS:US', 'RBK004'),
                    (2, 'SEEKWAY:US', 'SK001'),
                    (3, 'SIMARI:US', 'SK001'),
                    (4, 'RIVMOUNT:US', 'SWS002'),
                    (5, 'SIMARI:US', 'SWS002'),
                    (6, 'RIVMOUNT:US', 'WP001'),
                    (7, 'SIMARI:US', 'WP001')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_forecast (
                    id, store_site, site, listing, forecast_month, forecast_units
                )
                VALUES (1, 'RIVMOUNT:US', 'US', 'SWS002', '2026-09-01', 500)
                """
            )
        )

    monkeypatch.setattr(import_preview, "get_engine", lambda: engine)

    result = import_preview.commit_product_ops_source_import(tmp_path, changed_by="tester")

    assert result["success"] is True
    assert result["allocation"]["inserted_count"] == 2
    assert result["allocation"]["updated_count"] == 0
    assert result["allocation"]["skipped_count"] == 1
    assert result["forecast"]["inserted_count"] == 2
    assert result["forecast"]["updated_count"] == 1
    assert result["forecast"]["skipped_count"] == 1

    with engine.connect() as conn:
        allocations = conn.execute(
            text(
                """
                SELECT store_site, site, owner, listing, msku, sku, style_sales_ratio, sku_sales_ratio
                FROM amazon_sales_allocation
                ORDER BY store_site, listing, msku
                """
            )
        ).all()
        forecasts = conn.execute(
            text(
                """
                SELECT store_site, site, listing, forecast_month, forecast_units
                FROM amazon_sales_forecast
                ORDER BY store_site, listing, forecast_month
                """
            )
        ).all()

    assert allocations == [
        ("RIVBOS:US", "US", "张三", "RBK004", "RBK004-Pink", "SKU-1", 0.6, 0.3),
        ("SEEKWAY:US", "US", "李四", "SK001", "SK001-Black", "SKU-2", 0.4, 0.2),
    ]
    assert forecasts == [
        ("RIVBOS:US", "US", "RBK004", "2026-07-01", 1200),
        ("RIVMOUNT:US", "US", "SWS002", "2026-09-01", 600),
        ("SEEKWAY:US", "US", "SK001", "2026-08-01", 800),
    ]


def test_preview_sales_forecast_maintenance_import_classifies_rows(monkeypatch):
    from app.modules.product_ops import import_preview

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "销售预估维护"
    sheet.append(["店铺站点", "站点", "Listing", "月份", "Listing月度预估销量"])
    sheet.append(["RIVBOS:US", "US", "RBK004", "2026-07", 1200])
    sheet.append(["RIVBOS:US", "US", "RBK004", "2026-08", 1300])
    sheet.append(["", "US", "RBK004", "2026-09", 100])
    sheet.append(["RIVBOS:US", "US", "RBK004", "2026-10", "abc"])
    sheet.append(["NOOWNER:US", "US", "RBK004", "2026-11", 100])
    output = BytesIO()
    workbook.save(output)

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_forecast (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    listing TEXT,
                    forecast_month TEXT,
                    forecast_units REAL
                )
                """
            )
        )
        conn.execute(text("INSERT INTO amazon_listing_owner_config (id, store_site, listing) VALUES (1, 'RIVBOS:US', 'RBK004')"))
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_forecast (
                    id, store_site, site, listing, forecast_month, forecast_units
                )
                VALUES (1, 'RIVBOS:US', 'US', 'RBK004', '2026-07-01', 900)
                """
            )
        )

    monkeypatch.setattr(import_preview, "get_engine", lambda: engine)

    preview = import_preview.preview_sales_forecast_maintenance_import(output.getvalue())

    assert preview["total_rows"] == 5
    assert preview["valid_count"] == 2
    assert preview["insert_count"] == 1
    assert preview["update_count"] == 1
    assert preview["error_count"] == 3
    assert [row["action"] for row in preview["valid_rows"]] == ["更新", "新增"]
    assert {row["message"] for row in preview["error_rows"]} == {
        "店铺站点、Listing 或月份为空",
        "Listing月度预估销量不是有效数字",
        "店铺站点 + Listing 未匹配负责人配置",
    }


def test_preview_sales_allocation_maintenance_import_classifies_rows(monkeypatch):
    from app.modules.product_ops import import_preview

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "销占比维护"
    sheet.append(["店铺站点", "站点", "负责人", "Listing", "款式", "MSKU", "SKU", "款式销占比", "SKU销占比"])
    sheet.append(["RIVBOS:US", "US", "张三", "RBK004", "粉色款", "RBK004-Pink", "SKU-1", 0.35, 0.2])
    sheet.append(["RIVBOS:US", "US", "张三", "RBK004", "蓝色款", "RBK004-Blue", "SKU-2", 0.25, 0.1])
    sheet.append(["", "US", "张三", "RBK004", "绿色款", "RBK004-Green", "SKU-3", 0.1, 0.1])
    sheet.append(["RIVBOS:US", "US", "张三", "RBK004", "黄色款", "RBK004-Yellow", "SKU-4", "abc", 0.1])
    sheet.append(["NOOWNER:US", "US", "张三", "RBK004", "黑色款", "RBK004-Black", "SKU-5", 0.1, 0.1])
    output = BytesIO()
    workbook.save(output)

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_allocation (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    owner TEXT,
                    listing TEXT,
                    style TEXT,
                    msku TEXT,
                    sku TEXT,
                    style_sales_ratio REAL,
                    sku_sales_ratio REAL
                )
                """
            )
        )
        conn.execute(text("INSERT INTO amazon_listing_owner_config (id, store_site, listing) VALUES (1, 'RIVBOS:US', 'RBK004')"))
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_allocation (
                    id, store_site, site, owner, listing, style, msku, sku, style_sales_ratio, sku_sales_ratio
                )
                VALUES (1, 'RIVBOS:US', 'US', '张三', 'RBK004', '旧粉色', 'RBK004-Pink', 'SKU-1', 0.2, 0.1)
                """
            )
        )

    monkeypatch.setattr(import_preview, "get_engine", lambda: engine)

    preview = import_preview.preview_sales_allocation_maintenance_import(output.getvalue())

    assert preview["total_rows"] == 5
    assert preview["valid_count"] == 2
    assert preview["insert_count"] == 1
    assert preview["update_count"] == 1
    assert preview["error_count"] == 3
    assert [row["action"] for row in preview["valid_rows"]] == ["更新", "新增"]
    assert {row["message"] for row in preview["error_rows"]} == {
        "店铺站点、Listing、MSKU 或 SKU 为空",
        "款式销占比不是有效数字",
        "店铺站点 + Listing 未匹配负责人配置",
    }


def test_commit_sales_allocation_maintenance_import_upserts_and_logs(monkeypatch):
    from app.modules.product_ops import import_preview

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "销占比维护"
    sheet.append(["店铺站点", "站点", "负责人", "Listing", "款式", "MSKU", "SKU", "款式销占比", "SKU销占比", "备货定位", "总发货天数"])
    sheet.append(["RIVBOS:US", "US", "张三", "RBK004", "粉色款", "RBK004-Pink", "SKU-1", 0.35, 0.2, "备货", 95])
    sheet.append(["RIVBOS:US", "US", "张三", "RBK004", "蓝色款", "RBK004-Blue", "SKU-2", 0.25, 0.1, "少备", 60])
    output = BytesIO()
    workbook.save(output)

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_allocation (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    owner TEXT,
                    listing TEXT,
                    style TEXT,
                    msku TEXT,
                    sku TEXT,
                    scale_position TEXT,
                    style_sales_ratio REAL,
                    sku_sales_ratio REAL,
                    demand_position TEXT,
                    shipping_position TEXT,
                    stocking_position TEXT,
                    operation_min_order_days INTEGER,
                    total_shipping_days INTEGER
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_operation_log (
                    id INTEGER PRIMARY KEY,
                    table_name TEXT NOT NULL,
                    record_id INTEGER NOT NULL,
                    operation_type TEXT NOT NULL,
                    changed_by TEXT NOT NULL,
                    change_data TEXT NOT NULL
                )
                """
            )
        )
        conn.execute(text("INSERT INTO amazon_listing_owner_config (id, store_site, listing) VALUES (1, 'RIVBOS:US', 'RBK004')"))
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_allocation (
                    id, store_site, site, owner, listing, style, msku, sku, style_sales_ratio, sku_sales_ratio,
                    stocking_position, total_shipping_days
                )
                VALUES (1, 'RIVBOS:US', 'US', '张三', 'RBK004', '旧粉色', 'RBK004-Pink', 'SKU-1', 0.2, 0.1, '少备', 45)
                """
            )
        )

    monkeypatch.setattr(import_preview, "get_engine", lambda: engine)

    result = import_preview.commit_sales_allocation_maintenance_import(output.getvalue(), changed_by="tester")

    assert result["success"] is True
    assert result["inserted_count"] == 1
    assert result["updated_count"] == 1
    assert result["skipped_count"] == 0
    with engine.connect() as conn:
        allocations = conn.execute(
            text(
                """
                SELECT store_site, site, owner, listing, style, msku, sku, style_sales_ratio, sku_sales_ratio,
                       stocking_position, total_shipping_days
                FROM amazon_sales_allocation
                ORDER BY msku
                """
            )
        ).all()
        logs = conn.execute(
            text(
                """
                SELECT table_name, operation_type, changed_by
                FROM amazon_operation_log
                ORDER BY id
                """
            )
        ).all()

    assert allocations == [
        ("RIVBOS:US", "US", "张三", "RBK004", "蓝色款", "RBK004-Blue", "SKU-2", 0.25, 0.1, "少备", 60),
        ("RIVBOS:US", "US", "张三", "RBK004", "粉色款", "RBK004-Pink", "SKU-1", 0.35, 0.2, "备货", 95),
    ]
    assert logs == [
        ("amazon_sales_allocation", "IMPORT_UPDATE", "tester"),
        ("amazon_sales_allocation", "IMPORT_INSERT", "tester"),
    ]


def test_commit_sales_forecast_maintenance_import_upserts_and_logs(monkeypatch):
    from app.modules.product_ops import import_preview

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "销售预估维护"
    sheet.append(["店铺站点", "站点", "Listing", "月份", "Listing月度预估销量"])
    sheet.append(["RIVBOS:US", "US", "RBK004", "2026-07", 1200])
    sheet.append(["RIVBOS:US", "US", "RBK004", "2026-08", 1300])
    output = BytesIO()
    workbook.save(output)

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_forecast (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    listing TEXT,
                    forecast_month TEXT,
                    forecast_units REAL
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_operation_log (
                    id INTEGER PRIMARY KEY,
                    table_name TEXT NOT NULL,
                    record_id INTEGER NOT NULL,
                    operation_type TEXT NOT NULL,
                    changed_by TEXT NOT NULL,
                    change_data TEXT NOT NULL
                )
                """
            )
        )
        conn.execute(text("INSERT INTO amazon_listing_owner_config (id, store_site, listing) VALUES (1, 'RIVBOS:US', 'RBK004')"))
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_forecast (
                    id, store_site, site, listing, forecast_month, forecast_units
                )
                VALUES (1, 'RIVBOS:US', 'US', 'RBK004', '2026-07-01', 900)
                """
            )
        )

    monkeypatch.setattr(import_preview, "get_engine", lambda: engine)

    result = import_preview.commit_sales_forecast_maintenance_import(output.getvalue(), changed_by="tester")

    assert result["success"] is True
    assert result["inserted_count"] == 1
    assert result["updated_count"] == 1
    assert result["skipped_count"] == 0
    with engine.connect() as conn:
        forecasts = conn.execute(
            text(
                """
                SELECT store_site, site, listing, forecast_month, forecast_units
                FROM amazon_sales_forecast
                ORDER BY forecast_month
                """
            )
        ).all()
        logs = conn.execute(
            text(
                """
                SELECT table_name, operation_type, changed_by
                FROM amazon_operation_log
                ORDER BY id
                """
            )
        ).all()

    assert forecasts == [
        ("RIVBOS:US", "US", "RBK004", "2026-07-01", 1200),
        ("RIVBOS:US", "US", "RBK004", "2026-08-01", 1300),
    ]
    assert logs == [
        ("amazon_sales_forecast", "IMPORT_UPDATE", "tester"),
        ("amazon_sales_forecast", "IMPORT_INSERT", "tester"),
    ]


def test_maintenance_import_rows_are_visible_on_management_pages(monkeypatch):
    from app.modules.product_ops import import_preview, service

    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    listing_status TEXT,
                    listing_maintainer TEXT,
                    project_group TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_allocation (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    owner TEXT,
                    listing TEXT,
                    style TEXT,
                    msku TEXT,
                    sku TEXT,
                    scale_position TEXT,
                    style_sales_ratio REAL,
                    sku_sales_ratio REAL,
                    demand_position TEXT,
                    shipping_position TEXT,
                    stocking_position TEXT,
                    operation_min_order_days INTEGER,
                    total_shipping_days INTEGER,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_forecast (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    listing TEXT,
                    forecast_month TEXT,
                    forecast_units REAL,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_operation_log (
                    id INTEGER PRIMARY KEY,
                    table_name TEXT NOT NULL,
                    record_id INTEGER NOT NULL,
                    operation_type TEXT NOT NULL,
                    changed_by TEXT NOT NULL,
                    change_data TEXT NOT NULL
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (
                    id, store_site, listing, owner, listing_status, listing_maintainer, project_group
                )
                VALUES (1, 'RIVBOS:US', 'RBK004', '张三', '维护', '李四', '项目组A')
                """
            )
        )

    monkeypatch.setattr(import_preview, "get_engine", lambda: engine)
    monkeypatch.setattr(service, "get_engine", lambda: engine)

    allocation_workbook = Workbook()
    allocation_sheet = allocation_workbook.active
    allocation_sheet.title = "销占比维护"
    allocation_sheet.append(["店铺站点", "站点", "负责人", "Listing", "款式", "MSKU", "SKU", "款式销占比", "SKU销占比", "备货定位", "总发货天数"])
    allocation_sheet.append(["RIVBOS:US", "US", "张三", "RBK004", "粉色款", "RBK004-Pink", "SKU-1", 0.35, 0.2, "备货", 95])
    allocation_output = BytesIO()
    allocation_workbook.save(allocation_output)

    forecast_workbook = Workbook()
    forecast_sheet = forecast_workbook.active
    forecast_sheet.title = "销售预估维护"
    forecast_sheet.append(["店铺站点", "站点", "Listing", "月份", "Listing月度预估销量"])
    forecast_sheet.append(["RIVBOS:US", "US", "RBK004", "2026-07", 1200])
    forecast_output = BytesIO()
    forecast_workbook.save(forecast_output)

    allocation_result = import_preview.commit_sales_allocation_maintenance_import(
        allocation_output.getvalue(),
        changed_by="tester",
    )
    forecast_result = import_preview.commit_sales_forecast_maintenance_import(
        forecast_output.getvalue(),
        changed_by="tester",
    )

    assert allocation_result["success"] is True
    assert forecast_result["success"] is True

    allocation_page = client.get("/product-ops/allocations", params={"store_site": "RIVBOS:US", "listing": "RBK004"})
    forecast_page = client.get("/product-ops/forecasts", params={"store_site": "RIVBOS:US", "listing": "RBK004"})
    profile_page = client.get("/product-ops/listing-profile", params={"store_site": "RIVBOS:US", "listing": "RBK004"})

    assert allocation_page.status_code == 200
    assert "RBK004-Pink" in allocation_page.text
    assert "35.00%" in allocation_page.text
    assert "20.00%" in allocation_page.text
    assert forecast_page.status_code == 200
    assert "2026-07-01" in forecast_page.text
    assert "1200" in forecast_page.text
    assert profile_page.status_code == 200
    assert "张三" in profile_page.text
    assert "RBK004-Pink" in profile_page.text
    assert "35.00%" in profile_page.text
    assert "2026-07-01" in profile_page.text


def test_forecast_store_site_correction_preview_route_renders_result(monkeypatch):
    monkeypatch.setattr(
        "app.modules.product_ops.routes.preview_forecast_store_site_corrections",
        lambda content: {
            "total_rows": 2,
            "valid_count": 1,
            "error_count": 1,
            "valid_rows": [
                {
                    "row_number": 2,
                    "confirmed_store_site": "RIVMOUNT:CA",
                    "listing": "SWS002",
                    "forecast_month": "2025-05",
                    "forecast_units": 108,
                }
            ],
            "error_rows": [{"row_number": 3, "message": "确认店铺站点为空"}],
        },
    )

    response = client.post(
        "/product-ops/import-preview/forecast-store-site-review/preview",
        files={"file": ("review.xlsx", b"xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    assert "修正预览结果" in response.text
    assert "可导入 1 行" in response.text
    assert "错误行 1 行" in response.text
    assert "RIVMOUNT:CA" in response.text
    assert "确认店铺站点为空" in response.text


def test_forecast_store_site_correction_preview_route_exposes_commit_when_valid(monkeypatch):
    monkeypatch.setattr(
        "app.modules.product_ops.routes.preview_forecast_store_site_corrections",
        lambda content: {
            "total_rows": 1,
            "valid_count": 1,
            "error_count": 0,
            "valid_rows": [
                {
                    "row_number": 2,
                    "confirmed_store_site": "RIVMOUNT:CA",
                    "listing": "SWS002",
                    "forecast_month": "2025-05",
                    "forecast_units": 108,
                }
            ],
            "error_rows": [],
        },
    )
    monkeypatch.setattr("app.modules.product_ops.routes.save_import_upload", lambda content: "token123")

    response = client.post(
        "/product-ops/import-preview/forecast-store-site-review/preview",
        files={"file": ("review.xlsx", b"xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    assert "确认写入销售预估" in response.text
    assert 'name="import_token" value="token123"' in response.text


def test_forecast_store_site_correction_commit_route_writes_uploaded_file(monkeypatch):
    captured = {}

    monkeypatch.setattr("app.modules.product_ops.routes.load_import_upload", lambda token: b"xlsx")

    def fake_commit(content, changed_by):
        captured["content"] = content
        captured["changed_by"] = changed_by
        return {
            "success": True,
            "inserted_count": 1,
            "updated_count": 2,
            "skipped_count": 0,
            "message": "写入完成",
            "preview": {"total_rows": 3, "valid_count": 3, "error_count": 0, "valid_rows": [], "error_rows": []},
        }

    monkeypatch.setattr("app.modules.product_ops.routes.commit_forecast_store_site_corrections", fake_commit)

    response = client.post(
        "/product-ops/import-preview/forecast-store-site-review/commit",
        data={"import_token": "token123"},
    )

    assert response.status_code == 200
    assert "写入完成" in response.text
    assert "新增 1 行，更新 2 行" in response.text
    assert captured == {"content": b"xlsx", "changed_by": "test-admin"}


def test_product_ops_source_import_commit_route_renders_result(monkeypatch):
    captured = {}

    monkeypatch.setattr(
        "app.modules.product_ops.routes.build_product_ops_import_preview",
        lambda: {
            "base_dir": "待合并/src_data",
            "allocation": {
                "label": "销占比",
                "file_count": 1,
                "total_rows": 3,
                "duplicate_key_count": 0,
                "missing_key_count": 0,
                "owner_unmatched_count": 1,
                "owner_ambiguous_count": 0,
                "owner_resolved_by_allocation_count": 0,
                "owner_unresolved_ambiguous_count": 0,
                "issue_rows": [],
            },
            "forecast": {
                "label": "销售预估",
                "file_count": 1,
                "total_rows": 4,
                "duplicate_key_count": 0,
                "missing_key_count": 1,
                "store_site_missing_count": 3,
                "store_site_explicit_count": 1,
                "store_site_inferred_count": 1,
                "owner_unmatched_count": 0,
                "owner_ambiguous_count": 2,
                "owner_resolved_by_allocation_count": 1,
                "owner_unresolved_ambiguous_count": 1,
                "issue_rows": [],
            },
        },
    )

    def fake_commit(changed_by):
        captured["changed_by"] = changed_by
        return {
            "success": True,
            "message": "正式导入完成",
            "allocation": {"total_rows": 3, "inserted_count": 2, "updated_count": 0, "skipped_count": 1},
            "forecast": {"total_rows": 4, "inserted_count": 2, "updated_count": 1, "skipped_count": 1},
        }

    monkeypatch.setattr("app.modules.product_ops.routes.commit_product_ops_source_import", fake_commit)

    response = client.post("/product-ops/import-preview/commit")

    assert response.status_code == 200
    assert "正式导入完成" in response.text
    assert "销占比：新增 2 行，更新 0 行，跳过 1 行" in response.text
    assert "销售预估：新增 2 行，更新 1 行，跳过 1 行" in response.text
    assert captured == {"changed_by": "test-admin"}


def test_list_product_ops_rows_uses_listing_owner_as_overview_base(monkeypatch):
    from app.modules.product_ops import service

    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    msku TEXT,
                    sku TEXT,
                    brand TEXT,
                    category_a TEXT,
                    category_b TEXT,
                    listing TEXT,
                    sales_status TEXT,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    listing_status TEXT,
                    listing_maintainer TEXT,
                    include_inventory_age_assessment TEXT,
                    project_group TEXT,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_allocation (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    owner TEXT,
                    listing TEXT,
                    style TEXT,
                    msku TEXT,
                    sku TEXT,
                    style_sales_ratio REAL,
                    sku_sales_ratio REAL,
                    stocking_position TEXT,
                    total_shipping_days INTEGER,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_sales_forecast (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    site TEXT,
                    listing TEXT,
                    forecast_month TEXT,
                    forecast_units REAL,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, store_site, msku, sku, brand, category_a, category_b, listing, sales_status, updated_at
                )
                VALUES (7, 'RIVBOS:US', 'RBK004-2 W Pink', 'RBK004-2', 'RIVBOS', '手套', '冬季手套', 'RBK004', '在售', '2026-06-10')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, store_site, msku, sku, brand, category_a, category_b, listing, sales_status, updated_at
                )
                VALUES (8, 'RIVBOS:US', 'RBK004-2 W Blue', 'RBK004-2B', 'RIVBOS', '手套', '冬季手套', 'RBK004', '在售', '2026-06-10')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (
                    id, store_site, listing, owner, listing_status, listing_maintainer,
                    include_inventory_age_assessment, project_group, updated_at
                )
                VALUES (1, 'RIVBOS:US', 'RBK004', '张三', '正常', '李四', '是', '项目组A', '2026-06-10')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (
                    id, store_site, listing, owner, listing_status, listing_maintainer,
                    include_inventory_age_assessment, project_group, updated_at
                )
                VALUES (2, 'SEEKWAY:US', 'SK001', '李四', '维护', '王五', '是', '项目组B', '2026-06-10')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_allocation (
                    id, store_site, site, owner, listing, style, msku, sku,
                    style_sales_ratio, sku_sales_ratio, stocking_position, total_shipping_days, updated_at
                )
                VALUES (
                    1, 'RIVBOS:US', 'US', '张三', 'RBK004', '粉色款',
                    'RBK004-2 W Pink', 'RBK004-2', 0.35, 0.2, '备货', 95, '2026-06-10'
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_allocation (
                    id, store_site, site, owner, listing, style, msku, sku,
                    style_sales_ratio, sku_sales_ratio, stocking_position, total_shipping_days, updated_at
                )
                VALUES (
                    2, 'RIVBOS:US', 'US', '张三', 'RBK004', '蓝色款',
                    'RBK004-2 W Blue', 'RBK004-2B', 0.25, 0.1, '备货', 95, '2026-06-10'
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_sales_forecast (
                    id, store_site, site, listing, forecast_month, forecast_units, updated_at
                )
                VALUES
                    (1, 'RIVBOS:US', 'US', 'RBK004', '2026-07', 1200, '2026-06-10'),
                    (2, 'RIVBOS:US', 'US', 'RBK004', '2026-08', 1300, '2026-06-10'),
                    (3, 'OTHER:US', 'US', 'RBK004', '2026-07', 999, '2026-06-10')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    page = service.list_product_ops_rows(service.ProductOpsFilters(q="RBK004"))

    assert page["total"] == 1
    row = page["rows"][0]
    assert row["store_site"] == "RIVBOS:US"
    assert row["owner"] == "张三"
    assert row["listing_status"] == "正常"
    assert row["project_group"] == "项目组A"
    assert row["allocation_msku_count"] == 2
    assert row["product_msku_count"] == 2
    assert row["forecast_month_count"] == 2
    assert row["forecast_units_total"] == 2500
    assert row["data_status"] == "正常"
    assert row["purchase_readiness_status"] == "ready"
    assert row["purchase_readiness_label"] == "可进入采购判断"

    missing_page = service.list_product_ops_rows(service.ProductOpsFilters(q="SK001"))

    assert missing_page["total"] == 1
    missing_row = missing_page["rows"][0]
    assert missing_row["store_site"] == "SEEKWAY:US"
    assert missing_row["allocation_msku_count"] == 0
    assert missing_row["product_msku_count"] == 0
    assert missing_row["forecast_month_count"] == 0
    assert missing_row["data_status"] == "缺销占比/缺销售预估/缺产品信息"
    assert missing_row["purchase_readiness_status"] == "blocked"
    assert missing_row["purchase_readiness_label"] == "需补数据"
    assert "缺产品信息" in missing_row["purchase_readiness_reasons"]


def test_build_purchase_readiness_classifies_ready_blocked_and_review():
    from app.modules.product_ops.readiness import build_purchase_readiness

    ready = build_purchase_readiness(
        {
            "owner_config_id": 1,
            "product_msku_count": 2,
            "allocation_msku_count": 2,
            "forecast_month_count": 2,
            "zero_allocation_ratio_count": 0,
            "zero_forecast_units_count": 0,
        }
    )
    assert ready["status"] == "ready"
    assert ready["label"] == "可进入采购判断"

    blocked = build_purchase_readiness(
        {
            "owner_config_id": 1,
            "product_msku_count": 0,
            "allocation_msku_count": 0,
            "forecast_month_count": 0,
        }
    )
    assert blocked["status"] == "blocked"
    assert blocked["label"] == "需补数据"
    assert blocked["reasons"] == ["缺产品信息", "缺销占比", "缺销售预估"]

    review = build_purchase_readiness(
        {
            "owner_config_id": 1,
            "product_msku_count": 2,
            "allocation_msku_count": 2,
            "forecast_month_count": 2,
        },
        allocation_rows=[{"style_sales_ratio": 0, "sku_sales_ratio": 0.2}],
        forecast_rows=[{"forecast_units": 0}],
    )
    assert review["status"] == "review"
    assert review["label"] == "需人工确认"
    assert review["reasons"] == ["销占比为0", "销售预估为0"]
