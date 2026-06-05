from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, text

from app.main import app
from app.modules.product_import import service
from app.modules.product_import.service import commit_product_import, preview_product_import


client = TestClient(app)


PRODUCT_IMPORT_HEADERS = [
    "店铺/站点",
    "MSKU",
    "ASIN",
    "父ASIN",
    "产品名称",
    "SKU",
    "品牌",
    "FNSKU",
    "销售状态",
    "仓储类型",
    "一级品类",
    "品类A",
    "品类B",
    "Listing",
    "标签名",
    "MSKU发货备注",
    "借调备注",
    "锁仓MSKU",
]


def build_workbook_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_preview_product_import_validates_rows_without_writing(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    msku TEXT,
                    asin TEXT,
                    parent_asin TEXT,
                    product_name TEXT,
                    sku TEXT,
                    brand TEXT,
                    fnsku TEXT,
                    msku_lock_status TEXT,
                    listing TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, store_site, msku, asin, parent_asin, product_name, sku, brand, fnsku, listing
                )
                VALUES (
                    1, 'SAYOLA:US', 'MSKU-001', 'OLDASIN001', 'OLDPARENT1',
                    'Old Product', 'OLD-SKU', 'Old Brand', 'OLD-FNSKU', 'OLD-LISTING'
                )
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)
    content = build_workbook_bytes(
        [
            ["店铺/站点", "MSKU", "ASIN", "父ASIN", "产品名称", "SKU", "品牌", "FNSKU", "Listing"],
            [
                "SAYOLA:US",
                "MSKU-001",
                "NEWASIN001",
                "NEWPARENT1",
                "New Product",
                "NEW-SKU",
                "New Brand",
                "NEW-FNSKU",
                "NEW-LISTING",
            ],
            ["SAYOLA:US", "MSKU-404", "B002", "", "Missing Product", "SKU-404", "BrandB", "", ""],
            ["", "MSKU-EMPTY", "B003", "", "No Store", "SKU-EMPTY", "BrandC", "", ""],
        ]
    )

    preview = preview_product_import(content)

    assert preview["total_rows"] == 3
    assert preview["valid_count"] == 1
    assert preview["missing_product_count"] == 1
    assert preview["error_count"] == 1
    assert preview["blocked_fields"] == []
    assert preview["valid_rows"][0]["row_number"] == 2
    assert preview["valid_rows"][0]["store_site"] == "SAYOLA:US"
    assert preview["valid_rows"][0]["msku"] == "MSKU-001"
    assert preview["valid_rows"][0]["changes"] == {
        "asin": "NEWASIN001",
        "parent_asin": "NEWPARENT1",
        "product_name": "New Product",
        "sku": "NEW-SKU",
        "brand": "New Brand",
        "fnsku": "NEW-FNSKU",
        "listing": "NEW-LISTING",
    }
    assert preview["valid_rows"][0]["change_items"] == [
        {"field": "asin", "old": "OLDASIN001", "new": "NEWASIN001"},
        {"field": "parent_asin", "old": "OLDPARENT1", "new": "NEWPARENT1"},
        {"field": "product_name", "old": "Old Product", "new": "New Product"},
        {"field": "sku", "old": "OLD-SKU", "new": "NEW-SKU"},
        {"field": "brand", "old": "Old Brand", "new": "New Brand"},
        {"field": "fnsku", "old": "OLD-FNSKU", "new": "NEW-FNSKU"},
        {"field": "listing", "old": "OLD-LISTING", "new": "NEW-LISTING"},
    ]
    assert preview["missing_product_rows"][0]["row_number"] == 3
    assert preview["error_rows"][0]["message"] == "缺少店铺站点或 MSKU"


def test_build_product_import_issue_workbook_exports_errors_and_missing_rows():
    preview = {
        "missing_product_rows": [
            {"row_number": 3, "store_site": "SAYOLA:US", "msku": "MSKU-404"}
        ],
        "error_rows": [
            {"row_number": 4, "message": "缺少店铺站点或 MSKU"}
        ],
    }

    content = service.build_product_import_issue_workbook(preview)
    workbook = load_workbook(BytesIO(content), read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))

    assert rows[0] == ("类型", "行号", "店铺站点", "MSKU", "错误说明")
    assert rows[1] == ("未匹配产品", 3, "SAYOLA:US", "MSKU-404", "产品不存在")
    assert rows[2] == ("错误行", 4, None, None, "缺少店铺站点或 MSKU")


def test_commit_product_import_updates_existing_rows_and_writes_logs(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    msku TEXT,
                    sku TEXT,
                    msku_lock_status TEXT,
                    asin TEXT,
                    product_name TEXT,
                    brand TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_operation_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    table_name TEXT NOT NULL,
                    record_id INTEGER NOT NULL,
                    operation_type TEXT NOT NULL,
                    changed_by TEXT NOT NULL,
                    change_data TEXT NOT NULL,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (id, store_site, msku, asin, product_name, brand)
                VALUES (1, 'SAYOLA:US', 'MSKU-001', 'OLDASIN001', 'Old Product', 'Old Brand')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)
    cache_clears = {"list": 0, "filters": 0}
    monkeypatch.setattr(
        service,
        "clear_product_list_cache",
        lambda: cache_clears.__setitem__("list", cache_clears["list"] + 1),
    )
    monkeypatch.setattr(
        service,
        "clear_product_filter_options_cache",
        lambda: cache_clears.__setitem__("filters", cache_clears["filters"] + 1),
    )
    content = build_workbook_bytes(
        [
            ["店铺/站点", "MSKU", "ASIN", "产品名称", "品牌"],
            ["SAYOLA:US", "MSKU-001", "NEWASIN001", "New Product", "Old Brand"],
        ]
    )

    result = commit_product_import(content, changed_by="tester")

    assert result["success"] is True
    assert result["updated_count"] == 1
    assert cache_clears == {"list": 1, "filters": 1}
    with engine.connect() as conn:
        product = conn.execute(
            text("SELECT asin, product_name, brand FROM amazon_product_info WHERE id = 1")
        ).mappings().one()
        log = conn.execute(text("SELECT * FROM amazon_operation_log")).mappings().one()

    assert dict(product) == {
        "asin": "NEWASIN001",
        "product_name": "New Product",
        "brand": "Old Brand",
    }
    assert log["table_name"] == "amazon_product_info"
    assert log["record_id"] == 1
    assert log["operation_type"] == "IMPORT_UPDATE"
    assert log["changed_by"] == "tester"
    assert "NEWASIN001" in log["change_data"]


def test_commit_product_import_rejects_invalid_preview_without_writing(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    msku TEXT,
                    brand TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (id, store_site, msku, brand)
                VALUES (1, 'SAYOLA:US', 'MSKU-001', 'Old Brand')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)
    content = build_workbook_bytes(
        [
            ["店铺/站点", "MSKU", "品牌"],
            ["SAYOLA:US", "MSKU-404", "New Brand"],
        ]
    )

    result = commit_product_import(content)

    assert result["success"] is False
    assert result["updated_count"] == 0
    with engine.connect() as conn:
        brand = conn.execute(
            text("SELECT brand FROM amazon_product_info WHERE id = 1")
        ).scalar_one()
    assert brand == "Old Brand"


def test_preview_product_import_rejects_locked_conflict(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    msku TEXT,
                    sku TEXT,
                    msku_lock_status TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (id, store_site, msku, sku, msku_lock_status)
                VALUES
                    (1, 'SAYOLA:US', 'MSKU-001', 'SKU-001', '锁'),
                    (2, 'SAYOLA:US', 'MSKU-002', 'SKU-001', '否')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)
    content = build_workbook_bytes(
        [
            ["店铺/站点", "MSKU", "锁仓MSKU"],
            ["SAYOLA:US", "MSKU-002", "锁"],
        ]
    )

    preview = preview_product_import(content)

    assert preview["valid_count"] == 0
    assert preview["error_count"] == 1
    assert preview["error_rows"][0]["message"] == "同一店铺站点 + SKU 下最多只能有一个锁仓 MSKU 为“锁”。"


def test_product_import_page_renders_upload_form():
    response = client.get("/products/import")

    assert response.status_code == 200
    assert 'class="import-workbench"' in response.text
    assert "1 下载模板" in response.text
    assert "2 上传校验" in response.text
    assert "3 确认写入" in response.text
    assert "上传 Excel" in response.text
    assert "name=\"file\"" in response.text
    assert "/products/import/template" in response.text


def test_product_import_template_downloads_xlsx():
    response = client.get("/products/import/template")

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content), read_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    assert headers == PRODUCT_IMPORT_HEADERS


def test_product_import_preview_renders_result(monkeypatch):
    monkeypatch.setattr(
        "app.modules.product_import.routes.preview_product_import",
        lambda content: {
            "total_rows": 1,
            "valid_count": 1,
            "missing_product_count": 0,
            "error_count": 0,
            "blocked_fields": [],
            "valid_rows": [
                {
                    "row_number": 2,
                    "store_site": "SAYOLA:US",
                    "msku": "MSKU-001",
                    "changes": {"brand": "New Brand"},
                    "change_items": [{"field": "brand", "old": "Old Brand", "new": "New Brand"}],
                }
            ],
            "missing_product_rows": [],
            "error_rows": [],
        },
    )
    monkeypatch.setattr("app.modules.product_import.routes.save_import_upload", lambda content: "token-1")
    content = build_workbook_bytes([["店铺站点", "MSKU", "品牌"], ["SAYOLA:US", "MSKU-001", "New Brand"]])

    response = client.post(
        "/products/import/preview",
        files={
            "file": (
                "products.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert 'class="import-preview-panel"' in response.text
    assert "可更新 1 行" in response.text
    assert "校验结果" in response.text
    assert "校验通过，可确认写入数据库。" in response.text
    assert "MSKU-001" in response.text
    assert "Old Brand" in response.text
    assert "New Brand" in response.text
    assert "name=\"import_token\"" in response.text
    assert "token-1" in response.text


def test_product_import_preview_links_issue_download_when_preview_has_errors(monkeypatch):
    monkeypatch.setattr(
        "app.modules.product_import.routes.preview_product_import",
        lambda content: {
            "total_rows": 1,
            "valid_count": 0,
            "missing_product_count": 0,
            "error_count": 1,
            "blocked_fields": [],
            "valid_rows": [],
            "missing_product_rows": [],
            "error_rows": [{"row_number": 2, "message": "缺少店铺站点或 MSKU"}],
        },
    )
    monkeypatch.setattr("app.modules.product_import.routes.save_import_upload", lambda content: "token-1")
    content = build_workbook_bytes([["店铺站点", "MSKU"], ["", "MSKU-001"]])

    response = client.post(
        "/products/import/preview",
        files={
            "file": (
                "products.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert "/products/import/issues?import_token=token-1" in response.text


def test_product_import_issue_download_route_returns_workbook(monkeypatch):
    monkeypatch.setattr("app.modules.product_import.routes.load_import_upload", lambda token: b"xlsx")
    monkeypatch.setattr(
        "app.modules.product_import.routes.preview_product_import",
        lambda content: {
            "missing_product_rows": [],
            "error_rows": [{"row_number": 2, "message": "缺少店铺站点或 MSKU"}],
        },
    )

    response = client.get("/products/import/issues?import_token=token-1")

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]


def test_product_import_commit_route_renders_result(monkeypatch):
    captured = {}
    monkeypatch.setattr("app.modules.product_import.routes.load_import_upload", lambda token: b"xlsx")

    def fake_commit_product_import(content, changed_by="system"):
        captured["content"] = content
        captured["changed_by"] = changed_by
        return {"success": True, "updated_count": 2, "skipped_count": 1, "message": "写入完成"}

    monkeypatch.setattr("app.modules.product_import.routes.commit_product_import", fake_commit_product_import)

    response = client.post("/products/import/commit", data={"import_token": "token-1"})

    assert response.status_code == 200
    assert "写入完成" in response.text
    assert "更新 2 行" in response.text
    assert captured == {"content": b"xlsx", "changed_by": "test-admin"}
