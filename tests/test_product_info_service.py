import json
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from app.modules.product_info import service
from app.modules.product_info.service import (
    LockConflictError,
    ProductFilters,
    create_product,
    _build_where,
    export_products_to_xlsx,
    get_product_detail,
    list_products,
    list_products_for_export,
    normalize_filters,
    update_product,
)
from app.modules.product_import.service import IMPORT_TEMPLATE_HEADERS, preview_product_import
from app.modules.product_import import service as import_service
from app.modules.store_site.service import UnknownStoreSiteError


def test_normalize_filters_trims_values_and_forces_page_size():
    filters = normalize_filters(
        ProductFilters(
            q=" abc ",
            store_site=" SAYOLA:US ",
            brand=" ",
            sales_status="在售",
            listing="ListingA",
            listing_owner=" OwnerA ",
            listing_owner_status=" Active ",
            project_group=" GroupA ",
            page=0,
            page_size=999,
        )
    )

    assert filters.q == "abc"
    assert filters.store_site == "SAYOLA:US"
    assert filters.brand is None
    assert filters.sales_status == "在售"
    assert filters.listing == "ListingA"
    assert filters.listing_owner == "OwnerA"
    assert filters.listing_owner_status == "Active"
    assert filters.project_group == "GroupA"
    assert filters.page == 1
    assert filters.page_size == 20


def test_build_where_supports_search_and_exact_filters():
    where_sql, params = _build_where(
        ProductFilters(
            q="abc",
            store_site="SAYOLA:US",
            brand="BrandA",
            sales_status="在售",
            listing="ListingA",
            listing_owner="OwnerA",
            listing_owner_status="Active",
            project_group="GroupA",
        )
    )

    assert "msku LIKE :q" in where_sql
    assert "asin LIKE :q" in where_sql
    assert "product_name LIKE :q" in where_sql
    assert "store_site = :store_site" in where_sql
    assert "brand = :brand" in where_sql
    assert "p.sales_status = :sales_status" in where_sql
    assert "p.listing = :listing" in where_sql
    assert "lo.owner = :listing_owner" in where_sql
    assert "lo.listing_status = :listing_owner_status" in where_sql
    assert "lo.project_group = :project_group" in where_sql
    assert params == {
        "q": "%abc%",
        "store_site": "SAYOLA:US",
        "brand": "BrandA",
        "sales_status": "在售",
        "listing": "ListingA",
        "listing_owner": "OwnerA",
        "listing_owner_status": "Active",
        "project_group": "GroupA",
    }


def test_list_products_returns_lightweight_table_fields(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    asin TEXT,
                    msku TEXT,
                    store_site TEXT,
                    parent_asin TEXT,
                    product_name TEXT,
                    sku TEXT,
                    brand TEXT,
                    fnsku TEXT,
                    sales_status TEXT,
                    storage_type TEXT,
                    category_level_1 TEXT,
                    category_a TEXT,
                    category_b TEXT,
                    listing TEXT,
                    label_name TEXT,
                    msku_shipping_remark TEXT,
                    transfer_remark TEXT,
                    msku_lock_status TEXT,
                    created_at TEXT,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    listing_status TEXT,
                    listing_maintainer TEXT,
                    include_inventory_age_assessment TEXT,
                    project_group TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, asin, msku, store_site, parent_asin, product_name, sku, brand,
                    fnsku, sales_status, storage_type, category_level_1, category_a,
                    category_b, listing, label_name, msku_shipping_remark,
                    transfer_remark, msku_lock_status, created_at, updated_at
                )
                VALUES (
                    1, 'B001', 'MSKU-001', 'SAYOLA:US', 'PARENT001', 'Product A',
                    'SKU-001', 'BrandA', 'FNSKU-001', '在售', 'FBA', '服饰',
                    '眼镜', '太阳镜', 'RB833', '标签', '发货备注',
                    '借调备注', '否', '2026-06-01', '2026-06-02'
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (
                    store_site, listing, owner, listing_status, listing_maintainer,
                    include_inventory_age_assessment, project_group
                )
                VALUES (
                    'SAYOLA:US', 'RB833', 'OwnerA', 'Active', 'MaintainerA', '是', 'GroupA'
                )
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    rows = list_products(ProductFilters())["rows"]

    assert "parent_asin" not in rows[0]
    assert "fnsku" not in rows[0]
    assert "label_name" not in rows[0]
    assert "msku_shipping_remark" not in rows[0]
    assert "created_at" not in rows[0]
    assert rows[0]["msku"] == "MSKU-001"
    assert rows[0]["product_name"] == "Product A"
    assert rows[0]["listing_owner"] == "OwnerA"
    assert rows[0]["listing_owner_status"] == "Active"
    assert "listing_maintainer" not in rows[0]
    assert "include_inventory_age_assessment" not in rows[0]
    assert rows[0]["project_group"] == "GroupA"


def test_get_product_detail_returns_related_store_and_listing_owner(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    asin TEXT,
                    msku TEXT,
                    store_site TEXT,
                    parent_asin TEXT,
                    product_name TEXT,
                    sku TEXT,
                    brand TEXT,
                    fnsku TEXT,
                    sales_status TEXT,
                    storage_type TEXT,
                    category_level_1 TEXT,
                    category_a TEXT,
                    category_b TEXT,
                    listing TEXT,
                    label_name TEXT,
                    msku_shipping_remark TEXT,
                    transfer_remark TEXT,
                    msku_lock_status TEXT,
                    created_at TEXT,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_store_site (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    store TEXT,
                    country TEXT,
                    domain TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    listing_status TEXT,
                    listing_maintainer TEXT,
                    include_inventory_age_assessment TEXT,
                    project_group TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, asin, msku, store_site, parent_asin, product_name, sku, brand, fnsku,
                    sales_status, storage_type, category_level_1, category_a, category_b, listing,
                    label_name, msku_shipping_remark, transfer_remark, msku_lock_status, created_at, updated_at
                )
                VALUES (
                    1, 'B012345678', 'MSKU-001', 'SAYOLA:US', 'B0PARENT', 'Product 1', 'SKU-001',
                    'BrandA', 'FNSKU-001', '在售', 'FBA', '服饰', '眼镜', '太阳镜', 'ListingA',
                    '核心款', '发货备注', '借调备注', '否', '2026-06-01', '2026-06-02'
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_store_site (id, store_site, store, country, domain)
                VALUES (2, 'SAYOLA:US', 'SAYOLA', 'US', 'amazon.com')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (
                    id, store_site, listing, owner, listing_status, listing_maintainer,
                    include_inventory_age_assessment, project_group
                )
                VALUES (3, 'SAYOLA:US', 'ListingA', '张三', '正常', '李四', '是', '项目组A')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    detail = get_product_detail(1)

    assert detail is not None
    assert detail["store_site"]["domain"] == "amazon.com"
    assert detail["owner"]["id"] == 3
    assert detail["owner"]["owner"] == "张三"


def test_bulk_update_product_lock_status_updates_rows_and_logs(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    msku TEXT,
                    sku TEXT,
                    msku_lock_status TEXT,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_operation_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    table_name TEXT NOT NULL,
                    record_id INTEGER NOT NULL,
                    operation_type TEXT NOT NULL,
                    changed_by TEXT NOT NULL,
                    change_data TEXT NOT NULL,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (id, store_site, msku, sku, msku_lock_status)
                VALUES
                    (1, 'SAYOLA:US', 'MSKU-001', 'SKU-001', NULL),
                    (2, 'SAYOLA:US', 'MSKU-002', 'SKU-002', NULL)
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    result = service.bulk_update_product_lock_status([1, 2], "锁", changed_by="admin")

    assert result == {"updated": 2, "requested": 2}
    with engine.connect() as conn:
        rows = conn.execute(
            text("SELECT id, msku_lock_status FROM amazon_product_info ORDER BY id")
        ).mappings().all()
        logs = conn.execute(
            text("SELECT record_id, operation_type, changed_by, change_data FROM amazon_operation_log ORDER BY record_id")
        ).mappings().all()

    assert [dict(row) for row in rows] == [
        {"id": 1, "msku_lock_status": "锁"},
        {"id": 2, "msku_lock_status": "锁"},
    ]
    assert [log["record_id"] for log in logs] == [1, 2]
    assert all(log["operation_type"] == "BULK_UPDATE" for log in logs)
    assert all(log["changed_by"] == "admin" for log in logs)
    assert json.loads(logs[0]["change_data"]) == {
        "msku_lock_status": {"old": None, "new": "锁"}
    }


def test_bulk_update_product_lock_status_rejects_lock_conflict(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    msku TEXT,
                    sku TEXT,
                    msku_lock_status TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_operation_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    table_name TEXT NOT NULL,
                    record_id INTEGER NOT NULL,
                    operation_type TEXT NOT NULL,
                    changed_by TEXT NOT NULL,
                    change_data TEXT NOT NULL,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (id, store_site, msku, sku, msku_lock_status)
                VALUES
                    (1, 'SAYOLA:US', 'MSKU-001', 'SKU-001', '锁'),
                    (2, 'SAYOLA:US', 'MSKU-002', 'SKU-001', NULL)
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    with pytest.raises(LockConflictError):
        service.bulk_update_product_lock_status([2], "锁", changed_by="admin")

    with engine.connect() as conn:
        status = conn.execute(
            text("SELECT msku_lock_status FROM amazon_product_info WHERE id = 2")
        ).scalar_one()
        log_count = conn.execute(text("SELECT COUNT(*) FROM amazon_operation_log")).scalar_one()

    assert status is None
    assert log_count == 0


def test_list_products_reuses_cached_page(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    asin TEXT,
                    msku TEXT,
                    store_site TEXT,
                    product_name TEXT,
                    sku TEXT,
                    brand TEXT,
                    sales_status TEXT,
                    listing TEXT,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    listing_status TEXT,
                    project_group TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, asin, msku, store_site, product_name, sku, brand, sales_status, listing, updated_at
                )
                VALUES (1, 'B001', 'MSKU-001', 'SAYOLA:US', 'Product A', 'SKU-001', 'BrandA', '在售', 'RB833', '2026-06-05')
                """
            )
        )

    connect_count = 0
    original_connect = engine.connect

    def counted_connect(*args, **kwargs):
        nonlocal connect_count
        connect_count += 1
        return original_connect(*args, **kwargs)

    monkeypatch.setattr(service, "get_engine", lambda: engine)
    monkeypatch.setattr(engine, "connect", counted_connect)
    service.clear_product_list_cache()

    first = list_products(ProductFilters())
    second = list_products(ProductFilters())

    assert first == second
    assert connect_count == 1


def test_list_products_for_export_uses_filters_without_pagination(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    asin TEXT,
                    msku TEXT,
                    store_site TEXT,
                    product_name TEXT,
                    sku TEXT,
                    brand TEXT,
                    listing TEXT,
                    sales_status TEXT,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    listing_status TEXT,
                    listing_maintainer TEXT,
                    include_inventory_age_assessment TEXT,
                    project_group TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, asin, msku, store_site, product_name, sku,
                    brand, listing, sales_status, updated_at
                )
                VALUES
                    (1, 'B001', 'MSKU-001', 'SAYOLA:US', 'Product A', 'SKU-001', 'BrandA', 'RB833', '在售', '2026-06-01'),
                    (2, 'B002', 'MSKU-002', 'SAYOLA:US', 'Product B', 'SKU-002', 'BrandB', 'RB831', '在售', '2026-06-01')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (store_site, listing, owner, listing_status, project_group)
                VALUES
                    ('SAYOLA:US', 'RB833', 'OwnerA', 'Active', 'GroupA'),
                    ('SAYOLA:US', 'RB831', 'OwnerB', 'Paused', 'GroupB')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    rows = list_products_for_export(ProductFilters(brand="BrandA", page=9))

    assert len(rows) == 1
    assert rows[0]["msku"] == "MSKU-001"
    assert rows[0]["brand"] == "BrandA"
    assert rows[0]["listing_owner"] == "OwnerA"


def test_export_products_to_xlsx_uses_selected_safe_fields(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    msku TEXT,
                    asin TEXT,
                    store_site TEXT,
                    parent_asin TEXT,
                    product_name TEXT,
                    sku TEXT,
                    brand TEXT,
                    fnsku TEXT,
                    sales_status TEXT,
                    storage_type TEXT,
                    category_level_1 TEXT,
                    category_a TEXT,
                    category_b TEXT,
                    listing TEXT,
                    label_name TEXT,
                    msku_shipping_remark TEXT,
                    transfer_remark TEXT,
                    msku_lock_status TEXT,
                    created_at TEXT,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    listing_status TEXT,
                    project_group TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, msku, asin, store_site, product_name, storage_type, listing, updated_at
                )
                VALUES (1, 'MSKU-001', 'B001', 'SAYOLA:US', 'Product 1', 'FBA', 'RB833', '2026-06-04')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (store_site, listing, owner, listing_status, project_group)
                VALUES ('SAYOLA:US', 'RB833', 'OwnerA', 'Active', 'GroupA')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    content = export_products_to_xlsx(
        ProductFilters(),
        ["msku", "storage_type", "listing_owner", "project_group", "not_a_column"],
    )
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    assert [cell.value for cell in sheet[1]] == ["MSKU", "仓储类型", "Listing 负责人", "项目组"]
    assert [cell.value for cell in sheet[2]] == ["MSKU-001", "FBA", "OwnerA", "GroupA"]


def test_export_products_for_import_to_xlsx_matches_import_template_and_preview(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    msku TEXT,
                    asin TEXT,
                    store_site TEXT,
                    parent_asin TEXT,
                    product_name TEXT,
                    sku TEXT,
                    brand TEXT,
                    fnsku TEXT,
                    sales_status TEXT,
                    storage_type TEXT,
                    category_level_1 TEXT,
                    category_a TEXT,
                    category_b TEXT,
                    listing TEXT,
                    label_name TEXT,
                    msku_shipping_remark TEXT,
                    transfer_remark TEXT,
                    msku_lock_status TEXT,
                    created_at TEXT,
                    updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    listing TEXT,
                    owner TEXT,
                    listing_status TEXT,
                    project_group TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, store_site, msku, asin, parent_asin, product_name, sku,
                    brand, fnsku, sales_status, storage_type, category_level_1,
                    category_a, category_b, listing, label_name,
                    msku_shipping_remark, transfer_remark, msku_lock_status, updated_at
                )
                VALUES (
                    1, 'SAYOLA:US', 'MSKU-001', 'B001', 'PARENT1', 'Product 1', 'SKU-001',
                    'BrandA', 'FNSKU-001', '在售', 'FBA', 'Sports',
                    'A', 'B', 'RB833', 'LabelA',
                    'Ship note', 'Transfer note', '', '2026-06-04'
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (
                    id, store_site, msku, asin, product_name, sku, brand, updated_at
                )
                VALUES (2, 'SAYOLA:US', '', 'B002', 'Missing Key', 'SKU-002', 'BrandB', '2026-06-04')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)
    monkeypatch.setattr(import_service, "get_engine", lambda: engine)

    content = service.export_products_for_import_to_xlsx(ProductFilters())
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    preview = preview_product_import(content)

    assert [cell.value for cell in sheet[1]] == list(IMPORT_TEMPLATE_HEADERS)
    assert sheet.max_row == 2
    assert preview["blocked_fields"] == []
    assert preview["total_rows"] == 1
    assert preview["valid_count"] == 1
    assert preview["missing_product_count"] == 0
    assert preview["error_count"] == 0


def test_get_filter_options_reuses_cached_values(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    brand TEXT,
                    sales_status TEXT,
                    listing TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_listing_owner_config (
                    id INTEGER PRIMARY KEY,
                    owner TEXT,
                    listing_status TEXT,
                    project_group TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (store_site, brand, sales_status, listing)
                VALUES ('SAYOLA:US', 'BrandA', '在售', 'RB833')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_listing_owner_config (owner, listing_status, project_group)
                VALUES ('OwnerA', 'Active', 'GroupA')
                """
            )
        )

    connect_count = 0

    def fake_get_engine():
        return engine

    original_connect = engine.connect

    def counted_connect(*args, **kwargs):
        nonlocal connect_count
        connect_count += 1
        return original_connect(*args, **kwargs)

    monkeypatch.setattr(service, "get_engine", fake_get_engine)
    monkeypatch.setattr(engine, "connect", counted_connect)
    service.clear_filter_options_cache()

    first = service.get_filter_options()
    second = service.get_filter_options()

    assert first == second
    assert connect_count == 1


def test_create_product_rejects_second_locked_msku_for_same_store_site_sku(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_store_site (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT NOT NULL
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_store_site (id, store_site)
                VALUES (1, 'SAYOLA:US')
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    store_site TEXT,
                    msku TEXT,
                    sku TEXT,
                    msku_lock_status TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (store_site, msku, sku, msku_lock_status)
                VALUES ('SAYOLA:US', 'MSKU-001', 'SKU-001', '锁')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    with pytest.raises(LockConflictError):
        create_product(
            {
                "store_site": "SAYOLA:US",
                "msku": "MSKU-002",
                "sku": "SKU-001",
                "msku_lock_status": "锁",
            }
        )

    with engine.connect() as conn:
        row_count = conn.execute(text("SELECT COUNT(*) FROM amazon_product_info")).scalar_one()

    assert row_count == 1


def test_create_product_rejects_unknown_store_site(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_store_site (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT NOT NULL
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    store_site TEXT,
                    msku TEXT
                )
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    with pytest.raises(UnknownStoreSiteError):
        create_product({"store_site": "UNKNOWN:US", "msku": "MSKU-001"})

    with engine.connect() as conn:
        row_count = conn.execute(text("SELECT COUNT(*) FROM amazon_product_info")).scalar_one()

    assert row_count == 0


def test_update_product_rejects_second_locked_msku_for_same_store_site_sku(monkeypatch):
    engine = create_engine("sqlite://")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE amazon_product_info (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT,
                    msku TEXT,
                    sku TEXT,
                    product_name TEXT,
                    brand TEXT,
                    listing TEXT,
                    sales_status TEXT,
                    storage_type TEXT,
                    category_level_1 TEXT,
                    category_a TEXT,
                    category_b TEXT,
                    label_name TEXT,
                    msku_shipping_remark TEXT,
                    transfer_remark TEXT,
                    msku_lock_status TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO amazon_product_info (id, store_site, msku, sku, msku_lock_status)
                VALUES
                    (1, 'SAYOLA:US', 'MSKU-001', 'SKU-001', '锁'),
                    (2, 'SAYOLA:US', 'MSKU-002', 'SKU-001', '否')
                """
            )
        )

    monkeypatch.setattr(service, "get_engine", lambda: engine)

    with pytest.raises(LockConflictError):
        update_product(2, {"msku_lock_status": "锁"})

    with engine.connect() as conn:
        lock_status = conn.execute(
            text("SELECT msku_lock_status FROM amazon_product_info WHERE id = 2")
        ).scalar_one()

    assert lock_status == "否"
