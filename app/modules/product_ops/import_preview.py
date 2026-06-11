from collections import Counter
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from sqlalchemy import bindparam, text
from sqlalchemy.exc import SQLAlchemyError

from app.core.db import get_engine
from app.shared.audit import build_change_set, record_operation_log


WAITING_MERGE_DIR = "\u5f85\u5408\u5e76"
SOURCE_DATA_DIR = "src_data"
ALLOCATION_DIR = "\u9500\u5360\u6bd4\u53c2\u6570"
FORECAST_DIR = "\u9500\u552e\u9884\u4f30\u53c2\u6570"
ALLOCATION_SHEET = "parameter"
FORECAST_SHEET = "\u9500\u91cf\u9884\u4f30\u8868-parameter"

COL_SITE = "\u7ad9\u70b9"
COL_OWNER = "\u8d1f\u8d23\u4eba"
COL_LISTING = "Listing"
COL_STYLE = "\u6b3e\u5f0f"
COL_MSKU = "MSKU"
COL_SKU = "\u79ef\u52a0SKU"
COL_SKU_MAINTENANCE = "SKU"
COL_STORE = "\u5e97\u94fa"
COL_STORE_SITE_ALIASES = ("\u5e97\u94fa/\u7ad9\u70b9", "\u5e97\u94fa\u7ad9\u70b9", "\u5e97\u94fa-\u7ad9\u70b9")
COL_MONTH = "\u6708\u4efd"
COL_FORECAST_UNITS = "Listing_\u6708\u5ea6\u9884\u4f30\u9500\u91cf"
COL_FORECAST_MAINTENANCE_UNITS = "Listing\u6708\u5ea6\u9884\u4f30\u9500\u91cf"
COL_CONFIRMED_STORE_SITE = "\u786e\u8ba4\u5e97\u94fa\u7ad9\u70b9"
COL_SCALE_POSITION = "\u89c4\u6a21\u5b9a\u4f4d"
COL_STYLE_SALES_RATIO = "\u6b3e\u5f0f\u9500\u5360\u6bd4"
COL_SKU_SALES_RATIO = "SKU\u9500\u5360\u6bd4"
COL_DEMAND_POSITION = "\u9700\u6c42\u5b9a\u4f4d"
COL_SHIPPING_POSITION = "\u53d1\u8d27\u5b9a\u4f4d"
COL_STOCKING_POSITION = "\u5907\u8d27\u5b9a\u4f4d"
COL_OPERATION_MIN_ORDER_DAYS = "\u8fd0\u8425\u4fdd\u5e95\u4e0b\u5355\u5929\u6570"
COL_TOTAL_SHIPPING_DAYS = "\u603b\u53d1\u8d27\u5929\u6570"

ISSUE_LIMIT = 20
IMPORT_CHUNK_SIZE = 500

FORECAST_REVIEW_EXPORT_COLUMNS = (
    ("file", "\u6587\u4ef6"),
    ("row_number", "\u884c\u53f7"),
    ("site", "\u7ad9\u70b9"),
    ("listing", "Listing"),
    ("forecast_month", "\u6708\u4efd"),
    ("forecast_units", "\u9884\u4f30\u9500\u91cf"),
    ("candidate_store_sites", "\u5019\u9009\u5e97\u94fa\u7ad9\u70b9"),
    ("owner_candidates", "\u5019\u9009\u8d1f\u8d23\u4eba\u914d\u7f6e"),
    ("confirmed_store_site", COL_CONFIRMED_STORE_SITE),
)


def build_product_ops_import_preview(base_dir: Path | None = None) -> dict[str, object]:
    source_dir = base_dir or Path.cwd() / WAITING_MERGE_DIR / SOURCE_DATA_DIR
    owner_keys, owner_site_listing_counts = _load_owner_config_keys()
    allocation, allocation_site_listing_store_sites = _preview_allocation(source_dir / ALLOCATION_DIR, owner_keys)
    forecast = _preview_forecast(
        source_dir / FORECAST_DIR,
        owner_keys,
        owner_site_listing_counts,
        allocation_site_listing_store_sites,
    )
    return {
        "base_dir": str(source_dir),
        "allocation": allocation,
        "forecast": forecast,
    }


def build_forecast_store_site_review(base_dir: Path | None = None) -> dict[str, object]:
    source_dir = base_dir or Path.cwd() / WAITING_MERGE_DIR / SOURCE_DATA_DIR
    owner_keys, owner_site_listing_counts = _load_owner_config_keys()
    owner_candidates = _load_owner_config_candidates()
    _, allocation_site_listing_store_sites = _preview_allocation(source_dir / ALLOCATION_DIR, owner_keys)
    rows = _forecast_store_site_review_rows(
        source_dir / FORECAST_DIR,
        owner_site_listing_counts,
        allocation_site_listing_store_sites,
        owner_candidates,
    )
    return {
        "base_dir": str(source_dir),
        "total": len(rows),
        "rows": rows,
    }


def build_forecast_store_site_review_workbook(review: dict[str, object]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "\u9500\u552e\u9884\u4f30\u5f85\u6392\u67e5"
    sheet.append([label for _, label in FORECAST_REVIEW_EXPORT_COLUMNS])
    for row in review.get("rows", []):
        sheet.append([row.get(key) for key, _ in FORECAST_REVIEW_EXPORT_COLUMNS])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def preview_forecast_store_site_corrections(content: bytes) -> dict[str, object]:
    owner_keys, _ = _load_owner_config_keys()
    rows = _read_workbook_rows(content)
    valid_rows = []
    error_rows = []

    for row in rows:
        listing = _clean(row.get("Listing"))
        confirmed_store_site = _clean(row.get(COL_CONFIRMED_STORE_SITE))
        preview_row = {
            "row_number": row.get("\u884c\u53f7") or row["excel_row_number"],
            "site": _clean(row.get("\u7ad9\u70b9")).upper(),
            "listing": listing,
            "forecast_month": _month(row.get("\u6708\u4efd")),
            "forecast_units": row.get("\u9884\u4f30\u9500\u91cf"),
            "confirmed_store_site": confirmed_store_site,
        }
        if not confirmed_store_site:
            error_rows.append({**preview_row, "message": "\u786e\u8ba4\u5e97\u94fa\u7ad9\u70b9\u4e3a\u7a7a"})
        elif not listing:
            error_rows.append({**preview_row, "message": "Listing\u4e3a\u7a7a"})
        elif (confirmed_store_site, listing) not in owner_keys:
            error_rows.append({**preview_row, "message": "\u786e\u8ba4\u5e97\u94fa\u7ad9\u70b9\u672a\u5339\u914d\u8d1f\u8d23\u4eba\u914d\u7f6e"})
        else:
            valid_rows.append(preview_row)

    return {
        "total_rows": len(rows),
        "valid_count": len(valid_rows),
        "error_count": len(error_rows),
        "valid_rows": valid_rows,
        "error_rows": error_rows,
    }


def commit_forecast_store_site_corrections(content: bytes, changed_by: str = "system") -> dict[str, object]:
    preview = preview_forecast_store_site_corrections(content)
    if preview["error_count"] > 0:
        return {
            "success": False,
            "inserted_count": 0,
            "updated_count": 0,
            "skipped_count": preview["error_count"],
            "message": "\u4fee\u6b63\u6587\u4ef6\u8fd8\u6709\u9519\u8bef\uff0c\u672a\u5199\u5165\u6570\u636e\u5e93\u3002",
            "preview": preview,
        }

    engine = get_engine()
    if engine is None:
        return {
            "success": False,
            "inserted_count": 0,
            "updated_count": 0,
            "skipped_count": preview["valid_count"],
            "message": "\u6570\u636e\u5e93\u8fde\u63a5\u4e0d\u53ef\u7528\uff0c\u672a\u5199\u5165\u6570\u636e\u5e93\u3002",
            "preview": preview,
        }

    inserted_count = 0
    updated_count = 0
    skipped_count = 0
    with engine.begin() as conn:
        _ensure_sales_forecast_table(conn)
        for row in preview["valid_rows"]:
            forecast_month = _forecast_month_date(row["forecast_month"])
            params = {
                "store_site": row["confirmed_store_site"],
                "site": row["site"],
                "listing": row["listing"],
                "forecast_month": forecast_month,
                "forecast_units": row["forecast_units"] or 0,
            }
            before = conn.execute(
                text(
                    """
                    SELECT id, site, forecast_units
                    FROM amazon_sales_forecast
                    WHERE store_site = :store_site
                      AND listing = :listing
                      AND forecast_month = :forecast_month
                    """
                ),
                params,
            ).mappings().first()
            if before is None:
                result = conn.execute(
                    text(
                        """
                        INSERT INTO amazon_sales_forecast (
                            store_site, site, listing, forecast_month, forecast_units
                        )
                        VALUES (
                            :store_site, :site, :listing, :forecast_month, :forecast_units
                        )
                        """
                    ),
                    params,
                )
                inserted_count += 1 if result.rowcount > 0 else 0
                skipped_count += 0 if result.rowcount > 0 else 1
            else:
                result = conn.execute(
                    text(
                        """
                        UPDATE amazon_sales_forecast
                        SET site = :site,
                            forecast_units = :forecast_units
                        WHERE id = :forecast_id
                        """
                    ),
                    {**params, "forecast_id": before["id"]},
                )
                updated_count += 1 if result.rowcount > 0 else 0
                skipped_count += 0 if result.rowcount > 0 else 1

    return {
        "success": True,
        "inserted_count": inserted_count,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "message": "\u5199\u5165\u5b8c\u6210",
        "preview": preview,
    }


def preview_sales_forecast_maintenance_import(content: bytes) -> dict[str, object]:
    owner_keys, _ = _load_owner_config_keys()
    rows = _read_workbook_rows(content)
    valid_rows = []
    error_rows = []

    for row in rows:
        store_site = _forecast_store_site(row)
        site = _clean(row.get(COL_SITE)).upper() or _site_from_store_site(store_site)
        listing = _clean(row.get(COL_LISTING))
        forecast_month = _month(row.get(COL_MONTH))
        forecast_units = _number_or_none(_forecast_maintenance_units(row))
        preview_row = {
            "row_number": row["excel_row_number"],
            "store_site": store_site,
            "site": site,
            "listing": listing,
            "forecast_month": forecast_month,
            "forecast_units": forecast_units,
        }
        if not store_site or not listing or not forecast_month:
            error_rows.append({**preview_row, "message": "\u5e97\u94fa\u7ad9\u70b9\u3001Listing \u6216\u6708\u4efd\u4e3a\u7a7a"})
        elif forecast_units is None:
            error_rows.append({**preview_row, "message": "Listing\u6708\u5ea6\u9884\u4f30\u9500\u91cf\u4e0d\u662f\u6709\u6548\u6570\u5b57"})
        elif (store_site, listing) not in owner_keys:
            error_rows.append({**preview_row, "message": "\u5e97\u94fa\u7ad9\u70b9 + Listing \u672a\u5339\u914d\u8d1f\u8d23\u4eba\u914d\u7f6e"})
        else:
            valid_rows.append(preview_row)

    key_counts = Counter((row["store_site"], row["listing"], row["forecast_month"]) for row in valid_rows)
    deduped_rows = []
    for row in valid_rows:
        key = (row["store_site"], row["listing"], row["forecast_month"])
        if key_counts[key] > 1:
            error_rows.append({**row, "message": "\u91cd\u590d\u552f\u4e00\u952e"})
        else:
            deduped_rows.append(row)

    existing = _load_existing_forecast_rows(deduped_rows)
    for row in deduped_rows:
        key = (row["store_site"], row["listing"], _forecast_month_date(row["forecast_month"]))
        row["action"] = "\u66f4\u65b0" if key in existing else "\u65b0\u589e"
        if key in existing:
            row["forecast_id"] = existing[key]["id"]

    insert_count = sum(1 for row in deduped_rows if row["action"] == "\u65b0\u589e")
    update_count = sum(1 for row in deduped_rows if row["action"] == "\u66f4\u65b0")
    return {
        "total_rows": len(rows),
        "valid_count": len(deduped_rows),
        "insert_count": insert_count,
        "update_count": update_count,
        "error_count": len(error_rows),
        "valid_rows": deduped_rows,
        "error_rows": error_rows,
    }


def commit_sales_forecast_maintenance_import(content: bytes, changed_by: str = "system") -> dict[str, object]:
    preview = preview_sales_forecast_maintenance_import(content)
    if preview["error_count"] > 0:
        return {
            "success": False,
            "inserted_count": 0,
            "updated_count": 0,
            "skipped_count": preview["error_count"],
            "message": "\u6821\u9a8c\u672a\u901a\u8fc7\uff0c\u672a\u5199\u5165\u6570\u636e\u5e93\u3002",
            "preview": preview,
        }

    engine = get_engine()
    if engine is None:
        return {
            "success": False,
            "inserted_count": 0,
            "updated_count": 0,
            "skipped_count": preview["valid_count"],
            "message": "\u6570\u636e\u5e93\u8fde\u63a5\u4e0d\u53ef\u7528\uff0c\u672a\u5199\u5165\u6570\u636e\u5e93\u3002",
            "preview": preview,
        }

    inserted_count = 0
    updated_count = 0
    skipped_count = 0
    with engine.begin() as conn:
        _ensure_sales_forecast_table(conn)
        for row in preview["valid_rows"]:
            params = {
                "store_site": row["store_site"],
                "site": row["site"],
                "listing": row["listing"],
                "forecast_month": _forecast_month_date(row["forecast_month"]),
                "forecast_units": row["forecast_units"],
            }
            before = conn.execute(
                text(
                    """
                    SELECT id, site, forecast_units
                    FROM amazon_sales_forecast
                    WHERE store_site = :store_site
                      AND listing = :listing
                      AND forecast_month = :forecast_month
                    """
                ),
                params,
            ).mappings().first()
            if before is None:
                conn.execute(
                    text(
                        """
                        INSERT INTO amazon_sales_forecast (
                            store_site, site, listing, forecast_month, forecast_units
                        )
                        VALUES (
                            :store_site, :site, :listing, :forecast_month, :forecast_units
                        )
                        """
                    ),
                    params,
                )
                forecast_id = conn.execute(
                    text(
                        """
                        SELECT id
                        FROM amazon_sales_forecast
                        WHERE store_site = :store_site
                          AND listing = :listing
                          AND forecast_month = :forecast_month
                        """
                    ),
                    params,
                ).scalar_one()
                inserted_count += 1
                record_operation_log(
                    conn,
                    table_name="amazon_sales_forecast",
                    record_id=forecast_id,
                    operation_type="IMPORT_INSERT",
                    change_data={field: {"old": None, "new": value} for field, value in params.items()},
                    changed_by=changed_by,
                )
                continue

            changes = build_change_set(dict(before), {"site": params["site"], "forecast_units": params["forecast_units"]})
            if not changes:
                skipped_count += 1
                continue
            conn.execute(
                text(
                    """
                    UPDATE amazon_sales_forecast
                    SET site = :site,
                        forecast_units = :forecast_units
                    WHERE id = :forecast_id
                    """
                ),
                {**params, "forecast_id": before["id"]},
            )
            updated_count += 1
            record_operation_log(
                conn,
                table_name="amazon_sales_forecast",
                record_id=before["id"],
                operation_type="IMPORT_UPDATE",
                change_data=changes,
                changed_by=changed_by,
            )

    return {
        "success": True,
        "inserted_count": inserted_count,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "message": "\u5199\u5165\u5b8c\u6210",
        "preview": preview,
    }


def preview_sales_allocation_maintenance_import(content: bytes) -> dict[str, object]:
    owner_keys, _ = _load_owner_config_keys()
    rows = _read_workbook_rows(content)
    valid_rows = []
    error_rows = []

    for row in rows:
        store_site = _forecast_store_site(row)
        site = _clean(row.get(COL_SITE)).upper() or _site_from_store_site(store_site)
        listing = _clean(row.get(COL_LISTING))
        msku = _clean(row.get(COL_MSKU))
        sku = _allocation_maintenance_sku(row)
        style_sales_ratio, style_ratio_ok = _optional_number(row.get(COL_STYLE_SALES_RATIO), allow_percent=True)
        sku_sales_ratio, sku_ratio_ok = _optional_number(row.get(COL_SKU_SALES_RATIO), allow_percent=True)
        operation_min_order_days, min_days_ok = _optional_int(row.get(COL_OPERATION_MIN_ORDER_DAYS))
        total_shipping_days, total_days_ok = _optional_int(row.get(COL_TOTAL_SHIPPING_DAYS))
        preview_row = {
            "row_number": row["excel_row_number"],
            "store_site": store_site,
            "site": site,
            "owner": _clean(row.get(COL_OWNER)),
            "listing": listing,
            "style": _clean(row.get(COL_STYLE)),
            "msku": msku,
            "sku": sku,
            "scale_position": _clean(row.get(COL_SCALE_POSITION)),
            "style_sales_ratio": style_sales_ratio,
            "sku_sales_ratio": sku_sales_ratio,
            "demand_position": _clean(row.get(COL_DEMAND_POSITION)),
            "shipping_position": _clean(row.get(COL_SHIPPING_POSITION)),
            "stocking_position": _clean(row.get(COL_STOCKING_POSITION)),
            "operation_min_order_days": operation_min_order_days,
            "total_shipping_days": total_shipping_days,
        }
        if not store_site or not listing or not msku or not sku:
            error_rows.append({**preview_row, "message": "店铺站点、Listing、MSKU 或 SKU 为空"})
        elif not style_ratio_ok:
            error_rows.append({**preview_row, "message": "款式销占比不是有效数字"})
        elif not sku_ratio_ok:
            error_rows.append({**preview_row, "message": "SKU销占比不是有效数字"})
        elif not min_days_ok:
            error_rows.append({**preview_row, "message": "运营保底下单天数不是有效数字"})
        elif not total_days_ok:
            error_rows.append({**preview_row, "message": "总发货天数不是有效数字"})
        elif (store_site, listing) not in owner_keys:
            error_rows.append({**preview_row, "message": "店铺站点 + Listing 未匹配负责人配置"})
        else:
            valid_rows.append(preview_row)

    key_counts = Counter((row["store_site"], row["listing"], row["msku"]) for row in valid_rows)
    deduped_rows = []
    for row in valid_rows:
        key = (row["store_site"], row["listing"], row["msku"])
        if key_counts[key] > 1:
            error_rows.append({**row, "message": "重复唯一键"})
        else:
            deduped_rows.append(row)

    existing = _load_existing_allocation_rows(deduped_rows)
    for row in deduped_rows:
        key = (row["store_site"], row["listing"], row["msku"])
        row["action"] = "更新" if key in existing else "新增"
        if key in existing:
            row["allocation_id"] = existing[key]["id"]

    insert_count = sum(1 for row in deduped_rows if row["action"] == "新增")
    update_count = sum(1 for row in deduped_rows if row["action"] == "更新")
    return {
        "total_rows": len(rows),
        "valid_count": len(deduped_rows),
        "insert_count": insert_count,
        "update_count": update_count,
        "error_count": len(error_rows),
        "valid_rows": deduped_rows,
        "error_rows": error_rows,
    }


def commit_sales_allocation_maintenance_import(content: bytes, changed_by: str = "system") -> dict[str, object]:
    preview = preview_sales_allocation_maintenance_import(content)
    if preview["error_count"] > 0:
        return {
            "success": False,
            "inserted_count": 0,
            "updated_count": 0,
            "skipped_count": preview["error_count"],
            "message": "校验未通过，未写入数据库。",
            "preview": preview,
        }

    engine = get_engine()
    if engine is None:
        return {
            "success": False,
            "inserted_count": 0,
            "updated_count": 0,
            "skipped_count": preview["valid_count"],
            "message": "数据库连接不可用，未写入数据库。",
            "preview": preview,
        }

    inserted_count = 0
    updated_count = 0
    skipped_count = 0
    update_fields = (
        "site",
        "owner",
        "style",
        "sku",
        "scale_position",
        "style_sales_ratio",
        "sku_sales_ratio",
        "demand_position",
        "shipping_position",
        "stocking_position",
        "operation_min_order_days",
        "total_shipping_days",
    )
    with engine.begin() as conn:
        _ensure_sales_allocation_table(conn)
        for row in preview["valid_rows"]:
            params = {
                "store_site": row["store_site"],
                "site": row["site"],
                "owner": row["owner"],
                "listing": row["listing"],
                "style": row["style"],
                "msku": row["msku"],
                "sku": row["sku"],
                "scale_position": row["scale_position"],
                "style_sales_ratio": row["style_sales_ratio"],
                "sku_sales_ratio": row["sku_sales_ratio"],
                "demand_position": row["demand_position"],
                "shipping_position": row["shipping_position"],
                "stocking_position": row["stocking_position"],
                "operation_min_order_days": row["operation_min_order_days"],
                "total_shipping_days": row["total_shipping_days"],
            }
            before = conn.execute(
                text(
                    """
                    SELECT
                        id, site, owner, style, sku, scale_position,
                        style_sales_ratio, sku_sales_ratio, demand_position,
                        shipping_position, stocking_position,
                        operation_min_order_days, total_shipping_days
                    FROM amazon_sales_allocation
                    WHERE store_site = :store_site
                      AND listing = :listing
                      AND msku = :msku
                    """
                ),
                params,
            ).mappings().first()
            if before is None:
                conn.execute(
                    text(
                        """
                        INSERT INTO amazon_sales_allocation (
                            store_site, site, owner, listing, style, msku, sku,
                            scale_position, style_sales_ratio, sku_sales_ratio,
                            demand_position, shipping_position, stocking_position,
                            operation_min_order_days, total_shipping_days
                        )
                        VALUES (
                            :store_site, :site, :owner, :listing, :style, :msku, :sku,
                            :scale_position, :style_sales_ratio, :sku_sales_ratio,
                            :demand_position, :shipping_position, :stocking_position,
                            :operation_min_order_days, :total_shipping_days
                        )
                        """
                    ),
                    params,
                )
                allocation_id = conn.execute(
                    text(
                        """
                        SELECT id
                        FROM amazon_sales_allocation
                        WHERE store_site = :store_site
                          AND listing = :listing
                          AND msku = :msku
                        """
                    ),
                    params,
                ).scalar_one()
                inserted_count += 1
                record_operation_log(
                    conn,
                    table_name="amazon_sales_allocation",
                    record_id=allocation_id,
                    operation_type="IMPORT_INSERT",
                    change_data={field: {"old": None, "new": value} for field, value in params.items()},
                    changed_by=changed_by,
                )
                continue

            after = {field: params[field] for field in update_fields}
            changes = build_change_set(dict(before), after)
            if not changes:
                skipped_count += 1
                continue
            conn.execute(
                text(
                    """
                    UPDATE amazon_sales_allocation
                    SET site = :site,
                        owner = :owner,
                        style = :style,
                        sku = :sku,
                        scale_position = :scale_position,
                        style_sales_ratio = :style_sales_ratio,
                        sku_sales_ratio = :sku_sales_ratio,
                        demand_position = :demand_position,
                        shipping_position = :shipping_position,
                        stocking_position = :stocking_position,
                        operation_min_order_days = :operation_min_order_days,
                        total_shipping_days = :total_shipping_days
                    WHERE id = :allocation_id
                    """
                ),
                {**params, "allocation_id": before["id"]},
            )
            updated_count += 1
            record_operation_log(
                conn,
                table_name="amazon_sales_allocation",
                record_id=before["id"],
                operation_type="IMPORT_UPDATE",
                change_data=changes,
                changed_by=changed_by,
            )

    return {
        "success": True,
        "inserted_count": inserted_count,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "message": "写入完成",
        "preview": preview,
    }


def commit_product_ops_source_import(base_dir: Path | None = None, changed_by: str = "system") -> dict[str, object]:
    source_dir = base_dir or Path.cwd() / WAITING_MERGE_DIR / SOURCE_DATA_DIR
    owner_keys, owner_site_listing_counts = _load_owner_config_keys()
    engine = get_engine()
    if engine is None:
        return {
            "success": False,
            "message": "\u6570\u636e\u5e93\u8fde\u63a5\u4e0d\u53ef\u7528\uff0c\u672a\u5199\u5165\u6570\u636e\u5e93\u3002",
            "source_dir": str(source_dir),
            "allocation": _empty_commit_result(),
            "forecast": _empty_commit_result(),
        }

    allocation_rows, allocation_site_listing_store_sites, allocation_skipped = _source_allocation_import_rows(
        source_dir / ALLOCATION_DIR,
        owner_keys,
    )
    forecast_source_rows = _read_forecast_source_rows(source_dir / FORECAST_DIR)

    with engine.begin() as conn:
        _ensure_product_ops_import_tables(conn)
        existing_forecast_store_sites = _load_existing_forecast_store_sites(conn)
        forecast_rows, forecast_skipped = _source_forecast_import_rows(
            forecast_source_rows,
            owner_keys,
            owner_site_listing_counts,
            allocation_site_listing_store_sites,
            existing_forecast_store_sites,
        )
        allocation_result = _upsert_allocation_rows(conn, allocation_rows)
        forecast_result = _upsert_forecast_rows(conn, forecast_rows)

    allocation_result["skipped_count"] += allocation_skipped
    allocation_result["total_rows"] = allocation_result["inserted_count"] + allocation_result["updated_count"] + allocation_result["skipped_count"]
    forecast_result["skipped_count"] += forecast_skipped
    forecast_result["total_rows"] = forecast_result["inserted_count"] + forecast_result["updated_count"] + forecast_result["skipped_count"]
    return {
        "success": True,
        "message": "\u6b63\u5f0f\u5bfc\u5165\u5b8c\u6210",
        "source_dir": str(source_dir),
        "allocation": allocation_result,
        "forecast": forecast_result,
    }


def _empty_commit_result() -> dict[str, int]:
    return {
        "total_rows": 0,
        "inserted_count": 0,
        "updated_count": 0,
        "skipped_count": 0,
    }


def _source_allocation_import_rows(
    directory: Path,
    owner_keys: set[tuple[str, str]],
) -> tuple[list[dict[str, object]], dict[tuple[str, str], set[str]], int]:
    rows = []
    for file_path in _xlsx_files(directory):
        rows.extend(_read_sheet_rows(file_path, ALLOCATION_SHEET, [COL_SITE, COL_LISTING, COL_MSKU, COL_STORE]))

    import_rows = []
    skipped_count = 0
    allocation_site_listing_store_sites: dict[tuple[str, str], set[str]] = {}
    for row in rows:
        store_site = _store_site(row.get(COL_STORE), row.get(COL_SITE))
        site = _site_from_store_site(store_site)
        listing = _clean(row.get(COL_LISTING))
        msku = _clean(row.get(COL_MSKU))
        sku = _clean(row.get(COL_SKU))
        key = (store_site, listing, msku)
        owner_key = (store_site, listing)
        if not _complete_key(key) or not sku or owner_key not in owner_keys:
            skipped_count += 1
            continue

        allocation_site_listing_store_sites.setdefault((site, listing), set()).add(store_site)
        import_rows.append(
            {
                "store_site": store_site,
                "site": site,
                "owner": _clean(row.get(COL_OWNER)),
                "listing": listing,
                "style": _clean(row.get(COL_STYLE)),
                "msku": msku,
                "sku": sku,
                "scale_position": _clean(row.get(COL_SCALE_POSITION)),
                "style_sales_ratio": _number_or_none(row.get(COL_STYLE_SALES_RATIO)),
                "sku_sales_ratio": _number_or_none(row.get(COL_SKU_SALES_RATIO)),
                "demand_position": _clean(row.get(COL_DEMAND_POSITION)),
                "shipping_position": _clean(row.get(COL_SHIPPING_POSITION)),
                "stocking_position": _clean(row.get(COL_STOCKING_POSITION)),
                "operation_min_order_days": _int_or_none(row.get(COL_OPERATION_MIN_ORDER_DAYS)),
                "total_shipping_days": _int_or_none(row.get(COL_TOTAL_SHIPPING_DAYS)),
            }
        )

    key_counts = Counter((row["store_site"], row["listing"], row["msku"]) for row in import_rows)
    deduped_rows = [row for row in import_rows if key_counts[(row["store_site"], row["listing"], row["msku"])] == 1]
    skipped_count += len(import_rows) - len(deduped_rows)
    return deduped_rows, allocation_site_listing_store_sites, skipped_count


def _read_forecast_source_rows(directory: Path) -> list[dict[str, object]]:
    rows = []
    for file_path in _xlsx_files(directory):
        rows.extend(_read_sheet_rows(file_path, FORECAST_SHEET, [COL_SITE, COL_LISTING, COL_MONTH]))
    return rows


def _source_forecast_import_rows(
    rows: list[dict[str, object]],
    owner_keys: set[tuple[str, str]],
    owner_site_listing_counts: Counter[tuple[str, str]],
    allocation_site_listing_store_sites: dict[tuple[str, str], set[str]],
    existing_forecast_store_sites: dict[tuple[str, str, str], str],
) -> tuple[list[dict[str, object]], int]:
    import_rows = []
    skipped_count = 0
    for row in rows:
        site = _clean(row.get(COL_SITE)).upper()
        listing = _clean(row.get(COL_LISTING))
        month = _month(row.get(COL_MONTH))
        owner_key = (site, listing)
        explicit_store_site = _forecast_store_site(row)
        inferred_store_site = _inferred_store_site(owner_key, allocation_site_listing_store_sites)
        corrected_store_site = existing_forecast_store_sites.get((site, listing, month), "")
        store_site = explicit_store_site or inferred_store_site or corrected_store_site
        if not site or not listing or not month or not store_site:
            skipped_count += 1
            continue
        if (store_site, listing) not in owner_keys:
            skipped_count += 1
            continue
        if not explicit_store_site and not inferred_store_site and not corrected_store_site:
            skipped_count += 1
            continue
        if not explicit_store_site and not inferred_store_site and owner_site_listing_counts[owner_key] > 1 and not corrected_store_site:
            skipped_count += 1
            continue

        import_rows.append(
            {
                "store_site": store_site,
                "site": site,
                "listing": listing,
                "forecast_month": _forecast_month_date(month),
                "forecast_units": _number_or_zero(row.get(COL_FORECAST_UNITS)),
            }
        )

    key_counts = Counter((row["store_site"], row["listing"], row["forecast_month"]) for row in import_rows)
    deduped_rows = [row for row in import_rows if key_counts[(row["store_site"], row["listing"], row["forecast_month"])] == 1]
    skipped_count += len(import_rows) - len(deduped_rows)
    return deduped_rows, skipped_count


def _upsert_allocation_rows(conn, rows: list[dict[str, object]]) -> dict[str, int]:
    if conn.dialect.name == "mysql":
        return _bulk_upsert_allocation_rows(conn, rows)

    result = _empty_commit_result()
    for row in rows:
        before = conn.execute(
            text(
                """
                SELECT id
                FROM amazon_sales_allocation
                WHERE store_site = :store_site
                  AND listing = :listing
                  AND msku = :msku
                """
            ),
            row,
        ).mappings().first()
        if before is None:
            conn.execute(
                text(
                    """
                    INSERT INTO amazon_sales_allocation (
                        store_site, site, owner, listing, style, msku, sku,
                        scale_position, style_sales_ratio, sku_sales_ratio,
                        demand_position, shipping_position, stocking_position,
                        operation_min_order_days, total_shipping_days
                    )
                    VALUES (
                        :store_site, :site, :owner, :listing, :style, :msku, :sku,
                        :scale_position, :style_sales_ratio, :sku_sales_ratio,
                        :demand_position, :shipping_position, :stocking_position,
                        :operation_min_order_days, :total_shipping_days
                    )
                    """
                ),
                row,
            )
            result["inserted_count"] += 1
        else:
            conn.execute(
                text(
                    """
                    UPDATE amazon_sales_allocation
                    SET site = :site,
                        owner = :owner,
                        style = :style,
                        sku = :sku,
                        scale_position = :scale_position,
                        style_sales_ratio = :style_sales_ratio,
                        sku_sales_ratio = :sku_sales_ratio,
                        demand_position = :demand_position,
                        shipping_position = :shipping_position,
                        stocking_position = :stocking_position,
                        operation_min_order_days = :operation_min_order_days,
                        total_shipping_days = :total_shipping_days
                    WHERE id = :allocation_id
                    """
                ),
                {**row, "allocation_id": before["id"]},
            )
            result["updated_count"] += 1
    return result


def _upsert_forecast_rows(conn, rows: list[dict[str, object]]) -> dict[str, int]:
    if conn.dialect.name == "mysql":
        return _bulk_upsert_forecast_rows(conn, rows)

    result = _empty_commit_result()
    for row in rows:
        before = conn.execute(
            text(
                """
                SELECT id
                FROM amazon_sales_forecast
                WHERE store_site = :store_site
                  AND listing = :listing
                  AND forecast_month = :forecast_month
                """
            ),
            row,
        ).mappings().first()
        if before is None:
            conn.execute(
                text(
                    """
                    INSERT INTO amazon_sales_forecast (
                        store_site, site, listing, forecast_month, forecast_units
                    )
                    VALUES (
                        :store_site, :site, :listing, :forecast_month, :forecast_units
                    )
                    """
                ),
                row,
            )
            result["inserted_count"] += 1
        else:
            conn.execute(
                text(
                    """
                    UPDATE amazon_sales_forecast
                    SET site = :site,
                        forecast_units = :forecast_units
                    WHERE id = :forecast_id
                    """
                ),
                {**row, "forecast_id": before["id"]},
            )
            result["updated_count"] += 1
    return result


def _bulk_upsert_allocation_rows(conn, rows: list[dict[str, object]]) -> dict[str, int]:
    result = _empty_commit_result()
    if not rows:
        return result

    keys = sorted(set((row["store_site"], row["listing"], row["msku"]) for row in rows))
    before = _existing_allocation_keys(conn, keys)
    statement = text(
        """
        INSERT INTO amazon_sales_allocation (
            store_site, site, owner, listing, style, msku, sku,
            scale_position, style_sales_ratio, sku_sales_ratio,
            demand_position, shipping_position, stocking_position,
            operation_min_order_days, total_shipping_days
        )
        VALUES (
            :store_site, :site, :owner, :listing, :style, :msku, :sku,
            :scale_position, :style_sales_ratio, :sku_sales_ratio,
            :demand_position, :shipping_position, :stocking_position,
            :operation_min_order_days, :total_shipping_days
        )
        ON DUPLICATE KEY UPDATE
            site = VALUES(site),
            owner = VALUES(owner),
            style = VALUES(style),
            sku = VALUES(sku),
            scale_position = VALUES(scale_position),
            style_sales_ratio = VALUES(style_sales_ratio),
            sku_sales_ratio = VALUES(sku_sales_ratio),
            demand_position = VALUES(demand_position),
            shipping_position = VALUES(shipping_position),
            stocking_position = VALUES(stocking_position),
            operation_min_order_days = VALUES(operation_min_order_days),
            total_shipping_days = VALUES(total_shipping_days)
        """
    )
    for row_chunk in _chunks(rows):
        conn.execute(statement, row_chunk)

    after = _existing_allocation_keys(conn, keys)
    result["inserted_count"] = len(after - before)
    result["updated_count"] = len(after & before)
    result["skipped_count"] = len(set(keys) - after)
    return result


def _bulk_upsert_forecast_rows(conn, rows: list[dict[str, object]]) -> dict[str, int]:
    result = _empty_commit_result()
    if not rows:
        return result

    keys = sorted(set((row["store_site"], row["listing"], row["forecast_month"]) for row in rows))
    before = _existing_forecast_keys(conn, keys)
    statement = text(
        """
        INSERT INTO amazon_sales_forecast (store_site, site, listing, forecast_month, forecast_units)
        VALUES (:store_site, :site, :listing, :forecast_month, :forecast_units)
        ON DUPLICATE KEY UPDATE
            site = VALUES(site),
            forecast_units = VALUES(forecast_units)
        """
    )
    for row_chunk in _chunks(rows):
        conn.execute(statement, row_chunk)

    after = _existing_forecast_keys(conn, keys)
    result["inserted_count"] = len(after - before)
    result["updated_count"] = len(after & before)
    result["skipped_count"] = len(set(keys) - after)
    return result


def _existing_allocation_keys(conn, keys: list[tuple[str, str, str]]) -> set[tuple[str, str, str]]:
    if not keys:
        return set()

    statement = text(
        """
        SELECT store_site, listing, msku
        FROM amazon_sales_allocation
        WHERE (store_site, listing, msku) IN :keys
        """
    ).bindparams(bindparam("keys", expanding=True))
    existing = set()
    for key_chunk in _chunks(keys):
        existing.update(tuple(row) for row in conn.execute(statement, {"keys": key_chunk}).all())
    return existing


def _existing_forecast_keys(conn, keys: list[tuple[str, str, str]]) -> set[tuple[str, str, str]]:
    if not keys:
        return set()

    if conn.dialect.name == "mysql":
        forecast_month_expr = "DATE_FORMAT(forecast_month, '%Y-%m-%d')"
    else:
        forecast_month_expr = "forecast_month"
    statement = text(
        f"""
        SELECT store_site, listing, {forecast_month_expr} AS forecast_month
        FROM amazon_sales_forecast
        WHERE (store_site, listing, forecast_month) IN :keys
        """
    ).bindparams(bindparam("keys", expanding=True))
    existing = set()
    for key_chunk in _chunks(keys):
        existing.update(tuple(row) for row in conn.execute(statement, {"keys": key_chunk}).all())
    return existing


def _chunks(items: list, size: int = IMPORT_CHUNK_SIZE):
    for index in range(0, len(items), size):
        yield items[index:index + size]


def _load_existing_forecast_store_sites(conn) -> dict[tuple[str, str, str], str]:
    try:
        rows = conn.execute(
            text(
                """
                SELECT site, listing, forecast_month, store_site
                FROM amazon_sales_forecast
                WHERE store_site IS NOT NULL
                  AND TRIM(store_site) <> ''
                  AND site IS NOT NULL
                  AND TRIM(site) <> ''
                  AND listing IS NOT NULL
                  AND TRIM(listing) <> ''
                """
            )
        ).mappings()
    except SQLAlchemyError:
        return {}

    candidates: dict[tuple[str, str, str], set[str]] = {}
    for row in rows:
        key = (_clean(row["site"]).upper(), _clean(row["listing"]), _month(row["forecast_month"]))
        if _complete_key(key):
            candidates.setdefault(key, set()).add(_clean(row["store_site"]))
    return {key: next(iter(values)) for key, values in candidates.items() if len(values) == 1}


def _ensure_product_ops_import_tables(conn) -> None:
    _ensure_sales_allocation_table(conn)
    _ensure_sales_forecast_table(conn)


def _ensure_sales_allocation_table(conn) -> None:
    if conn.dialect.name == "sqlite":
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS amazon_sales_allocation (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT NOT NULL,
                    site TEXT NOT NULL,
                    owner TEXT,
                    listing TEXT NOT NULL,
                    style TEXT,
                    msku TEXT NOT NULL,
                    sku TEXT NOT NULL,
                    scale_position TEXT,
                    style_sales_ratio REAL,
                    sku_sales_ratio REAL,
                    demand_position TEXT,
                    shipping_position TEXT,
                    stocking_position TEXT,
                    operation_min_order_days INTEGER,
                    total_shipping_days INTEGER
                )
                """
            )
        )
        return

    conn.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS amazon_sales_allocation (
                id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT COMMENT '技术主键',
                store_site VARCHAR(50) NOT NULL COMMENT '店铺/站点',
                site VARCHAR(20) NOT NULL COMMENT '站点',
                owner VARCHAR(50) NULL COMMENT '负责人',
                listing VARCHAR(100) NOT NULL COMMENT 'Listing',
                style VARCHAR(100) NULL COMMENT '款式',
                msku VARCHAR(100) NOT NULL COMMENT 'MSKU',
                sku VARCHAR(100) NOT NULL COMMENT '积加SKU',
                scale_position VARCHAR(50) NULL COMMENT '规模定位',
                style_sales_ratio DECIMAL(12, 6) NULL COMMENT '款式销占比',
                sku_sales_ratio DECIMAL(12, 6) NULL COMMENT 'SKU销占比',
                demand_position VARCHAR(50) NULL COMMENT '需求定位',
                shipping_position VARCHAR(50) NULL COMMENT '发货定位',
                stocking_position VARCHAR(50) NULL COMMENT '备货定位',
                operation_min_order_days INT NULL COMMENT '运营保底下单天数',
                total_shipping_days INT NULL COMMENT '总发货天数',
                created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
                updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
                PRIMARY KEY (id),
                UNIQUE KEY uk_store_site_listing_msku (store_site, listing, msku),
                KEY idx_store_site_listing (store_site, listing),
                KEY idx_sku (sku),
                KEY idx_owner (owner)
            ) ENGINE=InnoDB
              DEFAULT CHARSET=utf8mb4
              COLLATE=utf8mb4_0900_ai_ci
              COMMENT='Amazon销占比配置表'
            """
        )
    )


def _ensure_sales_forecast_table(conn) -> None:
    if conn.dialect.name == "sqlite":
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS amazon_sales_forecast (
                    id INTEGER PRIMARY KEY,
                    store_site TEXT NOT NULL,
                    site TEXT NOT NULL,
                    listing TEXT NOT NULL,
                    forecast_month TEXT NOT NULL,
                    forecast_units REAL NOT NULL DEFAULT 0
                )
                """
            )
        )
        return

    conn.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS amazon_sales_forecast (
                id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT COMMENT '技术主键',
                store_site VARCHAR(50) NOT NULL COMMENT '店铺/站点',
                site VARCHAR(20) NOT NULL COMMENT '站点',
                listing VARCHAR(100) NOT NULL COMMENT 'Listing',
                forecast_month DATE NOT NULL COMMENT '预估月份，使用当月1日',
                forecast_units DECIMAL(14, 2) NOT NULL DEFAULT 0 COMMENT 'Listing月度预估销量',
                created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
                updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
                PRIMARY KEY (id),
                UNIQUE KEY uk_store_site_listing_month (store_site, listing, forecast_month),
                KEY idx_store_site_listing (store_site, listing),
                KEY idx_site_listing (site, listing),
                KEY idx_listing (listing),
                KEY idx_forecast_month (forecast_month)
            ) ENGINE=InnoDB
              DEFAULT CHARSET=utf8mb4
              COLLATE=utf8mb4_0900_ai_ci
              COMMENT='Amazon销售预估表'
            """
        )
    )


def _preview_allocation(
    directory: Path,
    owner_keys: set[tuple[str, str]],
) -> tuple[dict[str, object], dict[tuple[str, str], set[str]]]:
    rows = []
    files = _xlsx_files(directory)
    for file_path in files:
        rows.extend(_read_sheet_rows(file_path, ALLOCATION_SHEET, [COL_SITE, COL_LISTING, COL_MSKU, COL_STORE]))

    for row in rows:
        row["store_site"] = _store_site(row.get(COL_STORE), row.get(COL_SITE))
        row["listing"] = _clean(row.get(COL_LISTING))
        row["msku"] = _clean(row.get(COL_MSKU))
        row["key"] = (row["store_site"], row["listing"], row["msku"])
        row["owner_key"] = (row["store_site"], row["listing"])
        row["site_listing_key"] = (_site_from_store_site(row["store_site"]), row["listing"])
        row["display_key"] = f"{row['store_site']} / {row['listing']} / {row['msku']}"

    allocation_site_listing_store_sites: dict[tuple[str, str], set[str]] = {}
    for row in rows:
        if _complete_key(row["key"]):
            allocation_site_listing_store_sites.setdefault(row["site_listing_key"], set()).add(row["store_site"])

    key_counts = Counter(row["key"] for row in rows if _complete_key(row["key"]))
    missing_key_count = sum(1 for row in rows if not _complete_key(row["key"]))
    duplicate_key_count = sum(1 for row in rows if _complete_key(row["key"]) and key_counts[row["key"]] > 1)
    owner_unmatched_count = sum(
        1 for row in rows if _complete_key(row["key"]) and row["owner_key"] not in owner_keys
    )

    issue_rows = []
    for row in rows:
        if not _complete_key(row["key"]):
            _append_issue(issue_rows, "\u4e3b\u952e\u7f3a\u5931", row)
        elif key_counts[row["key"]] > 1:
            _append_issue(issue_rows, "\u91cd\u590d\u552f\u4e00\u952e", row)
        elif row["owner_key"] not in owner_keys:
            _append_issue(issue_rows, "\u672a\u5339\u914d\u8d1f\u8d23\u4eba\u914d\u7f6e", row)

    return {
        "label": "\u9500\u5360\u6bd4",
        "file_count": len(files),
        "total_rows": len(rows),
        "missing_key_count": missing_key_count,
        "duplicate_key_count": duplicate_key_count,
        "owner_unmatched_count": owner_unmatched_count,
        "owner_ambiguous_count": 0,
        "owner_resolved_by_allocation_count": 0,
        "owner_unresolved_ambiguous_count": 0,
        "issue_rows": issue_rows,
    }, allocation_site_listing_store_sites


def _preview_forecast(
    directory: Path,
    owner_keys: set[tuple[str, str]],
    owner_site_listing_counts: Counter[tuple[str, str]],
    allocation_site_listing_store_sites: dict[tuple[str, str], set[str]],
) -> dict[str, object]:
    rows = []
    files = _xlsx_files(directory)
    for file_path in files:
        rows.extend(_read_sheet_rows(file_path, FORECAST_SHEET, [COL_SITE, COL_LISTING, COL_MONTH]))

    for row in rows:
        row["site"] = _clean(row.get(COL_SITE)).upper()
        row["listing"] = _clean(row.get(COL_LISTING))
        row["month"] = _month(row.get(COL_MONTH))
        row["explicit_store_site"] = _forecast_store_site(row)
        row["owner_key"] = (row["site"], row["listing"])
        row["inferred_store_site"] = _inferred_store_site(row["owner_key"], allocation_site_listing_store_sites)
        row["final_store_site"] = row["explicit_store_site"] or row["inferred_store_site"]
        row["key"] = (row["final_store_site"], row["listing"], row["month"])
        if row["final_store_site"]:
            row["display_key"] = f"{row['final_store_site']} / {row['listing']} / {row['month']}"
        else:
            row["display_key"] = f"{row['site']} / {row['listing']} / {row['month']}"

    key_counts = Counter(row["key"] for row in rows if _complete_key(row["key"]))
    missing_key_count = sum(1 for row in rows if not _complete_key(row["key"]))
    duplicate_key_count = sum(1 for row in rows if _complete_key(row["key"]) and key_counts[row["key"]] > 1)
    owner_unmatched_count = sum(
        1 for row in rows if row["listing"] and row["site"] and _owner_match_count(row, owner_keys, owner_site_listing_counts) == 0
    )
    owner_ambiguous_count = sum(
        1
        for row in rows
        if not row["explicit_store_site"]
        and row["listing"]
        and row["site"]
        and owner_site_listing_counts[row["owner_key"]] > 1
    )
    owner_resolved_by_allocation_count = sum(
        1
        for row in rows
        if not row["explicit_store_site"]
        and row["listing"]
        and row["site"]
        and owner_site_listing_counts[row["owner_key"]] > 1
        and row["inferred_store_site"]
    )
    owner_unresolved_ambiguous_count = owner_ambiguous_count - owner_resolved_by_allocation_count
    store_site_explicit_count = sum(1 for row in rows if row["explicit_store_site"])
    store_site_missing_count = len(rows) - store_site_explicit_count
    store_site_inferred_count = sum(1 for row in rows if not row["explicit_store_site"] and row["inferred_store_site"])

    issue_rows = []
    for row in rows:
        owner_match_count = _owner_match_count(row, owner_keys, owner_site_listing_counts)
        if not row["site"] or not row["listing"] or not row["month"]:
            _append_issue(issue_rows, "\u4e3b\u952e\u7f3a\u5931", row)
        elif not row["final_store_site"]:
            if owner_match_count == 0:
                _append_issue(issue_rows, "\u672a\u5339\u914d\u8d1f\u8d23\u4eba\u914d\u7f6e", row)
            else:
                _append_issue(issue_rows, "\u7f3a\u5e97\u94fa\u7ad9\u70b9\u5f52\u5c5e", row)
        elif key_counts[row["key"]] > 1:
            _append_issue(issue_rows, "\u91cd\u590d\u552f\u4e00\u952e", row)
        elif owner_match_count == 0:
            _append_issue(issue_rows, "\u672a\u5339\u914d\u8d1f\u8d23\u4eba\u914d\u7f6e", row)

    return {
        "label": "\u9500\u552e\u9884\u4f30",
        "file_count": len(files),
        "total_rows": len(rows),
        "missing_key_count": missing_key_count,
        "duplicate_key_count": duplicate_key_count,
        "store_site_missing_count": store_site_missing_count,
        "store_site_explicit_count": store_site_explicit_count,
        "store_site_inferred_count": store_site_inferred_count,
        "owner_unmatched_count": owner_unmatched_count,
        "owner_ambiguous_count": owner_ambiguous_count,
        "owner_resolved_by_allocation_count": owner_resolved_by_allocation_count,
        "owner_unresolved_ambiguous_count": owner_unresolved_ambiguous_count,
        "issue_rows": issue_rows,
    }


def _forecast_store_site_review_rows(
    directory: Path,
    owner_site_listing_counts: Counter[tuple[str, str]],
    allocation_site_listing_store_sites: dict[tuple[str, str], set[str]],
    owner_candidates: dict[tuple[str, str], list[dict[str, str]]],
) -> list[dict[str, object]]:
    rows = []
    for file_path in _xlsx_files(directory):
        rows.extend(_read_sheet_rows(file_path, FORECAST_SHEET, [COL_SITE, COL_LISTING, COL_MONTH]))

    review_rows = []
    for row in rows:
        site = _clean(row.get(COL_SITE)).upper()
        listing = _clean(row.get(COL_LISTING))
        month = _month(row.get(COL_MONTH))
        owner_key = (site, listing)
        explicit_store_site = _forecast_store_site(row)
        inferred_store_site = _inferred_store_site(owner_key, allocation_site_listing_store_sites)
        if not site or not listing or not month:
            continue
        if explicit_store_site or inferred_store_site or owner_site_listing_counts[owner_key] <= 1:
            continue

        candidates = owner_candidates.get(owner_key, [])
        review_rows.append(
            {
                "file": row.get("file"),
                "row_number": row.get("row_number"),
                "site": site,
                "listing": listing,
                "forecast_month": month,
                "forecast_units": row.get(COL_FORECAST_UNITS),
                "candidate_store_sites": ", ".join(candidate["store_site"] for candidate in candidates),
                "owner_candidates": "\n".join(_format_owner_candidate(candidate) for candidate in candidates),
            }
        )
    return review_rows


def _forecast_store_site(row: dict[str, object]) -> str:
    for column in COL_STORE_SITE_ALIASES:
        value = _clean(row.get(column))
        if value:
            return value
    return _store_site(row.get(COL_STORE), row.get(COL_SITE))


def _forecast_maintenance_units(row: dict[str, object]) -> object:
    for column in (COL_FORECAST_MAINTENANCE_UNITS, COL_FORECAST_UNITS):
        if column in row:
            return row.get(column)
    return None


def _allocation_maintenance_sku(row: dict[str, object]) -> str:
    return _clean(row.get(COL_SKU_MAINTENANCE)) or _clean(row.get(COL_SKU))


def _optional_number(value: object, *, allow_percent: bool = False) -> tuple[float | None, bool]:
    text_value = _clean(value)
    if not text_value:
        return None, True
    try:
        if allow_percent and text_value.endswith("%"):
            return float(text_value[:-1].strip()) / 100, True
        return float(text_value), True
    except ValueError:
        return None, False


def _optional_int(value: object) -> tuple[int | None, bool]:
    number_value, is_valid = _optional_number(value)
    if not is_valid:
        return None, False
    return (None if number_value is None else int(number_value), True)


def _load_existing_allocation_rows(rows: list[dict[str, object]]) -> dict[tuple[str, str, str], dict[str, object]]:
    if not rows:
        return {}
    engine = get_engine()
    if engine is None:
        return {}

    existing: dict[tuple[str, str, str], dict[str, object]] = {}
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1 FROM amazon_sales_allocation LIMIT 1")).first()
            for row in rows:
                params = {
                    "store_site": row["store_site"],
                    "listing": row["listing"],
                    "msku": row["msku"],
                }
                before = conn.execute(
                    text(
                        """
                        SELECT id, store_site, listing, msku
                        FROM amazon_sales_allocation
                        WHERE store_site = :store_site
                          AND listing = :listing
                          AND msku = :msku
                        """
                    ),
                    params,
                ).mappings().first()
                if before:
                    key = (params["store_site"], params["listing"], params["msku"])
                    existing[key] = dict(before)
    except SQLAlchemyError:
        return {}
    return existing


def _load_existing_forecast_rows(rows: list[dict[str, object]]) -> dict[tuple[str, str, str], dict[str, object]]:
    if not rows:
        return {}
    engine = get_engine()
    if engine is None:
        return {}

    existing: dict[tuple[str, str, str], dict[str, object]] = {}
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1 FROM amazon_sales_forecast LIMIT 1")).first()
            for row in rows:
                params = {
                    "store_site": row["store_site"],
                    "listing": row["listing"],
                    "forecast_month": _forecast_month_date(row["forecast_month"]),
                }
                before = conn.execute(
                    text(
                        """
                        SELECT id, store_site, listing, forecast_month, site, forecast_units
                        FROM amazon_sales_forecast
                        WHERE store_site = :store_site
                          AND listing = :listing
                          AND forecast_month = :forecast_month
                        """
                    ),
                    params,
                ).mappings().first()
                if before:
                    key = (params["store_site"], params["listing"], params["forecast_month"])
                    existing[key] = dict(before)
    except SQLAlchemyError:
        return {}
    return existing


def _inferred_store_site(
    site_listing_key: tuple[str, str],
    allocation_site_listing_store_sites: dict[tuple[str, str], set[str]],
) -> str:
    store_sites = sorted(allocation_site_listing_store_sites.get(site_listing_key, set()))
    return store_sites[0] if len(store_sites) == 1 else ""


def _owner_match_count(
    row: dict[str, object],
    owner_keys: set[tuple[str, str]],
    owner_site_listing_counts: Counter[tuple[str, str]],
) -> int:
    if row["explicit_store_site"]:
        return 1 if (row["explicit_store_site"], row["listing"]) in owner_keys else 0
    return owner_site_listing_counts[row["owner_key"]]


def _read_sheet_rows(file_path: Path, sheet_name: str, required_columns: list[str]) -> list[dict[str, object]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        return []

    sheet = workbook[sheet_name]
    header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [_clean(value) for value in header_values]
    rows = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        data = {
            header: value
            for header, value in zip(headers, values)
            if header
        }
        if not any(_clean(data.get(column)) for column in required_columns):
            continue
        data["file"] = file_path.name
        data["row_number"] = row_number
        rows.append(data)

    workbook.close()
    return rows


def _read_workbook_rows(content: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [_clean(value) for value in header_values]
    rows = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        data = {
            header: value
            for header, value in zip(headers, values)
            if header
        }
        if not any(_clean(value) for value in data.values()):
            continue
        data["excel_row_number"] = row_number
        rows.append(data)

    workbook.close()
    return rows


def _load_owner_config_keys() -> tuple[set[tuple[str, str]], Counter[tuple[str, str]]]:
    engine = get_engine()
    if engine is None:
        return set(), Counter()

    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1 FROM amazon_listing_owner_config LIMIT 1")).first()
            rows = conn.execute(
                text(
                    """
                    SELECT store_site, listing
                    FROM amazon_listing_owner_config
                    WHERE store_site IS NOT NULL
                      AND TRIM(store_site) <> ''
                      AND listing IS NOT NULL
                      AND TRIM(listing) <> ''
                    """
                )
            ).mappings()
            owner_keys = {
                (_clean(row["store_site"]), _clean(row["listing"]))
                for row in rows
            }
    except SQLAlchemyError:
        return set(), Counter()

    site_listing_counts = Counter()
    for store_site, listing in owner_keys:
        site_listing_counts[(_site_from_store_site(store_site), listing)] += 1
    return owner_keys, site_listing_counts


def _load_owner_config_candidates() -> dict[tuple[str, str], list[dict[str, str]]]:
    engine = get_engine()
    if engine is None:
        return {}

    rows = []
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1 FROM amazon_listing_owner_config LIMIT 1")).first()
            rows = list(
                conn.execute(
                    text(
                        """
                        SELECT store_site, listing, owner, listing_status, project_group
                        FROM amazon_listing_owner_config
                        WHERE store_site IS NOT NULL
                          AND TRIM(store_site) <> ''
                          AND listing IS NOT NULL
                          AND TRIM(listing) <> ''
                        """
                    )
                ).mappings()
            )
    except SQLAlchemyError:
        try:
            with engine.connect() as conn:
                rows = list(
                    conn.execute(
                        text(
                            """
                            SELECT
                                store_site,
                                listing,
                                NULL AS owner,
                                NULL AS listing_status,
                                NULL AS project_group
                            FROM amazon_listing_owner_config
                            WHERE store_site IS NOT NULL
                              AND TRIM(store_site) <> ''
                              AND listing IS NOT NULL
                              AND TRIM(listing) <> ''
                            """
                        )
                    ).mappings()
                )
        except SQLAlchemyError:
            return {}

    candidates: dict[tuple[str, str], list[dict[str, str]]] = {}
    for row in rows:
        store_site = _clean(row["store_site"])
        listing = _clean(row["listing"])
        candidate = {
            "store_site": store_site,
            "owner": _clean(row.get("owner")),
            "listing_status": _clean(row.get("listing_status")),
            "project_group": _clean(row.get("project_group")),
        }
        candidates.setdefault((_site_from_store_site(store_site), listing), []).append(candidate)

    for values in candidates.values():
        values.sort(key=lambda item: item["store_site"])
    return candidates


def _format_owner_candidate(candidate: dict[str, str]) -> str:
    return " / ".join(
        [
            candidate.get("store_site") or "-",
            candidate.get("owner") or "-",
            candidate.get("listing_status") or "-",
            candidate.get("project_group") or "-",
        ]
    )


def _append_issue(issue_rows: list[dict[str, object]], issue: str, row: dict[str, object]) -> None:
    if len(issue_rows) >= ISSUE_LIMIT:
        return
    issue_rows.append(
        {
            "issue": issue,
            "file": row.get("file"),
            "row_number": row.get("row_number"),
            "key": row.get("display_key"),
        }
    )


def _xlsx_files(directory: Path) -> list[Path]:
    if not directory.exists():
        return []
    return sorted(path for path in directory.glob("*.xlsx") if not path.name.startswith("~$"))


def _store_site(store: object, site: object) -> str:
    store_text = _clean(store)
    site_text = _clean(site).upper()
    if not store_text or not site_text:
        return ""
    return f"{store_text}:{site_text}"


def _site_from_store_site(store_site: str) -> str:
    return _clean(store_site).split(":")[-1].upper()


def _month(value: object) -> str:
    if isinstance(value, datetime | date):
        return value.strftime("%Y-%m")
    text_value = _clean(value)
    if not text_value:
        return ""
    match = re.match(r"^(\d{4})[-/.\u5e74]?(\d{1,2})", text_value)
    if not match:
        return ""
    year = int(match.group(1))
    month = int(match.group(2))
    if month < 1 or month > 12:
        return ""
    return f"{year:04d}-{month:02d}"


def _forecast_month_date(month: str) -> str:
    month_text = _month(month)
    return f"{month_text}-01" if month_text else ""


def _number_or_none(value: object) -> object:
    text_value = _clean(value)
    if not text_value:
        return None
    try:
        return float(text_value)
    except ValueError:
        return None


def _number_or_zero(value: object) -> object:
    number_value = _number_or_none(value)
    return 0 if number_value is None else number_value


def _int_or_none(value: object) -> object:
    number_value = _number_or_none(value)
    return None if number_value is None else int(number_value)


def _complete_key(key: tuple[str, ...]) -> bool:
    return all(key)


def _clean(value: object) -> str:
    if value is None:
        return ""
    text_value = str(value).strip()
    return "" if text_value.lower() in {"nan", "nat", "none"} else text_value
