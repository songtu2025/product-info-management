from dataclasses import dataclass
from io import BytesIO
from math import ceil
from urllib.parse import urlencode

from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError

from app.core.db import get_engine
from app.shared.audit import build_change_set, record_operation_log
from app.modules.product_ops.readiness import build_purchase_readiness, flatten_purchase_readiness


PRODUCT_OPS_PAGE_SIZES = (50, 100, 200)
PRODUCT_OPS_EXPORT_COLUMNS = (
    ("store_site", "店铺站点"),
    ("listing", "Listing"),
    ("owner", "负责人"),
    ("listing_status", "Listing状态"),
    ("listing_maintainer", "Listing维护人"),
    ("project_group", "项目组"),
    ("product_brands", "品牌"),
    ("product_categories", "品类"),
    ("product_msku_count", "产品信息数"),
    ("allocation_msku_count", "销占比SKU数"),
    ("forecast_month_count", "销售预估月份数"),
    ("forecast_units_total", "预估销量合计"),
    ("first_forecast_month", "首个预估月份"),
    ("last_forecast_month", "最后预估月份"),
    ("data_status", "数据状态"),
    ("purchase_readiness_label", "采购准备度"),
    ("purchase_readiness_reasons", "准备度原因"),
)
SALES_ALLOCATION_EXPORT_COLUMNS = (
    ("store_site", "店铺站点"),
    ("site", "站点"),
    ("listing", "Listing"),
    ("msku", "MSKU"),
    ("sku", "SKU"),
    ("owner", "负责人"),
    ("style", "款式"),
    ("style_sales_ratio", "款式销占比"),
    ("sku_sales_ratio", "SKU销占比"),
    ("scale_position", "规模定位"),
    ("demand_position", "需求定位"),
    ("shipping_position", "发货定位"),
    ("stocking_position", "备货定位"),
    ("operation_min_order_days", "运营保底下单天数"),
    ("total_shipping_days", "总发货天数"),
    ("updated_at", "更新时间"),
)
SALES_ALLOCATION_MAINTENANCE_COLUMNS = (
    ("store_site", "店铺站点"),
    ("site", "站点"),
    ("owner", "负责人"),
    ("listing", "Listing"),
    ("style", "款式"),
    ("msku", "MSKU"),
    ("sku", "SKU"),
    ("scale_position", "规模定位"),
    ("style_sales_ratio", "款式销占比"),
    ("sku_sales_ratio", "SKU销占比"),
    ("demand_position", "需求定位"),
    ("shipping_position", "发货定位"),
    ("stocking_position", "备货定位"),
    ("operation_min_order_days", "运营保底下单天数"),
    ("total_shipping_days", "总发货天数"),
)
SALES_FORECAST_EXPORT_COLUMNS = (
    ("store_site", "店铺站点"),
    ("site", "站点"),
    ("listing", "Listing"),
    ("forecast_month", "月份"),
    ("forecast_units", "Listing月度预估销量"),
    ("updated_at", "更新时间"),
)
SALES_FORECAST_MAINTENANCE_COLUMNS = (
    ("store_site", "店铺站点"),
    ("site", "站点"),
    ("listing", "Listing"),
    ("forecast_month", "月份"),
    ("forecast_units", "Listing月度预估销量"),
)
PRODUCT_OPS_GAP_EXPORT_COLUMNS = (
    ("store_site", "店铺站点"),
    ("listing", "Listing"),
    ("owner", "负责人"),
    ("product_msku_count", "产品信息数"),
    ("allocation_msku_count", "销占比SKU数"),
    ("forecast_month_count", "销售预估月份数"),
    ("data_status", "状态"),
)


@dataclass(frozen=True)
class ProductOpsFilters:
    q: str | None = None
    store_site: str | None = None
    listing: str | None = None
    brand: str | None = None
    data_status: str | None = None
    page: int = 1
    page_size: int = 50


@dataclass(frozen=True)
class SalesAllocationFilters:
    q: str | None = None
    store_site: str | None = None
    listing: str | None = None
    ratio_status: str | None = None
    page: int = 1
    page_size: int = 50


@dataclass(frozen=True)
class SalesForecastFilters:
    q: str | None = None
    store_site: str | None = None
    site: str | None = None
    listing: str | None = None
    forecast_status: str | None = None
    page: int = 1
    page_size: int = 50


def list_product_ops_rows(filters: ProductOpsFilters) -> dict[str, object]:
    filters = normalize_product_ops_filters(filters)
    rows, total = _query_product_ops_rows(filters, paginate=True)
    return _page(rows, total, filters.page, filters.page_size)


def list_product_ops_rows_for_export(filters: ProductOpsFilters) -> list[dict[str, object]]:
    filters = normalize_product_ops_filters(filters)
    rows, _ = _query_product_ops_rows(filters, paginate=False)
    return rows


def _query_product_ops_rows(filters: ProductOpsFilters, *, paginate: bool) -> tuple[list[dict[str, object]], int]:
    engine = get_engine()
    if engine is None:
        return [], 0

    with engine.connect() as conn:
        if not _table_available(conn, "amazon_listing_owner_config"):
            return [], 0
        has_product = _table_available(conn, "amazon_product_info")
        has_allocation = _table_available(conn, "amazon_sales_allocation")
        has_forecast = _table_available(conn, "amazon_sales_forecast")

        where_sql, params = _build_product_ops_where(filters, has_product, has_allocation, has_forecast)
        joins = _product_ops_joins(has_product, has_allocation, has_forecast)
        columns = _product_ops_columns(has_product, has_allocation, has_forecast)
        limit_sql = ""
        if paginate:
            offset = (filters.page - 1) * filters.page_size
            params.update({"limit": filters.page_size, "offset": offset})
            limit_sql = "LIMIT :limit OFFSET :offset"

        total = conn.execute(
            text(
                f"""
                SELECT COUNT(*)
                FROM amazon_listing_owner_config lo
                {joins}
                {where_sql}
                """
            ),
            params,
        ).scalar_one()

        rows = [
            _decorate_product_ops_row(dict(row))
            for row in conn.execute(
                text(
                    f"""
                    SELECT
                        {columns}
                    FROM amazon_listing_owner_config lo
                    {joins}
                    {where_sql}
                    ORDER BY lo.store_site, lo.listing
                    {limit_sql}
                    """
                ),
                params,
            ).mappings()
        ]

    return rows, total


def list_sales_allocations(filters: SalesAllocationFilters) -> dict[str, object]:
    filters = normalize_sales_allocation_filters(filters)
    rows, total = _query_sales_allocation_rows(filters, paginate=True)
    return _page(rows, total, filters.page, filters.page_size)


def list_sales_allocations_for_export(filters: SalesAllocationFilters) -> list[dict[str, object]]:
    filters = normalize_sales_allocation_filters(filters)
    rows, _ = _query_sales_allocation_rows(filters, paginate=False)
    return rows


def bulk_update_sales_allocations(
    row_ids: list[int],
    updates: dict[str, object],
    changed_by: str = "system",
) -> dict[str, int]:
    allowed_fields = {"style_sales_ratio", "sku_sales_ratio", "stocking_position"}
    clean_ids = list(dict.fromkeys(row_id for row_id in row_ids if row_id > 0))
    clean_updates = {
        field: _clean(value) if field == "stocking_position" else value
        for field, value in updates.items()
        if field in allowed_fields and value not in (None, "")
    }
    if not clean_ids or not clean_updates:
        return {"updated": 0, "skipped": len(clean_ids), "requested": len(clean_ids)}

    engine = get_engine()
    if engine is None:
        return {"updated": 0, "skipped": len(clean_ids), "requested": len(clean_ids)}

    updated = 0
    skipped = 0
    selected_fields = tuple(clean_updates.keys())
    set_sql = ", ".join(f"{field} = :{field}" for field in selected_fields)
    with engine.begin() as conn:
        for row_id in clean_ids:
            before = conn.execute(
                text(
                    f"""
                    SELECT id, {", ".join(selected_fields)}
                    FROM amazon_sales_allocation
                    WHERE id = :id
                    """
                ),
                {"id": row_id},
            ).mappings().first()
            if before is None:
                skipped += 1
                continue
            changes = build_change_set(dict(before), clean_updates)
            if not changes:
                skipped += 1
                continue
            conn.execute(
                text(
                    f"""
                    UPDATE amazon_sales_allocation
                    SET {set_sql}
                    WHERE id = :id
                    """
                ),
                {**clean_updates, "id": row_id},
            )
            record_operation_log(
                conn,
                table_name="amazon_sales_allocation",
                record_id=row_id,
                operation_type="BULK_UPDATE",
                change_data=changes,
                changed_by=changed_by,
            )
            updated += 1
    return {"updated": updated, "skipped": skipped, "requested": len(clean_ids)}


def _query_sales_allocation_rows(filters: SalesAllocationFilters, *, paginate: bool) -> tuple[list[dict[str, object]], int]:
    engine = get_engine()
    if engine is None:
        return [], 0

    with engine.connect() as conn:
        if not _table_available(conn, "amazon_sales_allocation"):
            return [], 0
        where_sql, params = _build_sales_allocation_where(filters)
        limit_sql = ""
        if paginate:
            offset = (filters.page - 1) * filters.page_size
            params.update({"limit": filters.page_size, "offset": offset})
            limit_sql = "LIMIT :limit OFFSET :offset"

        total = conn.execute(text(f"SELECT COUNT(*) FROM amazon_sales_allocation {where_sql}"), params).scalar_one()
        rows = [
            dict(row)
            for row in conn.execute(
                text(
                    f"""
                    SELECT
                        id, store_site, site, owner, listing, style, msku, sku,
                        scale_position, style_sales_ratio, sku_sales_ratio,
                        demand_position, shipping_position, stocking_position,
                        operation_min_order_days, total_shipping_days, updated_at
                    FROM amazon_sales_allocation
                    {where_sql}
                    ORDER BY store_site, listing, msku
                    {limit_sql}
                    """
                ),
                params,
            ).mappings()
        ]

    return rows, total


def list_sales_forecasts(filters: SalesForecastFilters) -> dict[str, object]:
    filters = normalize_sales_forecast_filters(filters)
    rows, total = _query_sales_forecast_rows(filters, paginate=True)
    return _page(rows, total, filters.page, filters.page_size)


def list_sales_forecasts_for_export(filters: SalesForecastFilters) -> list[dict[str, object]]:
    filters = normalize_sales_forecast_filters(filters)
    rows, _ = _query_sales_forecast_rows(filters, paginate=False)
    return rows


def bulk_update_sales_forecasts(
    row_ids: list[int],
    forecast_units: int | float,
    changed_by: str = "system",
) -> dict[str, int]:
    clean_ids = list(dict.fromkeys(row_id for row_id in row_ids if row_id > 0))
    if not clean_ids:
        return {"updated": 0, "skipped": 0, "requested": 0}

    engine = get_engine()
    if engine is None:
        return {"updated": 0, "skipped": len(clean_ids), "requested": len(clean_ids)}

    updated = 0
    skipped = 0
    with engine.begin() as conn:
        for row_id in clean_ids:
            before = conn.execute(
                text(
                    """
                    SELECT id, forecast_units
                    FROM amazon_sales_forecast
                    WHERE id = :id
                    """
                ),
                {"id": row_id},
            ).mappings().first()
            if before is None:
                skipped += 1
                continue
            changes = build_change_set(dict(before), {"forecast_units": forecast_units})
            if not changes:
                skipped += 1
                continue
            conn.execute(
                text(
                    """
                    UPDATE amazon_sales_forecast
                    SET forecast_units = :forecast_units
                    WHERE id = :id
                    """
                ),
                {"id": row_id, "forecast_units": forecast_units},
            )
            record_operation_log(
                conn,
                table_name="amazon_sales_forecast",
                record_id=row_id,
                operation_type="BULK_UPDATE",
                change_data=changes,
                changed_by=changed_by,
            )
            updated += 1
    return {"updated": updated, "skipped": skipped, "requested": len(clean_ids)}


def upsert_sales_forecast(payload: dict[str, object], changed_by: str = "system") -> dict[str, object]:
    store_site = _clean(payload.get("store_site"))
    listing = _clean(payload.get("listing"))
    forecast_month = _forecast_month_start(payload.get("forecast_month"))
    forecast_units = payload.get("forecast_units")
    site = _clean(payload.get("site")).upper() or _site_from_store_site(store_site)
    if not store_site or not listing or not forecast_month or forecast_units in (None, ""):
        return {"action": "skipped", "id": None}

    engine = get_engine()
    if engine is None:
        return {"action": "skipped", "id": None}

    params = {
        "store_site": store_site,
        "site": site,
        "listing": listing,
        "forecast_month": forecast_month,
        "forecast_units": forecast_units,
    }
    with engine.begin() as conn:
        before = conn.execute(
            text(
                """
                SELECT id, site, forecast_units
                FROM amazon_sales_forecast
                WHERE store_site = :store_site
                  AND listing = :listing
                  AND forecast_month = :forecast_month
                """
            ),
            params,
        ).mappings().first()
        if before is None:
            conn.execute(
                text(
                    """
                    INSERT INTO amazon_sales_forecast (
                        store_site, site, listing, forecast_month, forecast_units
                    )
                    VALUES (
                        :store_site, :site, :listing, :forecast_month, :forecast_units
                    )
                    """
                ),
                params,
            )
            forecast_id = conn.execute(
                text(
                    """
                    SELECT id
                    FROM amazon_sales_forecast
                    WHERE store_site = :store_site
                      AND listing = :listing
                      AND forecast_month = :forecast_month
                    """
                ),
                params,
            ).scalar_one()
            record_operation_log(
                conn,
                table_name="amazon_sales_forecast",
                record_id=forecast_id,
                operation_type="MANUAL_INSERT",
                change_data={field: {"old": None, "new": value} for field, value in params.items()},
                changed_by=changed_by,
            )
            return {"action": "inserted", "id": forecast_id}

        changes = build_change_set(dict(before), {"site": site, "forecast_units": forecast_units})
        if not changes:
            return {"action": "skipped", "id": before["id"]}
        conn.execute(
            text(
                """
                UPDATE amazon_sales_forecast
                SET site = :site,
                    forecast_units = :forecast_units
                WHERE id = :forecast_id
                """
            ),
            {**params, "forecast_id": before["id"]},
        )
        record_operation_log(
            conn,
            table_name="amazon_sales_forecast",
            record_id=before["id"],
            operation_type="MANUAL_UPDATE",
            change_data=changes,
            changed_by=changed_by,
        )
        return {"action": "updated", "id": before["id"]}


def get_listing_profile(store_site: str, listing: str) -> dict[str, object]:
    store_site_value = _clean(store_site) or ""
    listing_value = _clean(listing) or ""
    overview_rows = list_product_ops_rows_for_export(
        ProductOpsFilters(store_site=store_site_value, listing=listing_value)
    )
    allocation_rows = list_sales_allocations_for_export(
        SalesAllocationFilters(store_site=store_site_value, listing=listing_value)
    )
    forecast_rows = list_sales_forecasts_for_export(
        SalesForecastFilters(store_site=store_site_value, listing=listing_value)
    )
    overview = overview_rows[0] if overview_rows else _empty_listing_profile_overview(
        store_site_value,
        listing_value,
        allocation_rows,
        forecast_rows,
    )
    health = _listing_profile_health(overview, allocation_rows, forecast_rows, store_site_value, listing_value)
    purchase_readiness = build_purchase_readiness(overview, allocation_rows, forecast_rows)
    return {
        "store_site": store_site_value,
        "listing": listing_value,
        "overview": overview,
        "issue_labels": _listing_profile_issue_labels(overview, allocation_rows, forecast_rows),
        "health_summary": health["summary"],
        "health_items": health["items"],
        "purchase_readiness": purchase_readiness,
        "allocation_rows": allocation_rows,
        "forecast_rows": forecast_rows,
    }


def _query_sales_forecast_rows(filters: SalesForecastFilters, *, paginate: bool) -> tuple[list[dict[str, object]], int]:
    engine = get_engine()
    if engine is None:
        return [], 0

    with engine.connect() as conn:
        if not _table_available(conn, "amazon_sales_forecast"):
            return [], 0
        where_sql, params = _build_sales_forecast_where(filters)
        limit_sql = ""
        if paginate:
            offset = (filters.page - 1) * filters.page_size
            params.update({"limit": filters.page_size, "offset": offset})
            limit_sql = "LIMIT :limit OFFSET :offset"

        total = conn.execute(text(f"SELECT COUNT(*) FROM amazon_sales_forecast {where_sql}"), params).scalar_one()
        rows = [
            dict(row)
            for row in conn.execute(
                text(
                    f"""
                    SELECT id, store_site, site, listing, forecast_month, forecast_units, updated_at
                    FROM amazon_sales_forecast
                    {where_sql}
                    ORDER BY store_site, listing, forecast_month
                    {limit_sql}
                    """
                ),
                params,
            ).mappings()
        ]

    return rows, total


def list_product_ops_gaps(limit: int | None = 20) -> dict[str, object]:
    rows = list_product_ops_rows_for_export(ProductOpsFilters())

    groups = [
        {"key": "missing_product", "label": "缺产品信息", "rows": []},
        {"key": "missing_allocation", "label": "缺销占比", "rows": []},
        {"key": "missing_forecast", "label": "缺销售预估", "rows": []},
        {"key": "zero_allocation_ratio", "label": "销占比为0", "rows": []},
        {"key": "zero_forecast_units", "label": "销售预估为0", "rows": []},
    ]
    group_map = {group["key"]: group for group in groups}

    for row in rows:
        if not row.get("product_msku_count"):
            group_map["missing_product"]["rows"].append(row)
        if not row.get("allocation_msku_count"):
            group_map["missing_allocation"]["rows"].append(row)
        if not row.get("forecast_month_count"):
            group_map["missing_forecast"]["rows"].append(row)

    zero_allocation_rows = _zero_allocation_ratio_rows(None)
    group_map["zero_allocation_ratio"]["rows"] = zero_allocation_rows
    zero_forecast_rows = _zero_forecast_units_rows(None)
    group_map["zero_forecast_units"]["rows"] = zero_forecast_rows
    summary = {group["key"]: len(group["rows"]) for group in groups}
    if limit is not None:
        for group in groups:
            group["rows"] = group["rows"][:limit]
    return {"summary": summary, "groups": groups}


def export_product_ops_rows_to_xlsx(filters: ProductOpsFilters) -> bytes:
    return _rows_to_workbook_bytes(
        "经营总览",
        PRODUCT_OPS_EXPORT_COLUMNS,
        list_product_ops_rows_for_export(filters),
    )


def export_sales_allocations_to_xlsx(filters: SalesAllocationFilters) -> bytes:
    return _rows_to_workbook_bytes(
        "销占比",
        SALES_ALLOCATION_EXPORT_COLUMNS,
        list_sales_allocations_for_export(filters),
        percent_fields={"style_sales_ratio", "sku_sales_ratio"},
    )


def export_sales_allocation_maintenance_template(filters: SalesAllocationFilters) -> bytes:
    return _rows_to_workbook_bytes(
        "销占比维护",
        SALES_ALLOCATION_MAINTENANCE_COLUMNS,
        list_sales_allocations_for_export(filters),
        percent_fields={"style_sales_ratio", "sku_sales_ratio"},
    )


def export_sales_forecasts_to_xlsx(filters: SalesForecastFilters) -> bytes:
    return _rows_to_workbook_bytes(
        "销售预估",
        SALES_FORECAST_EXPORT_COLUMNS,
        list_sales_forecasts_for_export(filters),
    )


def export_sales_forecast_maintenance_template(filters: SalesForecastFilters) -> bytes:
    return _rows_to_workbook_bytes(
        "销售预估维护",
        SALES_FORECAST_MAINTENANCE_COLUMNS,
        list_sales_forecasts_for_export(filters),
    )


def export_product_ops_gaps_to_xlsx() -> bytes:
    gaps = list_product_ops_gaps(limit=None)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "缺口汇总"
    summary_sheet.append(["类型", "数量"])
    for group in gaps["groups"]:
        summary_sheet.append([group["label"], gaps["summary"].get(group["key"], 0)])

    for group in gaps["groups"]:
        sheet = workbook.create_sheet(group["label"][:31])
        sheet.append([header for _, header in PRODUCT_OPS_GAP_EXPORT_COLUMNS])
        for row in group["rows"]:
            sheet.append([row.get(field) for field, _ in PRODUCT_OPS_GAP_EXPORT_COLUMNS])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def normalize_product_ops_filters(filters: ProductOpsFilters) -> ProductOpsFilters:
    return ProductOpsFilters(
        q=_clean(filters.q),
        store_site=_clean(filters.store_site),
        listing=_clean(filters.listing),
        brand=_clean(filters.brand),
        data_status=_clean(filters.data_status),
        page=max(int(filters.page or 1), 1),
        page_size=_normalize_page_size(filters.page_size),
    )


def normalize_sales_allocation_filters(filters: SalesAllocationFilters) -> SalesAllocationFilters:
    return SalesAllocationFilters(
        q=_clean(filters.q),
        store_site=_clean(filters.store_site),
        listing=_clean(filters.listing),
        ratio_status=_clean(filters.ratio_status),
        page=max(int(filters.page or 1), 1),
        page_size=_normalize_page_size(filters.page_size),
    )


def normalize_sales_forecast_filters(filters: SalesForecastFilters) -> SalesForecastFilters:
    return SalesForecastFilters(
        q=_clean(filters.q),
        store_site=_clean(filters.store_site),
        site=_clean(filters.site),
        listing=_clean(filters.listing),
        forecast_status=_clean(filters.forecast_status),
        page=max(int(filters.page or 1), 1),
        page_size=_normalize_page_size(filters.page_size),
    )


def _product_ops_columns(has_product: bool, has_allocation: bool, has_forecast: bool) -> str:
    product_columns = (
        """
            COALESCE(pi.product_msku_count, 0) AS product_msku_count,
            pi.product_brands,
            pi.product_categories
        """
        if has_product
        else """
            0 AS product_msku_count,
            NULL AS product_brands,
            NULL AS product_categories
        """
    )

    allocation_columns = (
        """
            COALESCE(sa.allocation_msku_count, 0) AS allocation_msku_count,
            COALESCE(sa.zero_allocation_ratio_count, 0) AS zero_allocation_ratio_count,
            sa.allocation_updated_at
        """
        if has_allocation
        else """
            0 AS allocation_msku_count,
            0 AS zero_allocation_ratio_count,
            NULL AS allocation_updated_at
        """
    )

    forecast_columns = (
        """
            COALESCE(sf.forecast_month_count, 0) AS forecast_month_count,
            COALESCE(sf.forecast_units_total, 0) AS forecast_units_total,
            COALESCE(sf.zero_forecast_units_count, 0) AS zero_forecast_units_count,
            sf.first_forecast_month,
            sf.last_forecast_month
        """
        if has_forecast
        else """
            0 AS forecast_month_count,
            0 AS forecast_units_total,
            0 AS zero_forecast_units_count,
            NULL AS first_forecast_month,
            NULL AS last_forecast_month
        """
    )

    return f"""
        lo.id AS owner_config_id,
        lo.store_site,
        lo.listing,
        lo.owner,
        lo.listing_status,
        lo.listing_maintainer,
        lo.project_group,
        {product_columns},
        {allocation_columns},
        {forecast_columns}
    """


def _product_ops_joins(has_product: bool, has_allocation: bool, has_forecast: bool) -> str:
    joins = []
    if has_product:
        joins.append(
            """
            LEFT JOIN (
                SELECT
                    store_site,
                    listing,
                    COUNT(*) AS product_msku_count,
                    GROUP_CONCAT(DISTINCT brand) AS product_brands,
                    GROUP_CONCAT(DISTINCT category_a) AS product_categories
                FROM amazon_product_info
                WHERE listing IS NOT NULL
                  AND TRIM(listing) <> ''
                GROUP BY store_site, listing
            ) pi
              ON pi.store_site = lo.store_site
             AND pi.listing = lo.listing
            """
        )
    if has_allocation:
        joins.append(
            """
            LEFT JOIN (
                SELECT
                    store_site,
                    listing,
                    COUNT(*) AS allocation_msku_count,
                    SUM(
                        CASE
                            WHEN COALESCE(style_sales_ratio, 0) = 0
                              OR COALESCE(sku_sales_ratio, 0) = 0
                            THEN 1
                            ELSE 0
                        END
                    ) AS zero_allocation_ratio_count,
                    MAX(updated_at) AS allocation_updated_at
                FROM amazon_sales_allocation
                GROUP BY store_site, listing
            ) sa
              ON sa.store_site = lo.store_site
             AND sa.listing = lo.listing
            """
        )
    if has_forecast:
        joins.append(
            """
            LEFT JOIN (
                SELECT
                    store_site,
                    listing,
                    COUNT(*) AS forecast_month_count,
                    SUM(forecast_units) AS forecast_units_total,
                    SUM(
                        CASE
                            WHEN COALESCE(forecast_units, 0) <= 0
                            THEN 1
                            ELSE 0
                        END
                    ) AS zero_forecast_units_count,
                    MIN(forecast_month) AS first_forecast_month,
                    MAX(forecast_month) AS last_forecast_month
                FROM amazon_sales_forecast
                GROUP BY store_site, listing
            ) sf
              ON sf.store_site = lo.store_site
             AND sf.listing = lo.listing
            """
        )
    return "\n".join(joins)


def _build_product_ops_where(
    filters: ProductOpsFilters,
    has_product: bool,
    has_allocation: bool,
    has_forecast: bool,
) -> tuple[str, dict[str, object]]:
    clauses = []
    params: dict[str, object] = {}
    if filters.q:
        q_columns = [
            "lo.store_site LIKE :q",
            "lo.listing LIKE :q",
            "lo.owner LIKE :q",
            "lo.project_group LIKE :q",
        ]
        if has_product:
            q_columns.extend(
                [
                    "pi.product_brands LIKE :q",
                    "pi.product_categories LIKE :q",
                ]
            )
        clauses.append(
            """
            (
                """
            + "\n                OR ".join(q_columns)
            + """
            )
            """
        )
        params["q"] = f"%{filters.q}%"
    for column, key, value in (
        ("lo.store_site", "store_site", filters.store_site),
        ("lo.listing", "listing", filters.listing),
    ):
        if value:
            clauses.append(f"{column} = :{key}")
            params[key] = value
    if filters.brand:
        if has_product:
            clauses.append("pi.product_brands LIKE :brand")
            params["brand"] = f"%{filters.brand}%"
        else:
            clauses.append("1 = 0")

    if filters.data_status == "missing_allocation":
        clauses.append("COALESCE(sa.allocation_msku_count, 0) = 0" if has_allocation else "1 = 1")
    elif filters.data_status == "missing_forecast":
        clauses.append("COALESCE(sf.forecast_month_count, 0) = 0" if has_forecast else "1 = 1")
    elif filters.data_status == "missing_product":
        clauses.append("COALESCE(pi.product_msku_count, 0) = 0" if has_product else "1 = 1")
    elif filters.data_status == "complete":
        clauses.append("COALESCE(pi.product_msku_count, 0) > 0" if has_product else "1 = 0")
        clauses.append("COALESCE(sa.allocation_msku_count, 0) > 0" if has_allocation else "1 = 0")
        clauses.append("COALESCE(sf.forecast_month_count, 0) > 0" if has_forecast else "1 = 0")

    return ("WHERE " + " AND ".join(clauses), params) if clauses else ("", params)


def _build_sales_allocation_where(filters: SalesAllocationFilters) -> tuple[str, dict[str, object]]:
    clauses = []
    params: dict[str, object] = {}
    if filters.q:
        clauses.append(
            """
            (
                store_site LIKE :q OR listing LIKE :q OR msku LIKE :q OR sku LIKE :q
                OR owner LIKE :q OR style LIKE :q
            )
            """
        )
        params["q"] = f"%{filters.q}%"
    for column, key, value in (
        ("store_site", "store_site", filters.store_site),
        ("listing", "listing", filters.listing),
    ):
        if value:
            clauses.append(f"{column} = :{key}")
            params[key] = value
    if filters.ratio_status == "zero":
        clauses.append("(COALESCE(style_sales_ratio, 0) = 0 OR COALESCE(sku_sales_ratio, 0) = 0)")
    return ("WHERE " + " AND ".join(clauses), params) if clauses else ("", params)


def _build_sales_forecast_where(filters: SalesForecastFilters) -> tuple[str, dict[str, object]]:
    clauses = []
    params: dict[str, object] = {}
    if filters.q:
        clauses.append("(store_site LIKE :q OR site LIKE :q OR listing LIKE :q)")
        params["q"] = f"%{filters.q}%"
    for column, key, value in (
        ("store_site", "store_site", filters.store_site),
        ("site", "site", filters.site),
        ("listing", "listing", filters.listing),
    ):
        if value:
            clauses.append(f"{column} = :{key}")
            params[key] = value
    if filters.forecast_status == "zero":
        clauses.append("COALESCE(forecast_units, 0) <= 0")
    return ("WHERE " + " AND ".join(clauses), params) if clauses else ("", params)


def _decorate_product_ops_row(row: dict[str, object]) -> dict[str, object]:
    missing = []
    if not row.get("allocation_msku_count"):
        missing.append("缺销占比")
    if not row.get("forecast_month_count"):
        missing.append("缺销售预估")
    if not row.get("product_msku_count"):
        missing.append("缺产品信息")
    row["data_status"] = "/".join(missing) if missing else "正常"
    flatten_purchase_readiness(row, build_purchase_readiness(row))
    return row


def _empty_listing_profile_overview(
    store_site: str,
    listing: str,
    allocation_rows: list[dict[str, object]],
    forecast_rows: list[dict[str, object]],
) -> dict[str, object]:
    return {
        "owner_config_id": None,
        "store_site": store_site,
        "listing": listing,
        "owner": None,
        "listing_status": None,
        "listing_maintainer": None,
        "project_group": None,
        "product_msku_count": 0,
        "product_brands": None,
        "product_categories": None,
        "allocation_msku_count": len(allocation_rows),
        "zero_allocation_ratio_count": sum(
            1
            for row in allocation_rows
            if _is_zero_ratio(row.get("style_sales_ratio")) or _is_zero_ratio(row.get("sku_sales_ratio"))
        ),
        "allocation_updated_at": None,
        "forecast_month_count": len(forecast_rows),
        "forecast_units_total": sum(row.get("forecast_units") or 0 for row in forecast_rows),
        "zero_forecast_units_count": sum(1 for row in forecast_rows if _is_zero_forecast_units(row.get("forecast_units"))),
        "first_forecast_month": forecast_rows[0]["forecast_month"] if forecast_rows else None,
        "last_forecast_month": forecast_rows[-1]["forecast_month"] if forecast_rows else None,
        "data_status": "缺负责人配置",
    }


def _listing_profile_issue_labels(
    overview: dict[str, object],
    allocation_rows: list[dict[str, object]],
    forecast_rows: list[dict[str, object]],
) -> list[str]:
    labels = [
        label
        for label in str(overview.get("data_status") or "").split("/")
        if label and label != "正常"
    ]
    if any(
        _is_zero_ratio(row.get("style_sales_ratio")) or _is_zero_ratio(row.get("sku_sales_ratio"))
        for row in allocation_rows
    ):
        labels.append("销占比为0")
    if any(_is_zero_forecast_units(row.get("forecast_units")) for row in forecast_rows):
        labels.append("销售预估为0")
    return labels or ["正常"]


def _listing_profile_health(
    overview: dict[str, object],
    allocation_rows: list[dict[str, object]],
    forecast_rows: list[dict[str, object]],
    store_site: str,
    listing: str,
) -> dict[str, object]:
    product_count = int(overview.get("product_msku_count") or 0)
    allocation_count = len(allocation_rows)
    forecast_count = len(forecast_rows) or int(overview.get("forecast_month_count") or 0)
    zero_allocation_count = sum(
        1
        for row in allocation_rows
        if _is_zero_ratio(row.get("style_sales_ratio")) or _is_zero_ratio(row.get("sku_sales_ratio"))
    )
    zero_forecast_count = sum(1 for row in forecast_rows if _is_zero_forecast_units(row.get("forecast_units")))
    has_owner_config = bool(
        overview.get("owner_config_id")
        or overview.get("owner")
        or overview.get("listing_status")
        or overview.get("listing_maintainer")
    )

    items = [
        _health_item(
            key="product_info",
            label="产品信息",
            status="normal" if product_count else "missing",
            status_label="正常" if product_count else "缺失",
            message=f"已关联 {product_count} 个产品信息" if product_count else "未找到产品信息",
            action_label="查看产品信息" if product_count else "维护产品信息",
            action_url=_listing_profile_url("/", store_site, listing),
        ),
        _health_item(
            key="owner_config",
            label="负责人配置",
            status="normal" if has_owner_config else "missing",
            status_label="正常" if has_owner_config else "缺失",
            message=f"负责人：{overview.get('owner') or '-'}" if has_owner_config else "缺负责人配置",
            action_label="查看负责人配置" if has_owner_config else "创建负责人配置",
            action_url=(
                _query_url("/listing-owners", {"store_site": store_site, "q": listing})
                if has_owner_config
                else _listing_profile_url("/listing-owners/new", store_site, listing)
            ),
        ),
        _health_item(
            key="sales_allocation",
            label="销占比",
            status="missing" if not allocation_count else "warning" if zero_allocation_count else "normal",
            status_label="缺失" if not allocation_count else "异常" if zero_allocation_count else "正常",
            message=(
                "未维护销占比"
                if not allocation_count
                else f"存在 {zero_allocation_count} 条销占比为0数据"
                if zero_allocation_count
                else f"已维护 {allocation_count} 个销占比 SKU"
            ),
            action_label="维护销占比",
            action_url=_listing_profile_url("/product-ops/allocations", store_site, listing),
        ),
        _health_item(
            key="sales_forecast",
            label="销售预估",
            status="missing" if not forecast_count else "warning" if zero_forecast_count else "normal",
            status_label="缺失" if not forecast_count else "异常" if zero_forecast_count else "正常",
            message=(
                "未维护销售预估"
                if not forecast_count
                else f"存在 {zero_forecast_count} 个月销售预估为0"
                if zero_forecast_count
                else f"已维护 {forecast_count} 个月销售预估"
            ),
            action_label="维护销售预估",
            action_url=_listing_profile_url("/product-ops/forecasts", store_site, listing),
        ),
    ]
    issue_count = sum(1 for item in items if item["status"] != "normal")
    summary = {
        "status": "normal" if issue_count == 0 else "warning",
        "status_label": "正常" if issue_count == 0 else "需维护",
        "message": "4 项均正常" if issue_count == 0 else f"{issue_count} 项待维护",
    }
    return {"summary": summary, "items": items}


def _health_item(
    *,
    key: str,
    label: str,
    status: str,
    status_label: str,
    message: str,
    action_label: str,
    action_url: str,
) -> dict[str, str]:
    return {
        "key": key,
        "label": label,
        "status": status,
        "status_label": status_label,
        "message": message,
        "action_label": action_label,
        "action_url": action_url,
    }


def _listing_profile_url(path: str, store_site: str, listing: str) -> str:
    return _query_url(path, {"store_site": store_site, "listing": listing})


def _query_url(path: str, params: dict[str, object]) -> str:
    query = urlencode({key: value for key, value in params.items() if value not in (None, "")})
    return f"{path}?{query}" if query else path


def _is_zero_ratio(value: object) -> bool:
    if value is None:
        return True
    try:
        return float(value) == 0
    except (TypeError, ValueError):
        return False


def _is_zero_forecast_units(value: object) -> bool:
    if value is None:
        return True
    try:
        return float(value) <= 0
    except (TypeError, ValueError):
        return False


def _forecast_month_start(value: object) -> str:
    month = _clean(value)
    if not month:
        return ""
    return f"{month[:7]}-01" if len(month) >= 7 else month


def _site_from_store_site(store_site: str) -> str:
    return store_site.split(":", 1)[1].upper() if ":" in store_site else ""


def _zero_allocation_ratio_rows(limit: int | None) -> list[dict[str, object]]:
    engine = get_engine()
    if engine is None:
        return []

    with engine.connect() as conn:
        if not _table_available(conn, "amazon_sales_allocation"):
            return []
        limit_sql = "LIMIT :limit" if limit is not None else ""
        params = {"limit": limit} if limit is not None else {}
        rows = [
            dict(row)
            for row in conn.execute(
                text(
                    f"""
                    SELECT
                        store_site,
                        listing,
                        MAX(owner) AS owner,
                        COUNT(*) AS allocation_msku_count,
                        0 AS product_msku_count,
                        0 AS forecast_month_count,
                        '销占比为0' AS data_status
                    FROM amazon_sales_allocation
                    WHERE COALESCE(style_sales_ratio, 0) = 0
                       OR COALESCE(sku_sales_ratio, 0) = 0
                    GROUP BY store_site, listing
                    ORDER BY store_site, listing
                    {limit_sql}
                    """
                ),
                params,
            ).mappings()
        ]
    return rows


def _zero_forecast_units_rows(limit: int | None) -> list[dict[str, object]]:
    engine = get_engine()
    if engine is None:
        return []

    with engine.connect() as conn:
        if not _table_available(conn, "amazon_sales_forecast"):
            return []
        owner_join = ""
        owner_column = "NULL AS owner"
        if _table_available(conn, "amazon_listing_owner_config"):
            owner_join = """
                LEFT JOIN amazon_listing_owner_config lo
                  ON lo.store_site = sf.store_site
                 AND lo.listing = sf.listing
            """
            owner_column = "MAX(lo.owner) AS owner"
        limit_sql = "LIMIT :limit" if limit is not None else ""
        params = {"limit": limit} if limit is not None else {}
        rows = [
            dict(row)
            for row in conn.execute(
                text(
                    f"""
                    SELECT
                        sf.store_site,
                        sf.listing,
                        {owner_column},
                        0 AS product_msku_count,
                        0 AS allocation_msku_count,
                        COUNT(*) AS forecast_month_count,
                        '销售预估为0' AS data_status
                    FROM amazon_sales_forecast sf
                    {owner_join}
                    WHERE COALESCE(sf.forecast_units, 0) <= 0
                    GROUP BY sf.store_site, sf.listing
                    ORDER BY sf.store_site, sf.listing
                    {limit_sql}
                    """
                ),
                params,
            ).mappings()
        ]
    return rows


def _rows_to_workbook_bytes(
    sheet_title: str,
    columns: tuple[tuple[str, str], ...],
    rows: list[dict[str, object]],
    percent_fields: set[str] | None = None,
) -> bytes:
    percent_fields = percent_fields or set()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title[:31]
    sheet.append([header for _, header in columns])
    for row in rows:
        sheet.append([row.get(field) for field, _ in columns])
        current_row = sheet.max_row
        for column_index, (field, _) in enumerate(columns, start=1):
            if field in percent_fields and row.get(field) is not None:
                sheet.cell(row=current_row, column=column_index).number_format = "0.00%"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _table_available(conn, table_name: str) -> bool:
    try:
        conn.execute(text(f"SELECT 1 FROM {table_name} LIMIT 1")).first()
    except SQLAlchemyError:
        return False
    return True


def _page(rows: list[dict[str, object]], total: int, page: int, page_size: int) -> dict[str, object]:
    return {
        "rows": rows,
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": ceil(total / page_size) if total else 0,
    }


def _empty_page(filters) -> dict[str, object]:
    return _page([], 0, filters.page, filters.page_size)


def _normalize_page_size(value: int) -> int:
    return value if value in PRODUCT_OPS_PAGE_SIZES else 50


def _clean(value: str | None) -> str | None:
    if value is None:
        return None
    value = value.strip()
    return value or None
