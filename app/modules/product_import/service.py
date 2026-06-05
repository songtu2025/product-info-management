from io import BytesIO
from pathlib import Path
from secrets import token_urlsafe
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import bindparam, text

from app.core.db import get_engine
from app.modules.product_info.service import (
    LOCK_CONFLICT_MESSAGE,
    clear_filter_options_cache as clear_product_filter_options_cache,
    clear_product_list_cache,
    is_locked_msku,
)
from app.shared.audit import build_change_set, record_operation_log


HEADER_MAP = {
    "店铺/站点": "store_site",
    "店铺站点": "store_site",
    "store_site": "store_site",
    "MSKU": "msku",
    "msku": "msku",
    "ASIN": "asin",
    "Asin": "asin",
    "asin": "asin",
    "父ASIN": "parent_asin",
    "父Asin": "parent_asin",
    "父 ASIN": "parent_asin",
    "parent_asin": "parent_asin",
    "产品名称": "product_name",
    "product_name": "product_name",
    "SKU": "sku",
    "sku": "sku",
    "品牌": "brand",
    "brand": "brand",
    "FNSKU": "fnsku",
    "fnsku": "fnsku",
    "销售状态": "sales_status",
    "sales_status": "sales_status",
    "仓储类型": "storage_type",
    "storage_type": "storage_type",
    "一级品类": "category_level_1",
    "category_level_1": "category_level_1",
    "品类A": "category_a",
    "category_a": "category_a",
    "品类B": "category_b",
    "category_b": "category_b",
    "Listing": "listing",
    "listing": "listing",
    "标签名": "label_name",
    "标签名称": "label_name",
    "label_name": "label_name",
    "MSKU发货备注": "msku_shipping_remark",
    "MSKU 发货备注": "msku_shipping_remark",
    "发货备注": "msku_shipping_remark",
    "msku_shipping_remark": "msku_shipping_remark",
    "借调备注": "transfer_remark",
    "transfer_remark": "transfer_remark",
    "锁仓MSKU": "msku_lock_status",
    "锁仓MKSU": "msku_lock_status",
    "锁仓状态": "msku_lock_status",
    "msku_lock_status": "msku_lock_status",
}

KEY_FIELDS = {"store_site", "msku"}
IMPORT_UPDATE_FIELDS = {
    "asin",
    "parent_asin",
    "product_name",
    "sku",
    "brand",
    "fnsku",
    "sales_status",
    "storage_type",
    "category_level_1",
    "category_a",
    "category_b",
    "listing",
    "label_name",
    "msku_shipping_remark",
    "transfer_remark",
    "msku_lock_status",
}
IMPORT_TEMPLATE_HEADERS = (
    "店铺/站点",
    "MSKU",
    "ASIN",
    "父ASIN",
    "产品名称",
    "SKU",
    "品牌",
    "FNSKU",
    "销售状态",
    "仓储类型",
    "一级品类",
    "品类A",
    "品类B",
    "Listing",
    "标签名",
    "MSKU发货备注",
    "借调备注",
    "锁仓MSKU",
)
IMPORT_UPLOAD_DIR = Path("data/imports")


def preview_product_import(content: bytes) -> dict[str, Any]:
    rows, blocked_fields = _read_rows(content)
    existing_keys = _load_existing_keys(rows)
    existing_products = _load_existing_products(rows)

    valid_rows = []
    missing_product_rows = []
    error_rows = []

    for row in rows:
        store_site = row["data"].get("store_site")
        msku = row["data"].get("msku")
        if not store_site or not msku:
            error_rows.append(
                {
                    "row_number": row["row_number"],
                    "message": "缺少店铺站点或 MSKU",
                }
            )
            continue

        if (store_site, msku) not in existing_keys:
            missing_product_rows.append(
                {
                    "row_number": row["row_number"],
                    "store_site": store_site,
                    "msku": msku,
                }
            )
            continue

        changes = {
            field: value
            for field, value in row["data"].items()
            if field in IMPORT_UPDATE_FIELDS and value is not None
        }
        before = existing_products.get((store_site, msku), {})
        change_set = build_change_set(before, changes)
        valid_rows.append(
            {
                "row_number": row["row_number"],
                "store_site": store_site,
                "msku": msku,
                "changes": changes,
                "change_items": [
                    {"field": field, "old": values["old"], "new": values["new"]}
                    for field, values in change_set.items()
                ],
            }
        )

    conflict_row_numbers = _find_lock_conflict_row_numbers(valid_rows)
    if conflict_row_numbers:
        valid_rows = [
            row
            for row in valid_rows
            if row["row_number"] not in conflict_row_numbers
        ]
        error_rows.extend(
            {
                "row_number": row_number,
                "message": LOCK_CONFLICT_MESSAGE,
            }
            for row_number in sorted(conflict_row_numbers)
        )

    return {
        "total_rows": len(rows),
        "valid_count": len(valid_rows),
        "missing_product_count": len(missing_product_rows),
        "error_count": len(error_rows),
        "blocked_fields": blocked_fields,
        "valid_rows": valid_rows,
        "missing_product_rows": missing_product_rows,
        "error_rows": error_rows,
    }


def build_product_import_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "产品信息导入模板"
    sheet.append(list(IMPORT_TEMPLATE_HEADERS))
    sheet.append(["SAYOLA:US", "MSKU-001"] + [""] * (len(IMPORT_TEMPLATE_HEADERS) - 2))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_product_import_issue_workbook(preview: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入问题"
    sheet.append(["类型", "行号", "店铺站点", "MSKU", "错误说明"])

    for row in preview.get("missing_product_rows", []):
        sheet.append(
            [
                "未匹配产品",
                row.get("row_number"),
                row.get("store_site"),
                row.get("msku"),
                "产品不存在",
            ]
        )
    for row in preview.get("error_rows", []):
        sheet.append(
            [
                "错误行",
                row.get("row_number"),
                row.get("store_site"),
                row.get("msku"),
                row.get("message"),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def save_import_upload(content: bytes) -> str:
    IMPORT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    token = token_urlsafe(24)
    _upload_path(token).write_bytes(content)
    return token


def load_import_upload(token: str) -> bytes | None:
    if not token or any(char not in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_" for char in token):
        return None

    path = _upload_path(token)
    if not path.exists():
        return None
    return path.read_bytes()


def commit_product_import(content: bytes, changed_by: str = "system") -> dict[str, Any]:
    preview = preview_product_import(content)
    if preview["blocked_fields"] or preview["missing_product_count"] or preview["error_count"]:
        return {
            "success": False,
            "updated_count": 0,
            "skipped_count": 0,
            "message": "校验未通过，未写入数据库。",
            "preview": preview,
        }

    engine = get_engine()
    if engine is None:
        return {
            "success": False,
            "updated_count": 0,
            "skipped_count": 0,
            "message": "数据库连接不可用，未写入数据库。",
            "preview": preview,
        }

    updated_count = 0
    skipped_count = 0
    with engine.begin() as conn:
        for row in preview["valid_rows"]:
            desired = row["changes"]
            if not desired:
                skipped_count += 1
                continue

            select_fields = ["id", *desired]
            select_sql = text(
                f"""
                SELECT {", ".join(select_fields)}
                FROM amazon_product_info
                WHERE store_site = :store_site AND msku = :msku
                """
            )
            before = conn.execute(
                select_sql,
                {"store_site": row["store_site"], "msku": row["msku"]},
            ).mappings().first()
            if before is None:
                skipped_count += 1
                continue

            changes = build_change_set(dict(before), desired)
            if not changes:
                skipped_count += 1
                continue

            update_fields = list(changes)
            update_sql = text(
                f"""
                UPDATE amazon_product_info
                SET {", ".join(f"{field} = :{field}" for field in update_fields)}
                WHERE id = :product_id
                """
            )
            params = {
                "product_id": before["id"],
                **{field: changes[field]["new"] for field in update_fields},
            }
            result = conn.execute(update_sql, params)
            if result.rowcount > 0:
                updated_count += 1
                record_operation_log(
                    conn,
                    table_name="amazon_product_info",
                    record_id=before["id"],
                    operation_type="IMPORT_UPDATE",
                    change_data=changes,
                    changed_by=changed_by,
                )
            else:
                skipped_count += 1
    if updated_count > 0:
        clear_product_list_cache()
        clear_product_filter_options_cache()

    return {
        "success": True,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "message": "写入完成",
        "preview": preview,
    }


def _upload_path(token: str) -> Path:
    return IMPORT_UPLOAD_DIR / f"{token}.xlsx"


def _read_rows(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    raw_headers = [_clean_cell(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    headers = [HEADER_MAP.get(header) for header in raw_headers]
    blocked_fields = [
        header
        for header, mapped in zip(raw_headers, headers)
        if header and mapped is None
    ]

    rows = []
    for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None and str(value).strip() for value in values):
            continue

        data = {}
        for mapped, value in zip(headers, values):
            if not mapped:
                continue
            data[mapped] = _clean_cell(value)
        rows.append({"row_number": index, "data": data})

    return rows, blocked_fields


def _load_existing_keys(rows: list[dict[str, Any]]) -> set[tuple[str, str]]:
    keys = {
        (row["data"].get("store_site"), row["data"].get("msku"))
        for row in rows
        if row["data"].get("store_site") and row["data"].get("msku")
    }
    if not keys:
        return set()

    engine = get_engine()
    if engine is None:
        return set()

    store_sites = sorted({store_site for store_site, _ in keys})
    mskus = sorted({msku for _, msku in keys})
    query = text(
        """
        SELECT store_site, msku
        FROM amazon_product_info
        WHERE store_site IN :store_sites AND msku IN :mskus
        """
    ).bindparams(
        bindparam("store_sites", expanding=True),
        bindparam("mskus", expanding=True),
    )

    with engine.connect() as conn:
        matched = {
            (row["store_site"], row["msku"])
            for row in conn.execute(
                query,
                {"store_sites": store_sites, "mskus": mskus},
            ).mappings()
        }

    return matched & keys


def _load_existing_products(rows: list[dict[str, Any]]) -> dict[tuple[str, str], dict[str, Any]]:
    keys = {
        (row["data"].get("store_site"), row["data"].get("msku"))
        for row in rows
        if row["data"].get("store_site") and row["data"].get("msku")
    }
    fields = sorted(
        {
            field
            for row in rows
            for field in row["data"]
            if field in IMPORT_UPDATE_FIELDS
        }
    )
    if not keys or not fields:
        return {}

    engine = get_engine()
    if engine is None:
        return {}

    store_sites = sorted({store_site for store_site, _ in keys})
    mskus = sorted({msku for _, msku in keys})
    select_fields = ["store_site", "msku", *fields]
    query = text(
        f"""
        SELECT {", ".join(select_fields)}
        FROM amazon_product_info
        WHERE store_site IN :store_sites AND msku IN :mskus
        """
    ).bindparams(
        bindparam("store_sites", expanding=True),
        bindparam("mskus", expanding=True),
    )

    with engine.connect() as conn:
        return {
            (row["store_site"], row["msku"]): dict(row)
            for row in conn.execute(
                query,
                {"store_sites": store_sites, "mskus": mskus},
            ).mappings()
        }


def _find_lock_conflict_row_numbers(rows: list[dict[str, Any]]) -> set[int]:
    if not rows:
        return set()

    engine = get_engine()
    if engine is None:
        return set()

    store_sites = sorted({row["store_site"] for row in rows})
    query = text(
        """
        SELECT id, store_site, msku, sku, msku_lock_status
        FROM amazon_product_info
        WHERE store_site IN :store_sites
        """
    ).bindparams(bindparam("store_sites", expanding=True))

    with engine.connect() as conn:
        products = [
            dict(row)
            for row in conn.execute(query, {"store_sites": store_sites}).mappings()
        ]

    by_key = {
        (product["store_site"], product["msku"]): product
        for product in products
    }
    imported_ids_by_row = {}
    for row in rows:
        product = by_key.get((row["store_site"], row["msku"]))
        if not product:
            continue
        imported_ids_by_row[row["row_number"]] = product["id"]
        if "sku" in row["changes"]:
            product["sku"] = row["changes"]["sku"]
        if "msku_lock_status" in row["changes"]:
            product["msku_lock_status"] = row["changes"]["msku_lock_status"]

    locked_ids_by_group: dict[tuple[str, str], list[int]] = {}
    for product in products:
        if not is_locked_msku(product.get("msku_lock_status")):
            continue
        store_site = product.get("store_site")
        sku = product.get("sku")
        if not store_site or not sku:
            continue
        locked_ids_by_group.setdefault((store_site, sku), []).append(product["id"])

    conflict_ids = {
        product_id
        for product_ids in locked_ids_by_group.values()
        if len(product_ids) > 1
        for product_id in product_ids
    }
    return {
        row_number
        for row_number, product_id in imported_ids_by_row.items()
        if product_id in conflict_ids
    }


def _clean_cell(value: object) -> str | None:
    if value is None:
        return None
    value = str(value).strip()
    return value or None
